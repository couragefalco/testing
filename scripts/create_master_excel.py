#!/usr/bin/env python3
"""
Create master Stanford Professors Outreach Excel - all 74+ professors.
Professional formatting with Stanford branding.
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Colors ────────────────────────────────────────────────────────────────────
DARK_NAVY    = "1B2A4A"
CARDINAL_RED = "8C1515"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F5F5F5"
BORDER_GRAY  = "BFBFBF"
DARK_GRAY    = "4A4A4A"
AREA_BLUE    = "1F4E79"
AREA_PURPLE  = "6B3FA0"
DEEP_PURPLE  = "4A2882"
SUBTITLE_BG  = "2C3E6B"
SALMON       = "FDDCCA"
GREEN_BG     = "C6EFCE"
GREEN_FT     = "006100"
YELLOW_BG    = "FFEB9C"
YELLOW_FT    = "9C6500"
RED_BG       = "FFC7CE"
RED_FT       = "9C0006"
BLUE_BG      = "BDD7EE"
BLUE_FT      = "1F4E79"
PRI_HIGH     = "E2EFDA"
PRI_MED      = "FCE4D6"
PRI_STD      = "D6DCE4"

thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)
header_border = Border(
    left=Side(style="thin", color=WHITE),
    right=Side(style="thin", color=WHITE),
    top=Side(style="medium", color=WHITE),
    bottom=Side(style="medium", color=WHITE),
)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Load original 33 professors ──────────────────────────────────────────────
original_professors = []
research_dossier = {}

# Parse the research dossier for key recent work
dossier_path = os.path.join(BASE_DIR, "stanford_professors_research_dossier.md")
if os.path.exists(dossier_path):
    with open(dossier_path, "r") as f:
        content = f.read()
    # Simple extraction of recent work per professor
    sections = content.split("\n## ")
    for section in sections[1:]:
        lines = section.strip().split("\n")
        name_line = lines[0]
        # Extract name (before the parenthesis)
        if "(" in name_line:
            # Format: "1. Kathleen Eisenhardt (MS&E) - Strategy..."
            parts = name_line.split("(")
            name = parts[0].strip()
            # Remove leading number and dot
            for i, c in enumerate(name):
                if c.isalpha():
                    name = name[i:]
                    break
            name = name.strip()
        else:
            name = name_line.strip()

        # Get outreach angle
        outreach = ""
        for i, line in enumerate(lines):
            if "**Outreach Angle:**" in line:
                outreach = line.replace("**Outreach Angle:**", "").strip()
                break

        # Get recent papers
        recent_work = ""
        in_papers = False
        paper_lines = []
        for line in lines:
            if "**Recent/Notable Papers:**" in line or "**Recent/Notable Papers/Projects" in line:
                in_papers = True
                continue
            if in_papers:
                if line.startswith("**") or line.startswith("---"):
                    in_papers = False
                    continue
                if line.strip().startswith("- "):
                    paper_lines.append(line.strip()[2:])

        if paper_lines:
            recent_work = "; ".join(paper_lines[:2])
        if outreach:
            if recent_work:
                recent_work = recent_work + " | ANGLE: " + outreach
            else:
                recent_work = outreach

        research_dossier[name] = recent_work

# Original CSV
orig_csv = os.path.join(BASE_DIR, "stanford_professors.csv")
with open(orig_csv, "r") as f:
    reader = csv.DictReader(f)
    for row in reader:
        name = row["Name"].strip()
        # Google Scholar URLs from verified data
        scholar_urls = {
            "Kathleen Eisenhardt": "https://scholar.google.com/citations?user=iTQHTwsAAAAJ",
            "Chuck Eesley": "https://scholar.google.com/citations?user=ENeJ_gkAAAAJ",
            "Ramesh Johari": "https://scholar.google.com/citations?user=fhLHgd8AAAAJ",
            "Pamela Hinds": "https://scholar.google.com/citations?user=m-aU0aIAAAAJ",
            "Melissa Valentine": "https://scholar.google.com/citations?user=yulZTtMAAAAJ",
            "Ashish Goel": "https://scholar.google.com/citations?user=B_rKfusAAAAJ",
            "Itai Ashlagi": "https://scholar.google.com/citations?user=t9iq5TwAAAAJ",
            "Amin Saberi": "https://scholar.google.com/citations?user=_PZKLYUAAAAJ",
            "Fei-Fei Li": "https://scholar.google.com/citations?user=rDfyQnIAAAAJ",
            "Percy Liang": "https://scholar.google.com/citations?user=pouyVyUAAAAJ",
            "Michael Bernstein": "https://scholar.google.com/citations?user=zkhHirIAAAAJ",
            "Sachin Katti": "https://scholar.google.com/citations?user=cc4Qi_IAAAAJ",
            "Larry Leifer": "https://scholar.google.com/citations?user=7qpQMZkAAAAJ",
            "Mark Cutkosky": "https://scholar.google.com/citations?user=qIg8KFYAAAAJ",
            "Allison Okamura": "https://scholar.google.com/citations?user=lD4Yjn4AAAAJ",
            "Erik Brynjolfsson": "https://scholar.google.com/citations?user=lqyGZpQAAAAJ",
            "Susan Athey": "https://scholar.google.com/citations?user=UdaJi94AAAAJ",
            "Tom Byers": "https://scholar.google.com/citations?user=8VSiPTkAAAAJ",
            "James Zou": "https://scholar.google.com/citations?user=23ZXZvEAAAAJ",
            "Nicholas Bloom": "https://scholar.google.com/citations?user=fJy1tloAAAAJ",
            "Robert Burgelman": "https://scholar.google.com/citations?user=xP8e32EAAAAJ",
            "Jesper Sorensen": "https://scholar.google.com/citations?user=aCkLjUoAAAAJ",
            "William Barnett": "https://scholar.google.com/citations?user=LttXXDQAAAAJ",
            "Ilya Strebulaev": "https://scholar.google.com/citations?user=j5HgL4MAAAAJ",
            "Peter DeMarzo": "https://scholar.google.com/citations?user=nJctXjIAAAAJ",
            "Jeffrey Pfeffer": "https://scholar.google.com/citations?user=vuPQD7sAAAAJ",
            "Hayagreeva Rao": "https://scholar.google.com/citations?user=jELAgFcAAAAJ",
            "Lindred Greer": "https://scholar.google.com/citations?user=OzDMHuIAAAAJ",
            "Robert Sutton": "https://scholar.google.com/citations?user=GKu6sNQAAAAJ",
            "Haim Mendelson": "https://scholar.google.com/citations?user=rtjF3_gAAAAJ",
            "Hau Lee": "https://scholar.google.com/citations?user=y1efcysAAAAJ",
            "Stefanos Zenios": "https://scholar.google.com/citations?user=_ZH86NYAAAAJ",
            "Sarah Soule": "https://scholar.google.com/citations?user=u_jPc_wAAAAJ",
        }
        personal_sites = {
            "Kathleen Eisenhardt": "https://stvp.stanford.edu",
            "Ramesh Johari": "http://web.stanford.edu/~rjohari/",
            "Melissa Valentine": "https://mvalentine.github.io/",
            "Ashish Goel": "https://web.stanford.edu/~ashishg/",
            "Itai Ashlagi": "http://web.stanford.edu/~iashlagi/",
            "Amin Saberi": "http://web.stanford.edu/~saberi",
            "Fei-Fei Li": "http://vision.stanford.edu/",
            "Percy Liang": "https://cs.stanford.edu/~pliang/",
            "Michael Bernstein": "http://hci.stanford.edu/msb/",
            "Larry Leifer": "https://centerfordesignresearch.stanford.edu/",
            "Mark Cutkosky": "http://bdml.stanford.edu",
            "Allison Okamura": "http://charm.stanford.edu/",
            "Erik Brynjolfsson": "http://brynjolfsson.com",
            "Susan Athey": "http://athey.people.stanford.edu/",
            "Tom Byers": "http://web.stanford.edu/~tbyers",
            "James Zou": "https://www.james-zou.com/",
            "Nicholas Bloom": "https://nbloom.people.stanford.edu/",
            "Jesper Sorensen": "https://web.stanford.edu/~sorensen/",
            "Ilya Strebulaev": "http://faculty-gsb.stanford.edu/strebulaev/",
            "Jeffrey Pfeffer": "https://jeffreypfeffer.com/",
            "Robert Sutton": "https://bobsutton.net/",
            "Hau Lee": "https://web.stanford.edu/~haulee/",
        }
        # Manual recent work fixes for dossier name mismatches
        manual_recent_work = {
            "Hayagreeva Rao": "The Friction Project (2024 bestseller with Sutton) - friction fixing framework for scaling organizations",
        }
        if name in manual_recent_work:
            research_dossier[name] = manual_recent_work[name]

        # Corrected emails
        email_corrections = {
            "Ashish Goel": "ashish.goel@stanford.edu",
            "Fei-Fei Li": "feifeili@stanford.edu",
        }
        email = email_corrections.get(name, row["Email"].strip())

        recent = research_dossier.get(name, "")
        # Truncate if too long
        if len(recent) > 300:
            recent = recent[:297] + "..."

        original_professors.append({
            "name": name,
            "department": row["Department"].strip(),
            "area": row["Area"].strip(),
            "research": row["Research Focus"].strip(),
            "recent_work": recent,
            "profile_url": row["Profile URL"].strip(),
            "email": email,
            "scholar_url": scholar_urls.get(name, ""),
            "personal_site": personal_sites.get(name, ""),
            "flag": "",
        })

# Flag unavailable professors
flags = {
    "Fei-Fei Li": "On partial leave (World Labs)",
    "Michael Bernstein": "On leave until Fall 2026 (Simile)",
    "Sachin Katti": "LEFT STANFORD - Intel CTO (Apr 2025)",
    "Lindred Greer": "Moved to Michigan Ross",
    "Robert Sutton": "Professor Emeritus (retired 2023)",
    "Sarah Soule": "Now GSB Dean (June 2025)",
}
for p in original_professors:
    if p["name"] in flags:
        p["flag"] = flags[p["name"]]

# ── Load additional professors ────────────────────────────────────────────────
additional_professors = []
add_csv = os.path.join(BASE_DIR, "stanford_professors_additional.csv")
if os.path.exists(add_csv):
    with open(add_csv, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("Name", "").strip()
            if not name:
                continue
            additional_professors.append({
                "name": name,
                "department": row.get("Department", "").strip(),
                "area": row.get("Area", "").strip(),
                "research": row.get("Research Focus", "").strip(),
                "recent_work": row.get("Recent Work", "").strip(),
                "profile_url": row.get("Profile URL", "").strip(),
                "email": row.get("Email", "").strip(),
                "scholar_url": row.get("Google Scholar URL", "").strip(),
                "personal_site": row.get("Personal/Lab Website", "").strip(),
                "flag": "",
            })

# Engineering additional professors (from first agent - hardcoded since not in CSV)
eng_additional = [
    {"name": "Yinyu Ye", "department": "MS&E (Emeritus)", "area": "Engineering", "research": "Optimization, data science, algorithmic game theory", "email": "yinyu-ye@stanford.edu", "profile_url": "https://profiles.stanford.edu/yinyu-ye", "scholar_url": "https://scholar.google.com/citations?user=BgOXDogAAAAJ", "personal_site": "https://web.stanford.edu/~yyye/", "recent_work": "Linear programming in the shadow of Simplex (Math Programming, 2024)"},
    {"name": "Peter Glynn", "department": "MS&E", "area": "Engineering", "research": "Simulation, computational probability, queueing theory, financial risk", "email": "glynn@stanford.edu", "profile_url": "https://profiles.stanford.edu/peter-glynn", "scholar_url": "", "personal_site": "https://web.stanford.edu/~glynn/", "recent_work": "Stochastic simulation and Monte Carlo methods (Operations Research, 2023)"},
    {"name": "Andrea Montanari", "department": "Statistics & MS&E", "area": "Engineering", "research": "High-dimensional statistics, ML, information theory", "email": "montanar@stanford.edu", "profile_url": "https://profiles.stanford.edu/andrea-montanari", "scholar_url": "https://scholar.google.com/citations?user=r3q68rcAAAAJ", "personal_site": "https://web.stanford.edu/~montanar/", "recent_work": "Sampling, diffusions, and stochastic localization (COLT, 2024)"},
    {"name": "Markus Pelger", "department": "MS&E", "area": "Engineering", "research": "ML for finance, asset pricing, high-dimensional statistics", "email": "mpelger@stanford.edu", "profile_url": "https://profiles.stanford.edu/markus-pelger", "scholar_url": "https://scholar.google.com/citations?user=FpNrPm8AAAAJ", "personal_site": "https://mpelger.people.stanford.edu/", "recent_work": "Deep Learning in Asset Pricing (Management Science, 2024)"},
    {"name": "Irene Lo", "department": "MS&E", "area": "Engineering", "research": "Market design, matching markets, mechanism design, equitable access", "email": "irene.lo@stanford.edu", "profile_url": "https://profiles.stanford.edu/irene-lo", "scholar_url": "https://scholar.google.com/citations?user=QVd3CugAAAAJ", "personal_site": "https://sites.google.com/view/irene-lo", "recent_work": "The Cutoff Structure of School Choice (AER, 2024)"},
    {"name": "Jure Leskovec", "department": "Computer Science", "area": "Engineering", "research": "Graph neural networks, social networks, data mining, knowledge graphs", "email": "jure@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/jure-leskovec", "scholar_url": "https://scholar.google.com/citations?user=Q_kKkIUAAAAJ", "personal_site": "https://cs.stanford.edu/people/jure/", "recent_work": "Relational Transformer (2025) - zero-shot foundation models for relational data; GraphMETRO (NeurIPS 2024)"},
    {"name": "Christopher Re", "department": "Computer Science", "area": "Engineering", "research": "Foundation models, data-centric AI, ML systems", "email": "chrismre@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/christopher-re", "scholar_url": "https://scholar.google.com/citations?user=DnnCWN0AAAAJ", "personal_site": "https://cs.stanford.edu/~chrismre/", "recent_work": "Evo 2: Genome modeling across all domains of life (2025); Thunderkittens GPU kernels (ICLR 2025)"},
    {"name": "Chelsea Finn", "department": "CS & EE", "area": "Engineering", "research": "Meta-learning, robot learning, deep reinforcement learning", "email": "cbfinn@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/chelsea-finn", "scholar_url": "https://scholar.google.com/citations?user=vfPE6hgAAAAJ", "personal_site": "https://ai.stanford.edu/~cbfinn/", "recent_work": "Mobile ALOHA (CoRL 2024) - bimanual mobile manipulation robot; OpenVLA open-source vision-language-action model"},
    {"name": "Emma Brunskill", "department": "Computer Science", "area": "Engineering", "research": "RL for education & healthcare, decision making under uncertainty", "email": "ebrun@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/emma-brunskill", "scholar_url": "https://scholar.google.com/citations?user=HaN8b2YAAAAJ", "personal_site": "https://cs.stanford.edu/people/ebrun/", "recent_work": "Offline RL for education policy optimization (NeurIPS, 2023)"},
    {"name": "Christopher Manning", "department": "CS & Linguistics", "area": "Engineering", "research": "NLP, deep learning for language, Associate Director of HAI", "email": "manning@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/chris-manning", "scholar_url": "", "personal_site": "https://nlp.stanford.edu/~manning/", "recent_work": "Elected to National Academy of Engineering (2025); IEEE von Neumann Medal (2024)"},
    {"name": "Dan Jurafsky", "department": "Linguistics & CS", "area": "Engineering", "research": "NLP, speech recognition, computational social science", "email": "jurafsky@stanford.edu", "profile_url": "https://profiles.stanford.edu/dan-jurafsky", "scholar_url": "https://scholar.google.com/citations?user=uZg9l58AAAAJ", "personal_site": "https://web.stanford.edu/~jurafsky/", "recent_work": "Speech & Language Processing, 3rd Ed (textbook, 2024 updates)"},
    {"name": "James Landay", "department": "Computer Science", "area": "Engineering", "research": "HCI, ubiquitous computing, Co-Director of Stanford HAI", "email": "landay@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/james-landay", "scholar_url": "https://scholar.google.com/citations?user=oQsObk0AAAAJ", "personal_site": "https://www.landay.org/", "recent_work": "AI-mediated communication & interaction design (CHI, 2024)"},
    {"name": "Stefano Ermon", "department": "Computer Science", "area": "Engineering", "research": "Generative AI, probabilistic reasoning, ML for social good", "email": "ermon@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/stefano-ermon", "scholar_url": "https://scholar.google.com/citations?user=ogXTOZ4AAAAJ", "personal_site": "https://cs.stanford.edu/~ermon/", "recent_work": "Score-based diffusion models via SDEs (ICLR, 2023)"},
    {"name": "Tatsunori Hashimoto", "department": "Computer Science", "area": "Engineering", "research": "LLM alignment & safety, NLP, distributional robustness", "email": "thashim@stanford.edu", "profile_url": "https://profiles.stanford.edu/tatsunori-hashimoto", "scholar_url": "https://scholar.google.com/citations?user=5ygiTwsAAAAJ", "personal_site": "https://thashim.github.io/", "recent_work": "LLM eval & alignment under distribution shift (NeurIPS, 2023)"},
    {"name": "Diyi Yang", "department": "Computer Science", "area": "Engineering", "research": "Socially aware NLP, LLMs, human-AI interaction", "email": "diyiy@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/diyi-yang", "scholar_url": "https://scholar.google.com/citations?user=j9jhYqQAAAAJ", "personal_site": "https://cs.stanford.edu/~diyiy/", "recent_work": "Social-aware NLP and LLM safety for online communities (ACL, 2024)"},
    {"name": "Carlos Guestrin", "department": "Computer Science", "area": "Engineering", "research": "ML explainability (LIME, XGBoost), Director of SAIL", "email": "guestrin@stanford.edu", "profile_url": "https://profiles.stanford.edu/carlos-guestrin", "scholar_url": "https://scholar.google.com/citations?user=DpLFv4gAAAAJ", "personal_site": "", "recent_work": "SHAP & interpretable ML at scale (SAIL, 2023-24)"},
    {"name": "Dorsa Sadigh", "department": "CS & EE", "area": "Engineering", "research": "Human-robot interaction, robot learning, autonomous systems", "email": "dorsa@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/dorsa-sadigh", "scholar_url": "https://scholar.google.com/citations?user=ZaJEZpYAAAAJ", "personal_site": "https://dorsa.fyi/", "recent_work": "Language-guided robot learning from human feedback (CoRL, 2023)"},
    {"name": "Noah Goodman", "department": "Psychology & CS", "area": "Interdisciplinary", "research": "Probabilistic programming, cognitive science, computational language", "email": "ngoodman@stanford.edu", "profile_url": "https://profiles.stanford.edu/noah-goodman", "scholar_url": "https://scholar.google.com/citations?user=OUpIbcQAAAAJ", "personal_site": "https://cocolab.stanford.edu/ndg", "recent_work": "Language agents as optimizers for probabilistic reasoning (NeurIPS, 2024)"},
    {"name": "Tengyu Ma", "department": "CS & Statistics", "area": "Engineering", "research": "Deep learning theory, RL, foundation models", "email": "tengyuma@stanford.edu", "profile_url": "https://profiles.stanford.edu/tengyu-ma", "scholar_url": "https://scholar.google.com/citations?user=i38QlUwAAAAJ", "personal_site": "https://ai.stanford.edu/~tengyuma/", "recent_work": "Theoretical foundations of in-context learning (ICML, 2024)"},
    {"name": "Sean Follmer", "department": "Mechanical Engineering", "area": "Engineering", "research": "HCI, haptics, human-robot interaction, tangible interfaces", "email": "sfollmer@stanford.edu", "profile_url": "https://profiles.stanford.edu/sean-follmer", "scholar_url": "https://scholar.google.com/citations?user=f3g5oeEAAAAJ", "personal_site": "https://shape.stanford.edu/", "recent_work": "Shape-changing haptic interfaces for accessibility (CHI, 2024)"},
    {"name": "Monroe Kennedy III", "department": "Mechanical Engineering", "area": "Engineering", "research": "Collaborative robotics, human-robot collaboration, assistive robotics", "email": "monroek@stanford.edu", "profile_url": "https://profiles.stanford.edu/monroe-kennedy", "scholar_url": "https://scholar.google.com/citations?user=x2ZPRfoAAAAJ", "personal_site": "https://monroekennedy3.com/", "recent_work": "Multi-agent collaborative manipulation planning (ICRA, 2024)"},
    {"name": "Steve Collins", "department": "Mechanical Engineering", "area": "Engineering", "research": "Wearable robots, exoskeletons, prosthetics, biomechatronics", "email": "stevecollins@stanford.edu", "profile_url": "https://profiles.stanford.edu/steven-collins", "scholar_url": "https://scholar.google.com/citations?user=eF5vfBAAAAAJ", "personal_site": "https://biomechatronics.stanford.edu/", "recent_work": "Personalized ankle exoskeleton via human-in-the-loop optimization (Nature, 2024)"},
    {"name": "Bernard Roth", "department": "ME & d.school", "area": "Engineering", "research": "Design thinking, kinematics, robotics, d.school co-founder", "email": "roth@stanford.edu", "profile_url": "https://profiles.stanford.edu/bernard-roth", "scholar_url": "", "personal_site": "https://dschool.stanford.edu/directory/bernie-roth", "recent_work": "The Achievement Habit & design thinking kinematics (ongoing)"},
    {"name": "Stephen Boyd", "department": "Electrical Engineering", "area": "Engineering", "research": "Convex optimization, control systems, AI/ML optimization", "email": "boyd@stanford.edu", "profile_url": "https://profiles.stanford.edu/stephen-boyd", "scholar_url": "https://scholar.google.com/citations?user=GExyiRkAAAAJ", "personal_site": "https://stanford.edu/~boyd/", "recent_work": "Fitting Multilevel Factor Models (SIAM J. Matrix Analysis, 2025)"},
    {"name": "Subhasish Mitra", "department": "EE & CS", "area": "Engineering", "research": "AI hardware, robust computing, nanosystems", "email": "subh@stanford.edu", "profile_url": "https://profiles.stanford.edu/subhasish-mitra", "scholar_url": "https://scholar.google.com/citations?user=SCdhzWoAAAAJ", "personal_site": "https://rsg.stanford.edu/", "recent_work": "MINOTAUR: Posit-Based Edge Transformer Accelerator (IEEE JSSC, 2025)"},
    {"name": "Gordon Wetzstein", "department": "EE", "area": "Engineering", "research": "Computational imaging, neural rendering, AR/VR", "email": "gordon.wetzstein@stanford.edu", "profile_url": "https://profiles.stanford.edu/gordon-wetzstein", "scholar_url": "https://scholar.google.com/citations?user=VOf45S0AAAAJ", "personal_site": "https://stanford.edu/~gordonwz/", "recent_work": "Full-colour 3D holographic AR displays with metasurface waveguides (Nature, 2024)"},
    {"name": "Russ Altman", "department": "Bioengineering & CS", "area": "Engineering", "research": "AI for medicine, pharmacogenomics, Associate Director of HAI", "email": "russ.altman@stanford.edu", "profile_url": "https://profiles.stanford.edu/russ-altman", "scholar_url": "https://scholar.google.com/citations?user=s6XjtCMAAAAJ", "personal_site": "https://rbaltman.people.stanford.edu/", "recent_work": "HOTPocket (2025) - 2.4M drug target pockets; AI misuse mitigation (Nature Machine Intelligence, 2024)"},
    {"name": "Manu Prakash", "department": "Bioengineering", "area": "Engineering", "research": "Frugal science, global health innovation, Foldscope", "email": "manup@stanford.edu", "profile_url": "https://profiles.stanford.edu/manu-prakash", "scholar_url": "https://scholar.google.com/citations?user=ROaO6BkAAAAJ", "personal_site": "https://web.stanford.edu/group/prakash-lab/", "recent_work": "Hidden comet tails of marine snow impede ocean carbon sequestration (Science, 2024)"},
    {"name": "Drew Endy", "department": "Bioengineering", "area": "Engineering", "research": "Synthetic biology, biotechnology policy", "email": "endy@stanford.edu", "profile_url": "https://profiles.stanford.edu/drew-endy", "scholar_url": "https://scholar.google.com/citations?user=6I2W3iwAAAAJ", "personal_site": "", "recent_work": "Risk mitigation of potential misuse of AI in biomedical research (Nat Mach Intell, 2024)"},
    {"name": "Kwabena Boahen", "department": "Bioengineering & EE", "area": "Engineering", "research": "Neuromorphic computing, brain-inspired chips", "email": "boahen@stanford.edu", "profile_url": "https://profiles.stanford.edu/kwabena-boahen", "scholar_url": "https://scholar.google.com/citations?user=WuOsGfsAAAAJ", "personal_site": "", "recent_work": "Hierarchical Event Readout with Async Pipelined Opportunistic Merges (ASYNC, 2025)"},
    {"name": "Mykel Kochenderfer", "department": "Aero & Astro", "area": "Engineering", "research": "Decision making under uncertainty, autonomous systems", "email": "mykel@stanford.edu", "profile_url": "https://profiles.stanford.edu/mykel-kochenderfer", "scholar_url": "https://scholar.google.com/citations?user=cAy9G6oAAAAJ", "personal_site": "https://mykel.kochenderfer.com/", "recent_work": "BetaZero: Belief-state planning for long-horizon POMDPs (RL Conf, 2024)"},
    {"name": "Marco Pavone", "department": "Aero & Astro, EE, CS", "area": "Engineering", "research": "Autonomous vehicles, robotics, self-driving, future mobility", "email": "pavone@stanford.edu", "profile_url": "https://profiles.stanford.edu/marco-pavone", "scholar_url": "", "personal_site": "https://web.stanford.edu/~pavone/", "recent_work": "Pseudo-Simulation for Autonomous Driving (JMLR, 2025)"},
    {"name": "Martin Fischer", "department": "Civil & Env Engineering", "area": "Engineering", "research": "Virtual design & construction, construction tech, CIFE director", "email": "fischer@stanford.edu", "profile_url": "https://profiles.stanford.edu/martin-fischer", "scholar_url": "https://scholar.google.com/citations?user=ACZeiEEAAAAJ", "personal_site": "https://cife.stanford.edu/people/martin-fischer", "recent_work": "Vitruvio: Conditional VAE to generate building meshes via perspective sketches (Autom Constr, 2024)"},
    {"name": "Tina Seelig", "department": "MS&E (Emerita)", "area": "Engineering", "research": "Creativity, innovation, entrepreneurship education", "email": "tseelig@stanford.edu", "profile_url": "https://profiles.stanford.edu/tina-seelig", "scholar_url": "", "personal_site": "", "recent_work": "What I Wish I Knew About Luck (Book, 2025)"},
    {"name": "Baba Shiv", "department": "GSB Marketing", "area": "GSB", "research": "Neuroeconomics, innovation leadership, entrepreneurial mindset", "email": "bshiv@stanford.edu", "profile_url": "https://profiles.stanford.edu/baba-shiv", "scholar_url": "https://scholar.google.com/citations?user=Wr1mOp0AAAAJ", "personal_site": "", "recent_work": "Emotional responses and marketing placebo effects on decisions (Emotion, 2023)"},
    {"name": "Maneesh Agrawala", "department": "Computer Science", "area": "Engineering", "research": "Visualization, HCI, computer graphics, cognitive design", "email": "maneesh@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/maneesh-agrawala", "scholar_url": "https://scholar.google.com/citations?user=YPzKczYAAAAJ", "personal_site": "http://graphics.stanford.edu/~maneesh/", "recent_work": "ScriptViz: Visualization Tool to Aid Scriptwriting from Large Movie DB (UIST, 2024)"},
    {"name": "Karen Liu", "department": "Computer Science", "area": "Engineering", "research": "Computer animation, robotics, physics simulation, RL", "email": "karenliu@cs.stanford.edu", "profile_url": "https://profiles.stanford.edu/c-karen-liu", "scholar_url": "https://scholar.google.com/citations?user=i28fU0MAAAAJ", "personal_site": "https://tml.stanford.edu/people/karen-liu", "recent_work": "PDP: Physics-Based Character Animation via Diffusion Policy (SIGGRAPH Asia, 2024)"},
    {"name": "Iro Armeni", "department": "Civil & Env Engineering", "area": "Engineering", "research": "Computer vision for built environment, data-driven design", "email": "armeni@stanford.edu", "profile_url": "https://profiles.stanford.edu/iro-armeni", "scholar_url": "", "personal_site": "https://ir0.github.io/", "recent_work": "Rectified Point Flow: Generic Point Cloud Pose Estimation (NeurIPS Spotlight, 2025)"},
]

# Add missing fields with defaults
for p in eng_additional:
    if "recent_work" not in p:
        p["recent_work"] = ""
    if "flag" not in p:
        p["flag"] = ""

# ── Deduplicate additional professors ─────────────────────────────────────────
orig_names = {p["name"].lower() for p in original_professors}
eng_names = {p["name"].lower() for p in eng_additional}

# Filter additional CSV professors to avoid duplicates
filtered_additional = []
for p in additional_professors:
    if p["name"].lower() not in orig_names and p["name"].lower() not in eng_names:
        filtered_additional.append(p)

# Also filter eng_additional against originals
filtered_eng = [p for p in eng_additional if p["name"].lower() not in orig_names]

# ── Combine all professors ────────────────────────────────────────────────────
all_professors = original_professors + filtered_eng + filtered_additional

print(f"Total professors: {len(all_professors)}")
print(f"  Original: {len(original_professors)}")
print(f"  Engineering additional: {len(filtered_eng)}")
print(f"  GSB/Social Science additional: {len(filtered_additional)}")

# ── Create Excel ──────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Outreach Tracker"

# Column definitions
columns = ["#", "Name", "Department", "Area", "Research Focus", "Key Recent Work (2024-2025)",
           "Profile URL", "Email", "Google Scholar", "Personal/Lab Website",
           "Priority", "Status", "Notes"]
col_widths = [5, 25, 22, 18, 40, 50, 35, 30, 35, 35, 14, 14, 25]

# ── Title rows ────────────────────────────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
title_cell = ws.cell(row=1, column=1, value="Stanford Professors - Visiting Research Position Outreach")
title_cell.font = Font(name="Calibri", size=20, bold=True, color=WHITE)
title_cell.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 45

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
sub_cell = ws.cell(row=2, column=1, value="Engineering | Business | Management | Entrepreneurship | AI | Innovation  -  April 2026")
sub_cell.font = Font(name="Calibri", size=11, italic=True, color=WHITE)
sub_cell.fill = PatternFill(start_color=SUBTITLE_BG, end_color=SUBTITLE_BG, fill_type="solid")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 25

# ── Header row ────────────────────────────────────────────────────────────────
header_row = 3
for col_idx, (col_name, width) in enumerate(zip(columns, col_widths), 1):
    cell = ws.cell(row=header_row, column=col_idx, value=col_name)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = header_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[header_row].height = 30

# ── Data rows ─────────────────────────────────────────────────────────────────
area_colors = {
    "Engineering": PatternFill(start_color=AREA_BLUE, end_color=AREA_BLUE, fill_type="solid"),
    "GSB": PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid"),
    "Interdisciplinary": PatternFill(start_color=AREA_PURPLE, end_color=AREA_PURPLE, fill_type="solid"),
    "Engineering + GSB": PatternFill(start_color=DEEP_PURPLE, end_color=DEEP_PURPLE, fill_type="solid"),
    "Social Sciences": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
    "Social Sciences + Law": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
    "Law": PatternFill(start_color="744B2F", end_color="744B2F", fill_type="solid"),
    "Education": PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
    "Sustainability": PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid"),
}

for idx, prof in enumerate(all_professors):
    row_num = header_row + 1 + idx
    is_flagged = bool(prof.get("flag"))
    is_alt = idx % 2 == 1

    # Determine priority
    if is_flagged:
        priority = "Unavailable"
    elif prof["area"] in ("Engineering + GSB", "Interdisciplinary"):
        priority = "High"
    elif prof["area"] == "Engineering" and "MS&E" in prof["department"]:
        priority = "High"
    elif prof["area"] == "GSB":
        priority = "Medium"
    else:
        priority = "Standard"

    row_data = [
        idx + 1,
        prof["name"],
        prof["department"],
        prof["area"],
        prof["research"],
        prof.get("recent_work", ""),
        prof["profile_url"],
        prof["email"],
        prof.get("scholar_url", ""),
        prof.get("personal_site", ""),
        priority,
        "",  # Status
        prof.get("flag", ""),  # Notes (put flag info here)
    ]

    # Background
    if is_flagged:
        row_fill = PatternFill(start_color=SALMON, end_color=SALMON, fill_type="solid")
    elif is_alt:
        row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    else:
        row_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Area column color
    area_cell = ws.cell(row=row_num, column=4)
    area_val = prof["area"]
    if area_val in area_colors:
        area_cell.fill = area_colors[area_val]
        area_cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        area_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Priority column color
    pri_cell = ws.cell(row=row_num, column=11)
    if priority == "High":
        pri_cell.fill = PatternFill(start_color=PRI_HIGH, end_color=PRI_HIGH, fill_type="solid")
        pri_cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN_FT)
    elif priority == "Medium":
        pri_cell.fill = PatternFill(start_color=PRI_MED, end_color=PRI_MED, fill_type="solid")
        pri_cell.font = Font(name="Calibri", size=10, bold=True, color=YELLOW_FT)
    elif priority == "Unavailable":
        pri_cell.fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
        pri_cell.font = Font(name="Calibri", size=10, bold=True, color=RED_FT)
    else:
        pri_cell.fill = PatternFill(start_color=PRI_STD, end_color=PRI_STD, fill_type="solid")
        pri_cell.font = Font(name="Calibri", size=10, color=AREA_BLUE)
    pri_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Number column centered
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_num, column=1).font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)

    # Name bold
    ws.cell(row=row_num, column=2).font = Font(name="Calibri", size=10, bold=True, color=DARK_NAVY)

    # Email as link font
    email_cell = ws.cell(row=row_num, column=8)
    if prof["email"]:
        email_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    ws.row_dimensions[row_num].height = 35

# ── Status data validation ────────────────────────────────────────────────────
last_data_row = header_row + len(all_professors)
dv = DataValidation(type="list", formula1='"Contacted,Pending,No Response,Scheduled,Declined"', allow_blank=True)
dv.prompt = "Select status"
ws.add_data_validation(dv)
dv.add(f"L{header_row+1}:L{last_data_row}")

# ── Conditional formatting for Status column ──────────────────────────────────
ws.conditional_formatting.add(f"L{header_row+1}:L{last_data_row}",
    CellIsRule(operator="equal", formula=['"Contacted"'],
              fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"),
              font=Font(color=GREEN_FT, bold=True)))
ws.conditional_formatting.add(f"L{header_row+1}:L{last_data_row}",
    CellIsRule(operator="equal", formula=['"Pending"'],
              fill=PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid"),
              font=Font(color=YELLOW_FT, bold=True)))
ws.conditional_formatting.add(f"L{header_row+1}:L{last_data_row}",
    CellIsRule(operator="equal", formula=['"No Response"'],
              fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"),
              font=Font(color=RED_FT, bold=True)))
ws.conditional_formatting.add(f"L{header_row+1}:L{last_data_row}",
    CellIsRule(operator="equal", formula=['"Scheduled"'],
              fill=PatternFill(start_color=BLUE_BG, end_color=BLUE_BG, fill_type="solid"),
              font=Font(color=BLUE_FT, bold=True)))

# ── Freeze panes & auto-filter ────────────────────────────────────────────────
ws.freeze_panes = f"A{header_row + 1}"
ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{last_data_row}"

# ── Print setup ───────────────────────────────────────────────────────────────
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.orientation = "landscape"
ws.print_title_rows = f"1:{header_row}"

# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY STATISTICS SHEET
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary Statistics")

# Count by area
area_counts = {}
dept_counts = {}
priority_counts = {"High": 0, "Medium": 0, "Standard": 0, "Unavailable": 0}
for p in all_professors:
    area = p["area"]
    area_counts[area] = area_counts.get(area, 0) + 1
    dept = p["department"].split("(")[0].split("/")[0].split("&")[0].strip()
    dept_counts[dept] = dept_counts.get(dept, 0) + 1
    # Recalculate priority
    if p.get("flag"):
        priority_counts["Unavailable"] += 1
    elif area in ("Engineering + GSB", "Interdisciplinary"):
        priority_counts["High"] += 1
    elif area == "Engineering" and "MS&E" in p["department"]:
        priority_counts["High"] += 1
    elif area == "GSB":
        priority_counts["Medium"] += 1
    else:
        priority_counts["Standard"] += 1

# Title
ws2.merge_cells("A1:E1")
t = ws2.cell(row=1, column=1, value="Summary Statistics")
t.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
t.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 35

# Total
ws2.cell(row=3, column=1, value="Total Professors").font = Font(bold=True, size=12)
ws2.cell(row=3, column=2, value=len(all_professors)).font = Font(bold=True, size=12, color=CARDINAL_RED)

# By Area
r = 5
ws2.cell(row=r, column=1, value="By Area").font = Font(bold=True, size=12, color=DARK_NAVY)
for area, count in sorted(area_counts.items(), key=lambda x: -x[1]):
    r += 1
    ws2.cell(row=r, column=1, value=area)
    ws2.cell(row=r, column=2, value=count)
    if area in area_colors:
        ws2.cell(row=r, column=1).fill = area_colors[area]
        ws2.cell(row=r, column=1).font = Font(color=WHITE, bold=True)

# By Priority
r += 2
ws2.cell(row=r, column=1, value="By Priority").font = Font(bold=True, size=12, color=DARK_NAVY)
for pri, count in priority_counts.items():
    r += 1
    ws2.cell(row=r, column=1, value=pri)
    ws2.cell(row=r, column=2, value=count)

# Flagged professors
r += 2
ws2.cell(row=r, column=1, value="Flagged Professors (Limited Availability)").font = Font(bold=True, size=12, color=RED_FT)
r += 1
for h, col in [("Name", 1), ("Department", 2), ("Reason", 3)]:
    c = ws2.cell(row=r, column=col, value=h)
    c.font = Font(bold=True, color=WHITE)
    c.fill = PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid")
for p in all_professors:
    if p.get("flag"):
        r += 1
        ws2.cell(row=r, column=1, value=p["name"])
        ws2.cell(row=r, column=2, value=p["department"])
        ws2.cell(row=r, column=3, value=p["flag"])
        for col in range(1, 4):
            ws2.cell(row=r, column=col).fill = PatternFill(start_color=SALMON, end_color=SALMON, fill_type="solid")

# Column widths
for col, w in [(1, 30), (2, 15), (3, 40), (4, 20), (5, 20)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# LEGEND SHEET
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Legend & Instructions")

ws3.merge_cells("A1:D1")
t = ws3.cell(row=1, column=1, value="Legend & Instructions")
t.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
t.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
t.alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 35

legend_data = [
    ("", ""),
    ("AREA COLORS", ""),
    ("Engineering", "Blue badge - School of Engineering faculty"),
    ("GSB", "Cardinal red badge - Graduate School of Business"),
    ("Interdisciplinary", "Purple badge - Cross-department appointments"),
    ("Engineering + GSB", "Deep purple badge - Joint appointments"),
    ("Social Sciences", "Medium blue - Economics, Political Science, Communication"),
    ("Law", "Brown - Law School"),
    ("Education", "Green - Graduate School of Education"),
    ("", ""),
    ("PRIORITY LEVELS", ""),
    ("High", "Interdisciplinary + MS&E faculty (best fit for visiting research)"),
    ("Medium", "GSB faculty"),
    ("Standard", "Other departments"),
    ("Unavailable", "Left Stanford, on leave, emeritus, or now in admin role"),
    ("", ""),
    ("STATUS TRACKING", ""),
    ("Contacted", "Email sent (green)"),
    ("Pending", "Drafting email or waiting for right timing (yellow)"),
    ("No Response", "No reply after follow-up (red)"),
    ("Scheduled", "Meeting or call scheduled (blue)"),
    ("Declined", "Declined or not accepting visitors"),
    ("", ""),
    ("TIPS", ""),
    ("", "Use auto-filter to sort by Area, Priority, or Status"),
    ("", "Salmon-highlighted rows = professors with limited availability"),
    ("", "Check Google Scholar profiles for their latest papers before emailing"),
    ("", "Reference specific recent work in your outreach email"),
]

for i, (label, desc) in enumerate(legend_data, 3):
    ws3.cell(row=i, column=1, value=label).font = Font(bold=True if label else False)
    ws3.cell(row=i, column=2, value=desc)
    if label in area_colors:
        ws3.cell(row=i, column=1).fill = area_colors[label]
        ws3.cell(row=i, column=1).font = Font(bold=True, color=WHITE)

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 60

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = os.path.join(BASE_DIR, "stanford_professors_outreach.xlsx")
wb.save(output_path)
print(f"\nSaved to: {output_path}")
print(f"File size: {os.path.getsize(output_path) / 1024:.1f} KB")
