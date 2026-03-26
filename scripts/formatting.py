"""
IB-standard formatting utilities for openpyxl workbooks.

Color conventions:
  - Blue font (0000FF)  = hard-coded inputs
  - Black font (000000) = formulas / calculations
  - Green font (008000) = cross-sheet references
  - Yellow fill (FFFF00) = key assumptions
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Color / Font constants
# ---------------------------------------------------------------------------
BLUE_FONT = Font(color="0000FF")
BLACK_FONT = Font(color="000000")
GREEN_FONT = Font(color="008000")
BLUE_BOLD = Font(color="0000FF", bold=True)
BLACK_BOLD = Font(color="000000", bold=True)

# ---------------------------------------------------------------------------
# Fill constants
# ---------------------------------------------------------------------------
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SECTION_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

# ---------------------------------------------------------------------------
# Border constants
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BOTTOM_DOUBLE = Border(bottom=Side(style="double"))

# ---------------------------------------------------------------------------
# Number format constants
# ---------------------------------------------------------------------------
FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_PERCENT = "0.0%"
FMT_MULTIPLE = '0.0"x"'
FMT_YEAR = "@"
FMT_RATIO = '0.00;(0.00);"-"'
FMT_INTEGER = '#,##0;(#,##0);"-"'

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def style_input_cell(cell, fmt=None):
    """Style a cell as a hard-coded input (blue font, optional number format)."""
    cell.font = BLUE_FONT
    if fmt is not None:
        cell.number_format = fmt


def style_formula_cell(cell, fmt=None):
    """Style a cell as a formula / calculation (black font)."""
    cell.font = BLACK_FONT
    if fmt is not None:
        cell.number_format = fmt


def style_crossref_cell(cell, fmt=None):
    """Style a cell as a cross-sheet reference (green font)."""
    cell.font = GREEN_FONT
    if fmt is not None:
        cell.number_format = fmt


def style_assumption_cell(cell, fmt=None):
    """Style a cell as a key assumption (blue font + yellow fill)."""
    cell.font = BLUE_FONT
    cell.fill = YELLOW_FILL
    if fmt is not None:
        cell.number_format = fmt


def style_header_row(ws, row, max_col):
    """Apply header styling (bold, header fill, thin borders) across a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BLACK_BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def style_section_header(ws, row, max_col, text):
    """Write a section header with section fill spanning the row."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = BLACK_BOLD
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = SECTION_FILL
        c.border = THIN_BORDER
        if col == 1:
            c.font = BLACK_BOLD


def style_total_row(ws, row, max_col):
    """Apply total-row styling (bold + top thin / bottom thin border)."""
    top_bottom = Border(
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BLACK_BOLD
        cell.border = top_bottom


def style_double_line_row(ws, row, max_col):
    """Apply a double-line bottom border across a row (classic IB total)."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = BOTTOM_DOUBLE


def set_column_widths(ws, widths=None, default_width=14):
    """Set column widths.

    Parameters
    ----------
    ws : Worksheet
    widths : dict | None
        Mapping of column number (1-based) to width.  If *None*, every
        column up to max_column gets *default_width*.
    default_width : float
        Fallback width applied when *widths* is None or a column is missing.
    """
    if widths is None:
        for col in range(1, (ws.max_column or 1) + 1):
            ws.column_dimensions[get_column_letter(col)].width = default_width
    else:
        for col_num, width in widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width


def write_year_headers(ws, row, start_col, years):
    """Write year labels across columns and style them as headers.

    Parameters
    ----------
    ws : Worksheet
    row : int
    start_col : int
        First column (1-based) for the first year.
    years : list[int | str]
    """
    for i, year in enumerate(years):
        cell = ws.cell(row=row, column=start_col + i, value=year)
        cell.font = BLACK_BOLD
        cell.number_format = FMT_YEAR
        cell.alignment = Alignment(horizontal="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER


def write_label(ws, row, col, text, bold=False, indent=0):
    """Write a text label into a cell with optional bold and indent.

    Parameters
    ----------
    ws : Worksheet
    row, col : int
    text : str
    bold : bool
    indent : int
        Number of leading spaces to prepend for visual indentation.
    """
    display_text = (" " * indent) + text if indent else text
    cell = ws.cell(row=row, column=col, value=display_text)
    if bold:
        cell.font = BLACK_BOLD
    else:
        cell.font = BLACK_FONT


def freeze_panes(ws, row, col):
    """Freeze panes at the given row/col intersection.

    Parameters
    ----------
    ws : Worksheet
    row : int
        First unfrozen row (e.g. 3 freezes rows 1-2).
    col : int
        First unfrozen column (e.g. 2 freezes column A).
    """
    ws.freeze_panes = ws.cell(row=row, column=col).coordinate
