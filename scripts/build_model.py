#!/usr/bin/env python3
"""Build the ASM International Valuation Model workbook."""

import sys
import os
import datetime
import copy as _copy

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from scripts.formatting import (
    BLUE_FONT,
    BLACK_FONT,
    BLACK_BOLD,
    GREEN_FONT,
    YELLOW_FILL,
    HEADER_FILL,
    THIN_BORDER,
    FMT_CURRENCY,
    FMT_PERCENT,
    FMT_YEAR,
    FMT_RATIO,
    FMT_MULTIPLE,
    style_input_cell,
    style_formula_cell,
    style_crossref_cell,
    style_assumption_cell,
    style_section_header,
    style_total_row,
    style_double_line_row,
    write_year_headers,
    write_label,
    set_column_widths,
    freeze_panes,
)
from openpyxl.utils import get_column_letter
from scripts.data_maps import SOURCE_FILE, COMPANY_INFO, YEARS

OUTPUT_FILE = "output/ASM_Valuation_Model.xlsx"

# ---------------------------------------------------------------------------
# Input tab mappings: (source sheet name, target sheet name)
# ---------------------------------------------------------------------------
INPUT_TAB_MAP = [
    ("ASM", "Input ASM"),
    ("ASM (as reported)", "Input ASM (Reported)"),
    ("AIXTRON", "Input AIXTRON"),
    ("Applied Materials", "Input AMAT"),
    ("Lam Research", "Input LRCX"),
]


# ---------------------------------------------------------------------------
# Cover tab
# ---------------------------------------------------------------------------
def build_cover(wb):
    """Build the Cover (title page) tab."""
    # Use the default sheet created with the workbook
    ws = wb.active
    ws.title = "Cover"

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 40

    # ---- Title block ----
    row = 3
    cell = ws.cell(row=row, column=2, value="ASM International NV")
    cell.font = Font(size=22, bold=True, color="000000")
    cell.alignment = Alignment(horizontal="left")

    row = 4
    cell = ws.cell(row=row, column=2, value="Equity Valuation Report")
    cell.font = Font(size=16, bold=False, color="333333")
    cell.alignment = Alignment(horizontal="left")

    row = 6
    cell = ws.cell(row=row, column=2, value="AC444 Valuation and Security Analysis")
    cell.font = Font(size=12, bold=False, color="000000")

    row = 7
    cell = ws.cell(row=row, column=2, value="London School of Economics")
    cell.font = Font(size=12, bold=False, color="000000")

    # ---- Key information ----
    row = 9
    info_items = [
        ("Ticker:", "ENXTAM:ASM"),
        ("Currency:", "EUR (Thousands)"),
        ("Data Period:", "FY 2019 - FY 2024"),
        ("Forecast Period:", "FY 2025E - FY 2030E"),
        ("Peers:", "AIXTRON SE, Applied Materials, Lam Research"),
    ]

    label_font = Font(size=11, bold=True, color="000000")
    value_font = Font(size=11, bold=False, color="000000")

    for label, value in info_items:
        ws.cell(row=row, column=2, value=label).font = label_font
        ws.cell(row=row, column=3, value=value).font = value_font
        row += 1

    # ---- Table of Contents ----
    row += 1  # blank row
    cell = ws.cell(row=row, column=2, value="Table of Contents")
    cell.font = Font(size=14, bold=True, color="000000")
    cell.border = Border(bottom=Side(style="medium"))
    ws.cell(row=row, column=3).border = Border(bottom=Side(style="medium"))
    row += 1

    toc_sections = [
        ("1.", "Cover", "Title page and overview"),
        ("2.", "Input ASM", "Raw financial data - ASM International"),
        ("3.", "Input ASM (Reported)", "As-reported financial data - ASM International"),
        ("4.", "Input AIXTRON", "Raw financial data - AIXTRON SE"),
        ("5.", "Input AMAT", "Raw financial data - Applied Materials"),
        ("6.", "Input LRCX", "Raw financial data - Lam Research"),
        ("7.", "Standardized ASM", "Standardized financial statements - ASM"),
        ("8.", "Standardized AIXTRON", "Standardized financial statements - AIXTRON"),
        ("9.", "Standardized AMAT", "Standardized financial statements - Applied Materials"),
        ("10.", "Standardized LRCX", "Standardized financial statements - Lam Research"),
        ("11.", "Adjustments", "Adjustments and adjusted financial statements"),
        ("12.", "Ratio Analysis", "Key financial ratios and analysis"),
        ("13.", "Forecasts", "Revenue, margin, and balance sheet forecasts"),
        ("14.", "Valuation", "WACC, DCF, and PVAOI valuation"),
        ("15.", "Sensitivity", "Sensitivity analysis tables"),
        ("16.", "Comps", "Comparable company analysis"),
        ("17.", "Football Field", "Valuation range summary"),
    ]

    num_font = Font(size=10, bold=True, color="000000")
    section_font = Font(size=10, bold=False, color="000000")
    desc_font = Font(size=10, bold=False, color="666666")

    for num, section, description in toc_sections:
        ws.cell(row=row, column=2, value=f"{num}  {section}").font = num_font
        ws.cell(row=row, column=3, value=description).font = desc_font
        row += 1

    # ---- Date stamp ----
    row += 1
    ws.cell(
        row=row,
        column=2,
        value=f"Generated: {datetime.date.today().strftime('%d %B %Y')}",
    ).font = Font(size=9, italic=True, color="888888")

    # Freeze panes - nothing to freeze on Cover, but set print area nicely
    ws.sheet_properties.tabColor = "1F4E79"

    return ws


# ---------------------------------------------------------------------------
# Copy source sheet helper
# ---------------------------------------------------------------------------
def copy_source_sheet(wb, source_wb, source_sheet_name, target_sheet_name):
    """Copy all non-empty rows from a source sheet into a new target sheet.

    - Column A (labels) = black font
    - Data columns (B onwards) with numeric values = blue font (hard-coded inputs)
    - String values = black font
    - Date values = kept as-is
    - Column A width = 45, data columns = 14
    - Freeze panes at B2
    """
    src_ws = source_wb[source_sheet_name]
    tgt_ws = wb.create_sheet(title=target_sheet_name)

    max_col = src_ws.max_column or 1
    max_row = src_ws.max_row or 1

    for row_idx in range(1, max_row + 1):
        # Check if this row has any non-empty cell
        row_has_data = False
        for col_idx in range(1, max_col + 1):
            val = src_ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                row_has_data = True
                break

        if not row_has_data:
            continue

        # Copy all cells in this row
        for col_idx in range(1, max_col + 1):
            src_cell = src_ws.cell(row=row_idx, column=col_idx)
            tgt_cell = tgt_ws.cell(row=row_idx, column=col_idx)
            val = src_cell.value
            tgt_cell.value = val

            if val is None:
                continue

            # Column A = labels -> black font
            if col_idx == 1:
                tgt_cell.font = BLACK_FONT
            else:
                # Data columns: numeric -> blue, string -> black, date -> keep
                if isinstance(val, (int, float)):
                    tgt_cell.font = BLUE_FONT
                elif isinstance(val, datetime.datetime):
                    # Keep date as-is (no special font override)
                    pass
                else:
                    tgt_cell.font = BLACK_FONT

    # Set column widths: A=45, data columns=14
    widths = {1: 45}
    for col_idx in range(2, max_col + 1):
        widths[col_idx] = 14
    set_column_widths(tgt_ws, widths=widths)

    # Freeze panes at B2 (row=2, col=2)
    freeze_panes(tgt_ws, row=2, col=2)

    return tgt_ws


# ---------------------------------------------------------------------------
# Build all input tabs
# ---------------------------------------------------------------------------
def build_input_tabs(wb, source_wb):
    """Build all Input tabs by copying data from source sheets."""
    for source_sheet_name, target_sheet_name in INPUT_TAB_MAP:
        print(f"  Copying '{source_sheet_name}' -> '{target_sheet_name}'...")
        copy_source_sheet(wb, source_wb, source_sheet_name, target_sheet_name)


# ---------------------------------------------------------------------------
# Standardized ASM tab
# ---------------------------------------------------------------------------
def build_standardized_tab(wb, company_key, input_sheet_name, reported_sheet_name, output_sheet_name):
    """Build a standardized financial statements tab with reformulated IS, BS, and CF.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The output workbook.
    company_key : str
        Key into COMPANY_INFO (e.g. 'ASM', 'AIXA', 'AMAT', 'LRCX').
    input_sheet_name : str
        Quoted sheet name for formula references (e.g. "'Input ASM'").
    reported_sheet_name : str or None
        Quoted sheet name for as-reported data (e.g. "'Input ASM (Reported)'").
        None for peer companies that have no reported sheet.
    output_sheet_name : str
        Name for the output tab (e.g. 'Std ASM', 'Std AIXTRON').

    The reformulated balance sheet separates operating, investment, and
    financing items.  The accounting identity used is:

        Business Assets (= NOA + Investment Assets)
        = Net Debt + Equity  (= Invested Capital)

    All data cells are Excel formulas referencing the Input tabs.
    """
    ws = wb.create_sheet(title=output_sheet_name)
    MAX_COL = 7  # columns A-G

    # Shortcuts to data maps
    info = COMPANY_INFO[company_key]
    is_map = info['is_map']
    cf_map = info['cf_map']
    bs_map = info['bs_map']

    input_sheet = input_sheet_name

    # --- Column widths ---
    widths = {1: 45}
    for c in range(2, 8):
        widths[c] = 14
    set_column_widths(ws, widths=widths)

    # --- Helper: write a cross-ref formula across years (with ISNUMBER guard) ---
    def write_crossref_row(row, sheet, src_row, fmt=FMT_CURRENCY, guard=True):
        """Write cross-sheet reference formulas across cols B-G."""
        for ci in range(2, 8):
            cl = get_column_letter(ci)
            cell = ws.cell(row=row, column=ci)
            if guard:
                cell.value = f"=IF(ISNUMBER({sheet}!{cl}{src_row}),{sheet}!{cl}{src_row},0)"
            else:
                cell.value = f"={sheet}!{cl}{src_row}"
            style_crossref_cell(cell, fmt)

    def write_zero_row(row, fmt=FMT_CURRENCY):
        """Write 0 across cols B-G for missing map keys."""
        for ci in range(2, 8):
            cell = ws.cell(row=row, column=ci, value=0)
            style_formula_cell(cell, fmt)

    def write_crossref_or_zero(row, sheet, map_dict, key, fmt=FMT_CURRENCY, guard=True):
        """Write crossref if key exists in map, else write 0."""
        if key in map_dict:
            write_crossref_row(row, sheet, map_dict[key], fmt=fmt, guard=guard)
        else:
            write_zero_row(row, fmt=fmt)

    def write_formula_row(row, formulas_by_col, fmt=FMT_CURRENCY, bold=False):
        """Write formula strings across cols B-G. formulas_by_col is a callable(col_letter)->str."""
        for ci in range(2, 8):
            cl = get_column_letter(ci)
            cell = ws.cell(row=row, column=ci)
            cell.value = formulas_by_col(cl)
            style_formula_cell(cell, fmt)
            if bold:
                cell.font = BLACK_BOLD

    # =========================================================================
    # Row 1-4: Title, subtitle, year headers
    # =========================================================================
    r = 1
    cell = ws.cell(row=r, column=1, value=f"Standardized Financial Statements - {info['name']}")
    cell.font = Font(size=14, bold=True, color="000000")

    r = 2
    ws.cell(row=r, column=1, value="(EUR 000)").font = Font(size=10, italic=True, color="666666")

    r = 3
    write_year_headers(ws, r, 2, YEARS)

    r = 4  # blank

    # =========================================================================
    # Section 1: Assumptions (rows 5-7)
    # =========================================================================
    r = 5
    write_label(ws, r, 1, "Operating cash %")
    for ci in range(2, 8):
        cell = ws.cell(row=r, column=ci, value=0.02)
        style_assumption_cell(cell, FMT_PERCENT)

    r = 6
    write_label(ws, r, 1, "Tax rate (effective)")
    write_crossref_or_zero(r, input_sheet, is_map, 'effective_tax_rate', fmt=FMT_PERCENT, guard=True)

    r = 7  # blank

    # =========================================================================
    # Section 2: INCOME STATEMENT (row 8 onwards)
    # =========================================================================
    r = 8
    style_section_header(ws, r, MAX_COL, "INCOME STATEMENT")

    # Row 9: Revenue
    r = 9
    write_label(ws, r, 1, "Revenue")
    write_crossref_row(r, input_sheet, is_map['revenue'])

    # Row 10: Cost of Sales
    r = 10
    write_label(ws, r, 1, "Cost of Sales")
    write_crossref_row(r, input_sheet, is_map['cogs'])

    # Row 11: Gross Profit = Revenue - COGS
    r = 11
    write_label(ws, r, 1, "Gross Profit", bold=True)
    write_formula_row(r, lambda cl: f"={cl}9-{cl}10", bold=True)

    # Row 12: SG&A
    r = 12
    write_label(ws, r, 1, "SG&A", indent=2)
    write_crossref_row(r, input_sheet, is_map['sga'])

    # Row 13: R&D
    r = 13
    write_label(ws, r, 1, "R&D", indent=2)
    write_crossref_row(r, input_sheet, is_map['rd'])

    # Row 14: D&A (from CF supplemental -- memo line, already embedded in COGS)
    r = 14
    write_label(ws, r, 1, "Depreciation & Amortization", indent=2)
    write_crossref_row(r, input_sheet, cf_map['da_total'])

    # Row 15: Other Operating Expense
    r = 15
    write_label(ws, r, 1, "Other Operating Expense", indent=2)
    write_crossref_or_zero(r, input_sheet, is_map, 'other_operating_exp')

    # Row 16: Total Operating Expense = COGS + SGA + RD + DA + Other
    r = 16
    write_label(ws, r, 1, "Total Operating Expense", bold=True)
    write_formula_row(r, lambda cl: f"={cl}10+{cl}12+{cl}13+{cl}14+{cl}15", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 17: Recurring Operating Profit = Revenue - Total OpEx
    r = 17
    write_label(ws, r, 1, "Recurring Operating Profit", bold=True)
    write_formula_row(r, lambda cl: f"={cl}9-{cl}16", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 18: blank
    r = 18

    # Row 19: Non-recurring items header
    r = 19
    write_label(ws, r, 1, "Non-recurring items:", bold=True)

    # Row 20: Gain/Loss on Sale of Investments
    r = 20
    write_label(ws, r, 1, "Gain/Loss on Sale of Investments", indent=2)
    write_crossref_or_zero(r, input_sheet, is_map, 'gain_loss_sale_invest')

    # Row 21: Gain/Loss on Sale of Assets
    r = 21
    write_label(ws, r, 1, "Gain/Loss on Sale of Assets", indent=2)
    write_crossref_or_zero(r, input_sheet, is_map, 'gain_loss_sale_assets')

    # Row 22: Asset Writedown
    r = 22
    write_label(ws, r, 1, "Asset Writedown", indent=2)
    write_crossref_or_zero(r, input_sheet, is_map, 'asset_writedown')

    # Row 23: Other Unusual Items
    r = 23
    write_label(ws, r, 1, "Other Unusual Items", indent=2)
    # Combine restructuring + other_unusual + merger_restruct if present
    _unusual_parts = []
    for _ukey in ['other_unusual', 'restructuring', 'merger_restruct', 'insurance_settlements']:
        if _ukey in is_map:
            _unusual_parts.append((_ukey, is_map[_ukey]))
    if len(_unusual_parts) == 0:
        write_zero_row(r)
    elif len(_unusual_parts) == 1:
        write_crossref_row(r, input_sheet, _unusual_parts[0][1])
    else:
        # Sum multiple unusual items
        for ci in range(2, 8):
            cl = get_column_letter(ci)
            cell = ws.cell(row=r, column=ci)
            parts = [f"IF(ISNUMBER({input_sheet}!{cl}{src_row}),{input_sheet}!{cl}{src_row},0)"
                     for _, src_row in _unusual_parts]
            cell.value = "=" + "+".join(parts)
            style_crossref_cell(cell, FMT_CURRENCY)

    # Row 24: Total Non-recurring = SUM(20:23)
    r = 24
    write_label(ws, r, 1, "Total Non-recurring", bold=True)
    write_formula_row(r, lambda cl: f"=SUM({cl}20:{cl}23)", bold=True)

    # Row 25: EBIT = Recurring Op Profit + Non-recurring
    r = 25
    write_label(ws, r, 1, "EBIT", bold=True)
    write_formula_row(r, lambda cl: f"={cl}17+{cl}24", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 26: blank
    r = 26

    # Row 27: Investment Income (Equity Method)
    r = 27
    write_label(ws, r, 1, "Investment Income (Equity Method)", indent=2)
    write_crossref_or_zero(r, input_sheet, is_map, 'income_from_affiliates')

    # Row 28: Interest Income
    r = 28
    write_label(ws, r, 1, "Interest Income", indent=2)
    write_crossref_row(r, input_sheet, is_map['interest_income'])

    # Row 29: Interest Expense
    r = 29
    write_label(ws, r, 1, "Interest Expense", indent=2)
    write_crossref_row(r, input_sheet, is_map['interest_expense'])

    # Row 30: FX Gains/(Losses)
    r = 30
    write_label(ws, r, 1, "FX Gains/(Losses)", indent=2)
    write_crossref_row(r, input_sheet, is_map['fx_gains'])

    # Row 31: Profit Before Tax = EBIT + 27 + 28 + 29 + 30
    r = 31
    write_label(ws, r, 1, "Profit Before Tax", bold=True)
    write_formula_row(r, lambda cl: f"={cl}25+{cl}27+{cl}28+{cl}29+{cl}30", bold=True)

    # Row 32: Tax Expense
    r = 32
    write_label(ws, r, 1, "Tax Expense", indent=2)
    write_crossref_row(r, input_sheet, is_map['tax_expense'])

    # Row 33: Net Income = PBT - Tax
    r = 33
    write_label(ws, r, 1, "Net Income", bold=True)
    write_formula_row(r, lambda cl: f"={cl}31-{cl}32", bold=True)
    style_double_line_row(ws, r, MAX_COL)

    # Row 34: blank
    r = 34

    # Row 35: Effective Tax Rate
    r = 35
    write_label(ws, r, 1, "Effective Tax Rate")
    write_formula_row(r, lambda cl: f"=IF({cl}31<>0,{cl}32/{cl}31,0)", fmt=FMT_PERCENT)

    # Row 36: NOPAT = Recurring Op Profit * (1 - ETR)
    r = 36
    write_label(ws, r, 1, "NOPAT", bold=True)
    write_formula_row(r, lambda cl: f"={cl}17*(1-{cl}35)", bold=True)

    # Row 37: blank
    r = 37

    # Row 38: Shares Outstanding (diluted)
    r = 38
    write_label(ws, r, 1, "Shares Outstanding (diluted)")
    write_crossref_row(r, input_sheet, is_map['diluted_shares'])

    # Row 39: EPS (diluted)  -- NI in thousands, shares in actuals
    r = 39
    write_label(ws, r, 1, "EPS (diluted)")
    write_formula_row(r, lambda cl: f"=IF({cl}38<>0,{cl}33/{cl}38*1000,0)", fmt='#,##0.00;(#,##0.00);"-"')

    # Row 40: Dividends per Share
    r = 40
    write_label(ws, r, 1, "Dividends per Share")
    write_crossref_or_zero(r, input_sheet, is_map, 'dividends_per_share', fmt='#,##0.00;(#,##0.00);"-"')

    # Row 41-42: blank
    r = 41
    r = 42

    # =========================================================================
    # Section 3: BALANCE SHEET (row 43 onwards)
    #
    # Uses CapIQ standardized data for consistency.
    # =========================================================================
    r = 43
    style_section_header(ws, r, MAX_COL, "BALANCE SHEET (Reformulated)")

    # --- Operating Working Capital ---
    r = 44
    write_label(ws, r, 1, "Operating Working Capital:", bold=True)

    # Row 45: Trade Receivables (CapIQ: includes contract assets)
    r = 45
    write_label(ws, r, 1, "Trade Receivables", indent=2)
    write_crossref_row(r, input_sheet, bs_map['accounts_receivable'])

    # Row 46: + Other Receivables (Income Tax Receivable etc.)
    r = 46
    write_label(ws, r, 1, "+ Other Receivables", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'other_receivables')

    # Row 47: + Inventories
    r = 47
    write_label(ws, r, 1, "+ Inventories", indent=2)
    write_crossref_row(r, input_sheet, bs_map['inventory'])

    # Row 48: + Prepaid & Other CA
    r = 48
    write_label(ws, r, 1, "+ Prepaid & Other CA", indent=2)
    for ci in range(2, 8):
        cl = get_column_letter(ci)
        cell = ws.cell(row=r, column=ci)
        prep_row = bs_map['prepaid_exp']
        oca_row = bs_map['other_current_assets']
        cell.value = (
            f"=IF(ISNUMBER({input_sheet}!{cl}{prep_row}),{input_sheet}!{cl}{prep_row},0)"
            f"+IF(ISNUMBER({input_sheet}!{cl}{oca_row}),{input_sheet}!{cl}{oca_row},0)"
        )
        style_crossref_cell(cell, FMT_CURRENCY)

    # Row 49: - Accounts Payable
    r = 49
    write_label(ws, r, 1, "- Accounts Payable", indent=2)
    write_crossref_row(r, input_sheet, bs_map['accounts_payable'])

    # Row 50: - Accrued Expenses
    r = 50
    write_label(ws, r, 1, "- Accrued Expenses", indent=2)
    write_crossref_row(r, input_sheet, bs_map['accrued_exp'])

    # Row 51: - Unearned Revenue / Contract Liabilities
    r = 51
    write_label(ws, r, 1, "- Unearned Revenue / Contract Liabilities", indent=2)
    write_crossref_row(r, input_sheet, bs_map['unearned_revenue_curr'])

    # Row 52: - Income Taxes Payable
    r = 52
    write_label(ws, r, 1, "- Income Taxes Payable", indent=2)
    write_crossref_row(r, input_sheet, bs_map['current_income_tax'])

    # Row 53: - Other Current Liabilities
    r = 53
    write_label(ws, r, 1, "- Other Current Liabilities", indent=2)
    write_crossref_row(r, input_sheet, bs_map['other_current_liab'])

    # Row 54: = Operating Working Capital
    OWC_ROW = 54
    r = OWC_ROW
    write_label(ws, r, 1, "= Operating Working Capital", bold=True)
    write_formula_row(r, lambda cl: f"={cl}45+{cl}46+{cl}47+{cl}48-{cl}49-{cl}50-{cl}51-{cl}52-{cl}53", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 55: blank
    r = 55

    # --- Net Non-Current Operating Assets ---
    r = 56
    write_label(ws, r, 1, "Net Non-Current Operating Assets:", bold=True)

    # Row 57: PP&E (net) - CapIQ net_ppe (= Reported PPE + ROU)
    r = 57
    write_label(ws, r, 1, "PP&E (net, incl. ROU)", indent=2)
    write_crossref_row(r, input_sheet, bs_map['net_ppe'])

    # Row 58: + Goodwill
    r = 58
    write_label(ws, r, 1, "+ Goodwill", indent=2)
    write_crossref_row(r, input_sheet, bs_map['goodwill'])

    # Row 59: + Other Intangible Assets
    r = 59
    write_label(ws, r, 1, "+ Other Intangible Assets", indent=2)
    write_crossref_row(r, input_sheet, bs_map['other_intangibles'])

    # Row 60: + Deferred Tax Assets
    r = 60
    write_label(ws, r, 1, "+ Deferred Tax Assets", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'deferred_tax_assets')

    # Row 61: + Deferred Charges & Other LT Intangibles
    r = 61
    write_label(ws, r, 1, "+ Deferred Charges & LT Intangibles", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'deferred_charges_lt')

    # Row 62: + Other Non-Current Assets
    r = 62
    write_label(ws, r, 1, "+ Other Non-Current Assets", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'other_lt_assets')

    # Row 63: - Deferred Tax Liabilities
    r = 63
    write_label(ws, r, 1, "- Deferred Tax Liabilities", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'deferred_tax_liab')

    # Row 64: - Other Non-Current Liabilities (incl. pension if applicable)
    r = 64
    write_label(ws, r, 1, "- Other Non-Current Liabilities", indent=2)
    if 'pension_post_retire' in bs_map:
        # Sum other_non_current_liab + pension_post_retire
        for ci in range(2, 8):
            cl = get_column_letter(ci)
            cell = ws.cell(row=r, column=ci)
            ncl_row = bs_map['other_non_current_liab']
            pen_row = bs_map['pension_post_retire']
            cell.value = (
                f"=IF(ISNUMBER({input_sheet}!{cl}{ncl_row}),{input_sheet}!{cl}{ncl_row},0)"
                f"+IF(ISNUMBER({input_sheet}!{cl}{pen_row}),{input_sheet}!{cl}{pen_row},0)"
            )
            style_crossref_cell(cell, FMT_CURRENCY)
    else:
        write_crossref_row(r, input_sheet, bs_map['other_non_current_liab'])

    # Note: Lease liabilities (non-current) are classified as financing,
    # not deducted from operating NCA.

    # Row 65: = Net Non-Current Operating Assets
    NET_NCA_ROW = 65
    r = NET_NCA_ROW
    write_label(ws, r, 1, "= Net Non-Current Operating Assets", bold=True)
    write_formula_row(r, lambda cl: f"={cl}57+{cl}58+{cl}59+{cl}60+{cl}61+{cl}62-{cl}63-{cl}64", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 66: blank
    r = 66

    # Row 67: = Net Operating Assets = OWC + Net NCA
    NOA_ROW = 67
    r = NOA_ROW
    write_label(ws, r, 1, "= Net Operating Assets", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{OWC_ROW}+{cl}{NET_NCA_ROW}", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 68: blank
    r = 68

    # --- Investment Assets ---
    r = 69
    write_label(ws, r, 1, "Investment Assets:", bold=True)

    # Row 70: Long-Term Investments
    r = 70
    write_label(ws, r, 1, "Long-Term Investments", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'lt_investments')

    # Row 71: = Total Investment Assets (single line item here)
    INVEST_ASSETS_ROW = 71
    r = INVEST_ASSETS_ROW
    write_label(ws, r, 1, "= Total Investment Assets", bold=True)
    write_formula_row(r, lambda cl: f"={cl}70", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 72: blank
    r = 72

    # Row 73: = Business Assets = NOA + Investment Assets
    BIZ_ASSETS_ROW = 73
    r = BIZ_ASSETS_ROW
    write_label(ws, r, 1, "= Business Assets", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{NOA_ROW}+{cl}{INVEST_ASSETS_ROW}", bold=True)
    style_double_line_row(ws, r, MAX_COL)

    # Row 74: blank
    r = 74

    # --- Financing ---
    r = 75
    style_section_header(ws, r, MAX_COL, "FINANCING")

    # Row 76: Debt items -- handle companies with explicit LT debt vs lease-only
    r = 76
    write_label(ws, r, 1, "Lease Liabilities (current)", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'current_lease_liab')

    # Row 77: + Lease Liabilities (non-current) / LT Debt
    r = 77
    write_label(ws, r, 1, "+ Lease Liabilities (non-current)", indent=2)
    write_crossref_or_zero(r, input_sheet, bs_map, 'lt_leases')

    # Row 78: = Total Debt
    TOTAL_DEBT_ROW = 78
    r = TOTAL_DEBT_ROW
    write_label(ws, r, 1, "= Total Debt", bold=True)
    # For companies with explicit debt (lt_debt, current_lt_debt, st_borrowings),
    # use the CapIQ total_debt figure; for ASM/AIXA (lease-only), sum row 76+77.
    if 'total_debt' in bs_map and ('lt_debt' in bs_map or 'st_borrowings' in bs_map):
        # Company has financial debt beyond leases -- use CapIQ total_debt
        write_crossref_row(r, input_sheet, bs_map['total_debt'])
    else:
        # Lease-only debt (ASM, AIXA) -- sum current + non-current leases
        write_formula_row(r, lambda cl: f"={cl}76+{cl}77", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 79: Cash & Equivalents
    r = 79
    write_label(ws, r, 1, "Cash & Equivalents", indent=2)
    write_crossref_row(r, input_sheet, bs_map['cash_st_investments'])

    # Row 80: = Net Debt = Debt - Cash
    NET_DEBT_ROW = 80
    r = NET_DEBT_ROW
    write_label(ws, r, 1, "= Net Debt", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{TOTAL_DEBT_ROW}-{cl}79", bold=True)

    # Row 81: Group Equity
    EQUITY_ROW = 81
    r = EQUITY_ROW
    write_label(ws, r, 1, "Group Equity")
    write_crossref_row(r, input_sheet, bs_map['total_equity'])

    # Row 82: = Invested Capital = Net Debt + Equity
    # (Business Assets = NOA + Investment = Net Debt + Equity)
    INVESTED_CAP_ROW = 82
    r = INVESTED_CAP_ROW
    write_label(ws, r, 1, "= Invested Capital", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{NET_DEBT_ROW}+{cl}{EQUITY_ROW}", bold=True)
    style_double_line_row(ws, r, MAX_COL)

    # Row 83: BS Check (should = 0)
    BS_CHECK_ROW = 83
    r = BS_CHECK_ROW
    write_label(ws, r, 1, "BS Check (should = 0)", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{BIZ_ASSETS_ROW}-{cl}{INVESTED_CAP_ROW}", bold=True)

    # Row 84-85: blank
    r = 84
    r = 85

    # =========================================================================
    # Section 4: CASH FLOW (row 86 onwards)
    # =========================================================================
    r = 86
    style_section_header(ws, r, MAX_COL, "FREE CASH FLOW")

    # Row 87: NOPAT (reference to IS NOPAT row 36)
    NOPAT_CF_ROW = 87
    r = NOPAT_CF_ROW
    write_label(ws, r, 1, "NOPAT")
    write_formula_row(r, lambda cl: f"={cl}36")

    # Row 88: + D&A
    DA_CF_ROW = 88
    r = DA_CF_ROW
    write_label(ws, r, 1, "+ D&A")
    write_crossref_row(r, input_sheet, cf_map['da_total'])

    # Row 89: - Change in OWC (OWC(t) - OWC(t-1)), 2019 = 0
    CHG_OWC_ROW = 89
    r = CHG_OWC_ROW
    write_label(ws, r, 1, "- Change in OWC")
    # Col B (2019) = 0
    cell = ws.cell(row=r, column=2, value=0)
    style_formula_cell(cell, FMT_CURRENCY)
    # Cols C-G = OWC(t) - OWC(t-1)
    for ci in range(3, 8):
        cl = get_column_letter(ci)
        prev_cl = get_column_letter(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={cl}{OWC_ROW}-{prev_cl}{OWC_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)

    # Row 90: - CapEx (absolute value -- source is negative)
    CAPEX_CF_ROW = 90
    r = CAPEX_CF_ROW
    write_label(ws, r, 1, "- CapEx")
    for ci in range(2, 8):
        cl = get_column_letter(ci)
        cell = ws.cell(row=r, column=ci)
        src_row = cf_map['capex']
        cell.value = f"=ABS(IF(ISNUMBER({input_sheet}!{cl}{src_row}),{input_sheet}!{cl}{src_row},0))"
        style_crossref_cell(cell, FMT_CURRENCY)

    # Row 91: - Purchase of Intangibles (absolute value)
    PURCH_INTANG_ROW = 91
    r = PURCH_INTANG_ROW
    write_label(ws, r, 1, "- Purchase of Intangibles")
    if 'purchase_intangibles' in cf_map:
        for ci in range(2, 8):
            cl = get_column_letter(ci)
            cell = ws.cell(row=r, column=ci)
            src_row = cf_map['purchase_intangibles']
            cell.value = f"=ABS(IF(ISNUMBER({input_sheet}!{cl}{src_row}),{input_sheet}!{cl}{src_row},0))"
            style_crossref_cell(cell, FMT_CURRENCY)
    else:
        write_zero_row(r)

    # Row 92: = Operating CF after investment
    OP_CF_ROW = 92
    r = OP_CF_ROW
    write_label(ws, r, 1, "= Operating CF after investment", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{NOPAT_CF_ROW}+{cl}{DA_CF_ROW}-{cl}{CHG_OWC_ROW}-{cl}{CAPEX_CF_ROW}-{cl}{PURCH_INTANG_ROW}", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 93: blank
    r = 93

    # Row 94: + Net Investment Profit after tax = Invest Income * (1 - ETR)
    NET_INV_PROFIT_ROW = 94
    r = NET_INV_PROFIT_ROW
    write_label(ws, r, 1, "+ Net Investment Profit after tax")
    write_formula_row(r, lambda cl: f"={cl}27*(1-{cl}35)")

    # Row 95: - Change in Investment Assets = InvestAssets(t) - InvestAssets(t-1)
    CHG_INVEST_ROW = 95
    r = CHG_INVEST_ROW
    write_label(ws, r, 1, "- Change in Investment Assets")
    # Col B (2019) = 0
    cell = ws.cell(row=r, column=2, value=0)
    style_formula_cell(cell, FMT_CURRENCY)
    for ci in range(3, 8):
        cl = get_column_letter(ci)
        prev_cl = get_column_letter(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={cl}{INVEST_ASSETS_ROW}-{prev_cl}{INVEST_ASSETS_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)

    # Row 96: = FCF to Debt & Equity
    FCF_DE_ROW = 96
    r = FCF_DE_ROW
    write_label(ws, r, 1, "= FCF to Debt & Equity", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{OP_CF_ROW}+{cl}{NET_INV_PROFIT_ROW}-{cl}{CHG_INVEST_ROW}", bold=True)
    style_total_row(ws, r, MAX_COL)

    # Row 97: - Interest Expense after tax = Interest Exp * (1 - ETR)
    INT_AFTER_TAX_ROW = 97
    r = INT_AFTER_TAX_ROW
    write_label(ws, r, 1, "- Interest Expense after tax")
    # Interest expense is row 29 in IS; typically negative (expense)
    write_formula_row(r, lambda cl: f"={cl}29*(1-{cl}35)")

    # Row 98: + Change in Debt = Debt(t) - Debt(t-1)
    CHG_DEBT_ROW = 98
    r = CHG_DEBT_ROW
    write_label(ws, r, 1, "+ Change in Debt")
    cell = ws.cell(row=r, column=2, value=0)
    style_formula_cell(cell, FMT_CURRENCY)
    for ci in range(3, 8):
        cl = get_column_letter(ci)
        prev_cl = get_column_letter(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={cl}{TOTAL_DEBT_ROW}-{prev_cl}{TOTAL_DEBT_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)

    # Row 99: = FCF to Equity
    FCF_EQ_ROW = 99
    r = FCF_EQ_ROW
    write_label(ws, r, 1, "= FCF to Equity", bold=True)
    write_formula_row(r, lambda cl: f"={cl}{FCF_DE_ROW}-{cl}{INT_AFTER_TAX_ROW}+{cl}{CHG_DEBT_ROW}", bold=True)
    style_double_line_row(ws, r, MAX_COL)

    # --- Freeze panes at B5 ---
    freeze_panes(ws, row=5, col=2)

    return ws


def build_standardized_asm(wb):
    """Build the 'Std ASM' sheet -- delegates to generic function."""
    return build_standardized_tab(wb, 'ASM', "'Input ASM'", "'Input ASM (Reported)'", 'Std ASM')


# ---------------------------------------------------------------------------
# Adjustments documentation tab
# ---------------------------------------------------------------------------
def build_adjustments_tab(wb):
    """Build the 'Adjustments' tab documenting accounting analysis performed.

    This is a text-only documentation tab -- no Excel formulas needed.
    """
    ws = wb.create_sheet(title="Adjustments")

    # Column widths
    set_column_widths(ws, widths={1: 5, 2: 30, 3: 80})

    # Row 1: Title
    r = 1
    cell = ws.cell(row=r, column=1, value="Accounting Analysis & Adjustments - ASM International NV")
    cell.font = Font(size=14, bold=True, color="000000")

    # Row 3: Section header
    r = 3
    cell = ws.cell(row=r, column=1, value="Areas Examined")
    cell.font = Font(size=12, bold=True, color="000000")

    # ----- Areas data -----
    areas = [
        {
            "name": "1. R&D Capitalization Policy",
            "description": "ASM capitalizes development expenditure (significant amounts).",
            "finding": "Compare with peers: all semi equipment companies capitalize similarly.",
            "decision": "No Adjustment",
            "justification": "Consistent with industry practice and peers.",
        },
        {
            "name": "2. Equity Method Investment (ASMPT)",
            "description": "~25% stake in ASM Pacific Technology.",
            "finding": "Already separated as Investment Income in standardized IS.",
            "decision": "No Adjustment",
            "justification": "No further adjustment needed.",
        },
        {
            "name": "3. FX Gains/Losses",
            "description": "Highly volatile: ranged from -23M (2020) to +45M (2024).",
            "finding": "Already classified below EBIT in standardized IS (non-operating).",
            "decision": "No Adjustment",
            "justification": "Already handled in standardization.",
        },
        {
            "name": "4. Stock-Based Compensation",
            "description": "EUR 42M in 2024, growing over time.",
            "finding": "Already included in operating expenses in CapIQ data.",
            "decision": "No Adjustment",
            "justification": "No adjustment needed.",
        },
        {
            "name": "5. Goodwill (LPE Acquisition 2022)",
            "description": "EUR 321M from LPE acquisition.",
            "finding": "No impairment charges taken.",
            "decision": "No Adjustment",
            "justification": "Monitor for future impairment risk.",
        },
        {
            "name": "6. Asset Writedowns & Unusual Items",
            "description": "Small and sporadic.",
            "finding": "Already classified as non-recurring in standardized IS.",
            "decision": "No Adjustment",
            "justification": "No adjustment needed.",
        },
    ]

    r = 5  # start writing areas
    for area in areas:
        # Area name (bold)
        cell = ws.cell(row=r, column=1, value=area["name"])
        cell.font = BLACK_BOLD

        r += 1
        ws.cell(row=r, column=2, value="Description:").font = BLACK_BOLD
        ws.cell(row=r, column=3, value=area["description"]).font = BLACK_FONT

        r += 1
        ws.cell(row=r, column=2, value="Finding:").font = BLACK_BOLD
        ws.cell(row=r, column=3, value=area["finding"]).font = BLACK_FONT

        r += 1
        ws.cell(row=r, column=2, value="Decision:").font = BLACK_BOLD
        ws.cell(row=r, column=3, value=area["decision"]).font = BLACK_FONT

        r += 1
        ws.cell(row=r, column=2, value="Justification:").font = BLACK_BOLD
        ws.cell(row=r, column=3, value=area["justification"]).font = BLACK_FONT

        r += 2  # blank row between areas

    # Summary
    r += 1
    cell = ws.cell(row=r, column=1, value="Summary")
    cell.font = Font(size=12, bold=True, color="000000")
    r += 1
    ws.cell(
        row=r,
        column=1,
        value="No material adjustments required. The standardization process already handles "
              "all analytical reclassifications. Adjusted tabs pass through from Standardized tabs.",
    ).font = BLACK_FONT

    return ws


# ---------------------------------------------------------------------------
# Adjusted statement tabs (pass-through from Standardized)
# ---------------------------------------------------------------------------
def build_adjusted_tab(wb, company_key, std_sheet_name, output_sheet_name):
    """Build an Adjusted tab that mirrors a Standardized tab via cross-sheet refs.

    Since no material adjustments are being made, every data cell is a simple
    cross-sheet reference to the corresponding Standardized tab cell.

    Parameters
    ----------
    wb : openpyxl.Workbook
    company_key : str
        Key into COMPANY_INFO (e.g. 'ASM').
    std_sheet_name : str
        Name of the source Standardized sheet (e.g. 'Std ASM').
    output_sheet_name : str
        Name for the new Adjusted sheet (e.g. 'Adj ASM').
    """
    src_ws = wb[std_sheet_name]
    ws = wb.create_sheet(title=output_sheet_name)

    MAX_COL = 7  # columns A-G

    # Column widths
    widths = {1: 45}
    for c in range(2, 8):
        widths[c] = 14
    set_column_widths(ws, widths=widths)

    # Iterate through every row/col in the Standardized sheet and mirror
    max_row = src_ws.max_row or 1

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, MAX_COL + 1):
            src_cell = src_ws.cell(row=row_idx, column=col_idx)
            tgt_cell = ws.cell(row=row_idx, column=col_idx)

            if src_cell.value is None:
                # Preserve formatting for section header fills, borders, etc.
                if src_cell.fill and src_cell.fill.fill_type == "solid":
                    tgt_cell.fill = _copy.copy(src_cell.fill)
                if src_cell.border:
                    tgt_cell.border = _copy.copy(src_cell.border)
                continue

            if col_idx == 1:
                # Column A: copy label text with same font/formatting
                tgt_cell.value = src_cell.value
                if src_cell.font:
                    tgt_cell.font = _copy.copy(src_cell.font)
                if src_cell.alignment:
                    tgt_cell.alignment = _copy.copy(src_cell.alignment)
            else:
                # Data columns (B-G): write cross-sheet reference
                cl = get_column_letter(col_idx)
                tgt_cell.value = f"='{std_sheet_name}'!{cl}{row_idx}"
                style_crossref_cell(tgt_cell, FMT_CURRENCY)

                # Preserve special number formats from source
                if src_cell.number_format and src_cell.number_format != 'General':
                    tgt_cell.number_format = src_cell.number_format

                # Preserve bold from source (for total rows)
                if src_cell.font and src_cell.font.bold:
                    tgt_cell.font = Font(color="008000", bold=True)

            # Preserve fills (section headers, etc.) and borders
            if src_cell.fill and src_cell.fill.fill_type == "solid":
                tgt_cell.fill = _copy.copy(src_cell.fill)
            if src_cell.border and src_cell.border != Border():
                tgt_cell.border = _copy.copy(src_cell.border)

    # Freeze panes at B5 (same as standardized)
    freeze_panes(ws, row=5, col=2)

    return ws


# ---------------------------------------------------------------------------
# Ratio Definitions tab (reference / documentation)
# ---------------------------------------------------------------------------
def build_ratio_definitions(wb):
    """Build the 'Ratio Definitions' tab -- a text-only reference of ratio formulas."""
    ws = wb.create_sheet(title="Ratio Definitions")

    # Column widths
    set_column_widths(ws, widths={1: 5, 2: 40, 3: 70})

    # Title
    cell = ws.cell(row=1, column=1, value="Ratio Definitions & Formulas")
    cell.font = Font(size=14, bold=True, color="000000")

    sections = [
        ("Profitability", [
            ("ROE", "Net Income / Average Total Equity"),
            ("DuPont Decomposition", "ROS x Asset Turnover x Financial Leverage"),
            ("RNOA", "NOPAT Margin x NOA Turnover"),
            ("Gross Profit Margin", "Gross Profit / Revenue"),
            ("EBIT Margin", "EBIT / Revenue"),
            ("NOPAT Margin", "NOPAT / Revenue"),
        ]),
        ("Asset Management", [
            ("Receivables Turnover", "Revenue / Avg Trade Receivables"),
            ("Days Receivable", "365 / Receivables Turnover"),
            ("Inventory Turnover", "COGS / Avg Inventory"),
            ("Days Inventory", "365 / Inventory Turnover"),
            ("Payables Turnover", "COGS / Avg Trade Payables"),
            ("Days Payable", "365 / Payables Turnover"),
            ("Cash Conversion Cycle", "Days Inventory + Days Receivable - Days Payable"),
        ]),
        ("Liquidity", [
            ("OWC / Revenue", "Operating Working Capital / Revenue"),
            ("Cash / Total Debt", "Cash & Equivalents / Total Debt"),
        ]),
        ("Solvency", [
            ("Debt-to-Equity", "Total Debt / Total Equity"),
            ("Interest Coverage", "EBIT / |Interest Expense|"),
        ]),
        ("Growth", [
            ("Revenue Growth", "(Revenue_t / Revenue_{t-1}) - 1"),
            ("Sustainable Growth", "(1 - Payout Ratio) x ROE"),
        ]),
    ]

    r = 3
    for section_name, ratios in sections:
        # Section header
        cell = ws.cell(row=r, column=1, value=section_name)
        cell.font = Font(size=12, bold=True, color="000000")
        cell.border = Border(bottom=Side(style="medium"))
        ws.cell(row=r, column=2).border = Border(bottom=Side(style="medium"))
        ws.cell(row=r, column=3).border = Border(bottom=Side(style="medium"))
        r += 1

        for name, formula in ratios:
            ws.cell(row=r, column=2, value=name).font = BLACK_BOLD
            ws.cell(row=r, column=3, value=formula).font = BLACK_FONT
            r += 1

        r += 1  # blank row between sections

    return ws


# ---------------------------------------------------------------------------
# Ratio Analysis tab (per company)
# ---------------------------------------------------------------------------
def build_ratio_tab(wb, adj_sheet_name, output_sheet_name):
    """Build a ratio analysis tab with Excel formulas referencing an Adjusted tab.

    Parameters
    ----------
    wb : openpyxl.Workbook
    adj_sheet_name : str
        Name of the Adjusted sheet (e.g. 'Adj ASM').
    output_sheet_name : str
        Name for the output ratio sheet (e.g. 'Ratios ASM').
    """
    ws = wb.create_sheet(title=output_sheet_name)
    MAX_COL = 7  # A-G

    adj = f"'{adj_sheet_name}'"

    # Column widths
    widths = {1: 40}
    for c in range(2, 8):
        widths[c] = 14
    set_column_widths(ws, widths=widths)

    # --- Title ---
    ws.cell(row=1, column=1,
            value=f"Ratio Analysis - {adj_sheet_name.replace('Adj ', '')}").font = Font(
        size=14, bold=True, color="000000")
    ws.cell(row=2, column=1, value="(derived from Adjusted statements)").font = Font(
        size=10, italic=True, color="666666")

    # Year headers in row 3
    write_year_headers(ws, 3, 2, YEARS)

    # --- Helpers ---
    def _label(row, text, bold=False, indent=0):
        write_label(ws, row, 1, text, bold=bold, indent=indent)

    def _formula(row, col_idx, formula_str, fmt=FMT_PERCENT):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = formula_str
        style_formula_cell(cell, fmt)

    def _na(row, col_idx, fmt=FMT_PERCENT):
        """Write N/A for the first year where averages are not computable."""
        cell = ws.cell(row=row, column=col_idx, value="N/A")
        cell.font = Font(color="999999", italic=True)
        if fmt:
            cell.number_format = fmt

    def _cols():
        """Return list of (col_index, col_letter, prev_col_letter_or_None)."""
        result = []
        for ci in range(2, 8):
            cl = get_column_letter(ci)
            prev_cl = get_column_letter(ci - 1) if ci > 2 else None
            result.append((ci, cl, prev_cl))
        return result

    # =========================================================================
    # Section: PROFITABILITY
    # =========================================================================
    r = 5
    style_section_header(ws, r, MAX_COL, "PROFITABILITY")

    # --- ROE ---
    r = 6
    _label(r, "ROE", bold=True)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}81,{adj}!{cl}81)=0,0,"
                     f"{adj}!{cl}33/AVERAGE({adj}!{prev_cl}81,{adj}!{cl}81))")

    # --- DuPont: Net Profit Margin (ROS) ---
    r = 7
    _label(r, "  Net Profit Margin (ROS)", indent=2)
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}33/{adj}!{cl}9)")

    # --- DuPont: Asset Turnover ---
    r = 8
    _label(r, "  x Asset Turnover", indent=2)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}73,{adj}!{cl}73)=0,0,"
                     f"{adj}!{cl}9/AVERAGE({adj}!{prev_cl}73,{adj}!{cl}73))",
                     fmt=FMT_RATIO)

    # --- DuPont: ROA = ROS x Asset Turnover ---
    r = 9
    _label(r, "  = ROA", indent=2)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci)
        else:
            _formula(r, ci, f"={cl}7*{cl}8")

    # --- DuPont: Financial Leverage ---
    r = 10
    _label(r, "  x Financial Leverage", indent=2)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}81,{adj}!{cl}81)=0,0,"
                     f"AVERAGE({adj}!{prev_cl}73,{adj}!{cl}73)/AVERAGE({adj}!{prev_cl}81,{adj}!{cl}81))",
                     fmt=FMT_RATIO)

    # --- DuPont: ROE check ---
    r = 11
    _label(r, "  = ROE check (ROA x Leverage)", indent=2)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci)
        else:
            _formula(r, ci, f"={cl}9*{cl}10")

    # --- RNOA ---
    r = 13
    _label(r, "RNOA", bold=True)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}67,{adj}!{cl}67)=0,0,"
                     f"{adj}!{cl}36/AVERAGE({adj}!{prev_cl}67,{adj}!{cl}67))")

    # --- NOPAT Margin (for RNOA decomposition) ---
    r = 14
    _label(r, "  NOPAT Margin", indent=2)
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}36/{adj}!{cl}9)")

    # --- NOA Turnover ---
    r = 15
    _label(r, "  x NOA Turnover", indent=2)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}67,{adj}!{cl}67)=0,0,"
                     f"{adj}!{cl}9/AVERAGE({adj}!{prev_cl}67,{adj}!{cl}67))",
                     fmt=FMT_RATIO)

    # --- Gross Profit Margin ---
    r = 17
    _label(r, "Gross Profit Margin")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}11/{adj}!{cl}9)")

    # --- EBIT Margin ---
    r = 18
    _label(r, "EBIT Margin")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}25/{adj}!{cl}9)")

    # --- NOPAT Margin ---
    r = 19
    _label(r, "NOPAT Margin")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}36/{adj}!{cl}9)")

    # =========================================================================
    # Section: COMMON-SIZE INCOME STATEMENT
    # =========================================================================
    r = 21
    style_section_header(ws, r, MAX_COL, "COMMON-SIZE INCOME STATEMENT")

    r = 22
    _label(r, "Revenue")
    for ci in range(2, 8):
        _formula(r, ci, "=1", fmt=FMT_PERCENT)

    r = 23
    _label(r, "Cost of Sales / Revenue")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}10/{adj}!{cl}9)")

    r = 24
    _label(r, "SG&A / Revenue")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}12/{adj}!{cl}9)")

    r = 25
    _label(r, "R&D / Revenue")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}13/{adj}!{cl}9)")

    # =========================================================================
    # Section: ASSET MANAGEMENT
    # =========================================================================
    r = 27
    style_section_header(ws, r, MAX_COL, "ASSET MANAGEMENT")

    # --- Receivables Turnover ---
    r = 28
    _label(r, "Receivables Turnover")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}45,{adj}!{cl}45)=0,0,"
                     f"{adj}!{cl}9/AVERAGE({adj}!{prev_cl}45,{adj}!{cl}45))",
                     fmt=FMT_RATIO)

    # --- Days Receivable ---
    r = 29
    _label(r, "Days Receivable")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci, f"=IF({cl}28=0,0,365/{cl}28)", fmt=FMT_RATIO)

    # --- Inventory Turnover ---
    r = 30
    _label(r, "Inventory Turnover")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}47,{adj}!{cl}47)=0,0,"
                     f"{adj}!{cl}10/AVERAGE({adj}!{prev_cl}47,{adj}!{cl}47))",
                     fmt=FMT_RATIO)

    # --- Days Inventory ---
    r = 31
    _label(r, "Days Inventory")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci, f"=IF({cl}30=0,0,365/{cl}30)", fmt=FMT_RATIO)

    # --- Payables Turnover ---
    r = 32
    _label(r, "Payables Turnover")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci,
                     f"=IF(AVERAGE({adj}!{prev_cl}49,{adj}!{cl}49)=0,0,"
                     f"{adj}!{cl}10/AVERAGE({adj}!{prev_cl}49,{adj}!{cl}49))",
                     fmt=FMT_RATIO)

    # --- Days Payable ---
    r = 33
    _label(r, "Days Payable")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci, f"=IF({cl}32=0,0,365/{cl}32)", fmt=FMT_RATIO)

    # --- Cash Conversion Cycle ---
    r = 34
    _label(r, "Cash Conversion Cycle", bold=True)
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci, fmt=FMT_RATIO)
        else:
            _formula(r, ci, f"={cl}31+{cl}29-{cl}33", fmt=FMT_RATIO)

    # --- OWC / Revenue ---
    r = 36
    _label(r, "OWC / Revenue")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}54/{adj}!{cl}9)")

    # --- PP&E / Revenue ---
    r = 37
    _label(r, "PP&E / Revenue")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}9=0,0,{adj}!{cl}57/{adj}!{cl}9)")

    # =========================================================================
    # Section: LIQUIDITY & SOLVENCY
    # =========================================================================
    r = 39
    style_section_header(ws, r, MAX_COL, "LIQUIDITY & SOLVENCY")

    # --- Cash / Total Debt ---
    r = 40
    _label(r, "Cash / Total Debt")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}78=0,0,{adj}!{cl}79/{adj}!{cl}78)",
                 fmt=FMT_RATIO)

    # --- Debt-to-Equity ---
    r = 41
    _label(r, "Debt-to-Equity")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF({adj}!{cl}81=0,0,{adj}!{cl}78/{adj}!{cl}81)",
                 fmt=FMT_RATIO)

    # --- Interest Coverage ---
    r = 42
    _label(r, "Interest Coverage")
    for ci, cl, _ in _cols():
        _formula(r, ci,
                 f"=IF(ABS({adj}!{cl}29)=0,0,{adj}!{cl}25/ABS({adj}!{cl}29))",
                 fmt=FMT_RATIO)

    # =========================================================================
    # Section: GROWTH
    # =========================================================================
    r = 44
    style_section_header(ws, r, MAX_COL, "GROWTH")

    # --- Revenue Growth ---
    r = 45
    _label(r, "Revenue Growth")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci)
        else:
            _formula(r, ci,
                     f"=IF({adj}!{prev_cl}9=0,0,{adj}!{cl}9/{adj}!{prev_cl}9-1)")

    # --- Sustainable Growth ---
    r = 46
    _label(r, "Sustainable Growth")
    for ci, cl, prev_cl in _cols():
        if prev_cl is None:
            _na(r, ci)
        else:
            # Payout = DPS * shares / NI
            # Sust Growth = (1 - payout) * ROE
            # Guard: if NI=0, show 0
            _formula(r, ci,
                     f"=IF({adj}!{cl}33=0,0,"
                     f"(1-{adj}!{cl}40*{adj}!{cl}38/1000/{adj}!{cl}33)*{cl}6)")

    # Freeze panes at B5
    freeze_panes(ws, row=5, col=2)

    return ws


# ---------------------------------------------------------------------------
# Ratio Comparison tab (peer side-by-side)
# ---------------------------------------------------------------------------
def build_ratio_comparison(wb):
    """Build the 'Ratio Comparison' tab pulling key ratios from all 4 ratio tabs."""
    ws = wb.create_sheet(title="Ratio Comparison")
    MAX_COL = 7  # A-G

    # Column widths
    widths = {1: 40}
    for c in range(2, 8):
        widths[c] = 14
    set_column_widths(ws, widths=widths)

    # Title
    ws.cell(row=1, column=1,
            value="Ratio Comparison - Peer Analysis").font = Font(
        size=14, bold=True, color="000000")
    ws.cell(row=2, column=1,
            value="(key ratios side-by-side)").font = Font(
        size=10, italic=True, color="666666")

    # Year headers in row 3
    write_year_headers(ws, 3, 2, YEARS)

    # Ratio tabs and their display names
    ratio_tabs = [
        ("Ratios ASM", "ASM"),
        ("Ratios AIXTRON", "AIXTRON"),
        ("Ratios AMAT", "AMAT"),
        ("Ratios LRCX", "LRCX"),
    ]

    # Key ratios to compare: (label, source_row, format)
    key_ratios = [
        ("ROE", 6, FMT_PERCENT),
        ("RNOA", 13, FMT_PERCENT),
        ("NOPAT Margin", 19, FMT_PERCENT),
        ("Gross Profit Margin", 17, FMT_PERCENT),
        ("Revenue Growth", 45, FMT_PERCENT),
        ("OWC / Revenue", 36, FMT_PERCENT),
        ("Days Receivable", 29, FMT_RATIO),
        ("Debt-to-Equity", 41, FMT_RATIO),
        ("Interest Coverage", 42, FMT_RATIO),
    ]

    r = 5
    for ratio_label, src_row, fmt in key_ratios:
        # Section label for this ratio
        style_section_header(ws, r, MAX_COL, ratio_label)
        r += 1

        for tab_name, company_label in ratio_tabs:
            write_label(ws, r, 1, f"  {company_label}", indent=2)
            for ci in range(2, 8):
                cl = get_column_letter(ci)
                cell = ws.cell(row=r, column=ci)
                cell.value = f"='{tab_name}'!{cl}{src_row}"
                style_crossref_cell(cell, fmt)
            r += 1

        r += 1  # blank row between ratio groups

    # Freeze panes at B5
    freeze_panes(ws, row=5, col=2)

    return ws


# ---------------------------------------------------------------------------
# Forecast tab
# ---------------------------------------------------------------------------
def build_forecast_tab(wb):
    """Build the 'Forecast' sheet with assumptions, condensed IS, BS, and FCF.

    Historical columns (B-D) reference 'Adj ASM' rows E-G (2022-2024).
    Forecast columns (E-J) use formulas referencing the assumption rows.
    """
    ws = wb.create_sheet(title="Forecast")
    MAX_COL = 10  # columns A-J

    # Column widths: A=45, B-J=14
    widths = {1: 45}
    for c in range(2, MAX_COL + 1):
        widths[c] = 14
    set_column_widths(ws, widths=widths)

    # Adj ASM source sheet (quoted for formulas)
    ADJ = "'Adj ASM'"

    # Historical column mapping: Forecast col (2=B, 3=C, 4=D) -> Adj ASM col letter
    HIST_COL_MAP = {2: 'E', 3: 'F', 4: 'G'}

    # Adj ASM row references
    ADJ_REV = 9
    ADJ_COGS = 10
    ADJ_GP = 11
    ADJ_SGA = 12
    ADJ_RD = 13
    ADJ_DA = 14
    ADJ_EBIT = 25
    ADJ_INV_INC = 27
    ADJ_INT_EXP = 29
    ADJ_PBT = 31
    ADJ_TAX = 32
    ADJ_NI = 33
    ADJ_ETR = 35
    ADJ_NOPAT = 36
    ADJ_OWC = 54
    ADJ_NET_NCA = 65
    ADJ_NOA = 67
    ADJ_INVEST_ASSETS = 71
    ADJ_BIZ_ASSETS = 73
    ADJ_TOTAL_DEBT = 78
    ADJ_CASH = 79
    ADJ_NET_DEBT = 80
    ADJ_EQUITY = 81
    ADJ_INVESTED_CAP = 82
    ADJ_FCF_DE = 96
    ADJ_FCF_EQ = 99

    # Year headers
    hist_years = [2022, 2023, 2024]
    fcast_years = ["2025E", "2026E", "2027E", "2028E", "2029E", "2030E"]
    all_years = hist_years + fcast_years

    # Helper: get column letter for data columns (col 2 = B, ... col 10 = J)
    def cl(col_idx):
        return get_column_letter(col_idx)

    # Helper: write historical cross-ref row (cols B-D referencing Adj ASM)
    def write_hist_ref(row, adj_row, fmt=FMT_CURRENCY):
        for ci in range(2, 5):  # cols B, C, D
            cell = ws.cell(row=row, column=ci)
            adj_col = HIST_COL_MAP[ci]
            cell.value = f"={ADJ}!{adj_col}{adj_row}"
            style_crossref_cell(cell, fmt)

    # Helper: write forecast formula across cols E-J
    def write_fcast_formula(row, formula_fn, fmt=FMT_CURRENCY, bold=False):
        """formula_fn(col_letter, col_idx) -> formula string"""
        for ci in range(5, 11):  # cols E through J
            c = cl(ci)
            cell = ws.cell(row=row, column=ci)
            cell.value = formula_fn(c, ci)
            style_formula_cell(cell, fmt)
            if bold:
                cell.font = BLACK_BOLD

    # Helper: write full-row formula (hist derived + forecast)
    def write_hist_formula(row, formula_fn, fmt=FMT_CURRENCY, bold=False):
        """Write formula across historical cols B-D."""
        for ci in range(2, 5):
            c = cl(ci)
            cell = ws.cell(row=row, column=ci)
            cell.value = formula_fn(c, ci)
            style_formula_cell(cell, fmt)
            if bold:
                cell.font = BLACK_BOLD

    # =========================================================================
    # Row 1-3: Title, subtitle, year headers
    # =========================================================================
    r = 1
    cell = ws.cell(row=r, column=1, value="ASM International NV - Financial Forecast")
    cell.font = Font(size=14, bold=True, color="000000")

    r = 2
    ws.cell(row=r, column=1, value="(EUR 000)").font = Font(size=10, italic=True, color="666666")

    r = 3
    write_year_headers(ws, r, 2, all_years)
    # Also style column A header
    cell = ws.cell(row=r, column=1)
    cell.font = BLACK_BOLD

    r = 4  # blank

    # =========================================================================
    # Section 1: FORECAST ASSUMPTIONS (rows 5-17)
    # =========================================================================
    r = 5
    style_section_header(ws, r, MAX_COL, "FORECAST ASSUMPTIONS")

    # Row 6: Revenue Growth Rate
    r = 6
    write_label(ws, r, 1, "Revenue Growth Rate")
    # Historical: calculated from Adj ASM revenue
    for ci in range(2, 5):
        c = cl(ci)
        adj_col = HIST_COL_MAP[ci]
        # Previous year col in Adj ASM
        prev_adj_col = chr(ord(adj_col) - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{prev_adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_REV}/{ADJ}!{prev_adj_col}{ADJ_REV}-1,0)"
        style_formula_cell(cell, FMT_PERCENT)
    # Forecast assumptions (blue + yellow)
    growth_rates = [0.12, 0.10, 0.08, 0.06, 0.04, 0.03]
    for i, rate in enumerate(growth_rates):
        cell = ws.cell(row=r, column=5 + i, value=rate)
        style_assumption_cell(cell, FMT_PERCENT)
    REV_GROWTH_ROW = r

    # Row 7: Gross Profit Margin
    r = 7
    write_label(ws, r, 1, "Gross Profit Margin")
    # Historical: GP / Revenue from Adj ASM
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_GP}/{ADJ}!{adj_col}{ADJ_REV},0)"
        style_formula_cell(cell, FMT_PERCENT)
    gp_margins = [0.505, 0.505, 0.510, 0.510, 0.510, 0.510]
    for i, val in enumerate(gp_margins):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    GP_MARGIN_ROW = r

    # Row 8: SG&A / Revenue
    r = 8
    write_label(ws, r, 1, "SG&A / Revenue")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_SGA}/{ADJ}!{adj_col}{ADJ_REV},0)"
        style_formula_cell(cell, FMT_PERCENT)
    sga_pcts = [0.108, 0.105, 0.102, 0.100, 0.100, 0.100]
    for i, val in enumerate(sga_pcts):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    SGA_PCT_ROW = r

    # Row 9: R&D / Revenue
    r = 9
    write_label(ws, r, 1, "R&D / Revenue")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_RD}/{ADJ}!{adj_col}{ADJ_REV},0)"
        style_formula_cell(cell, FMT_PERCENT)
    rd_pcts = [0.126, 0.125, 0.123, 0.120, 0.120, 0.120]
    for i, val in enumerate(rd_pcts):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    RD_PCT_ROW = r

    # Row 10: D&A / Revenue
    r = 10
    write_label(ws, r, 1, "D&A / Revenue")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_DA}/{ADJ}!{adj_col}{ADJ_REV},0)"
        style_formula_cell(cell, FMT_PERCENT)
    da_pcts = [0.037, 0.037, 0.037, 0.035, 0.035, 0.035]
    for i, val in enumerate(da_pcts):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    DA_PCT_ROW = r

    # Row 11: Tax Rate
    r = 11
    write_label(ws, r, 1, "Tax Rate")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={ADJ}!{adj_col}{ADJ_ETR}"
        style_formula_cell(cell, FMT_PERCENT)
    tax_rates = [0.21, 0.21, 0.21, 0.21, 0.21, 0.21]
    for i, val in enumerate(tax_rates):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    TAX_RATE_ROW = r

    # Row 12: OWC / Revenue
    r = 12
    write_label(ws, r, 1, "OWC / Revenue")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_OWC}/{ADJ}!{adj_col}{ADJ_REV},0)"
        style_formula_cell(cell, FMT_PERCENT)
    owc_pcts = [0.05, 0.05, 0.05, 0.05, 0.05, 0.05]
    for i, val in enumerate(owc_pcts):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    OWC_PCT_ROW = r

    # Row 13: Net NCA / Revenue
    r = 13
    write_label(ws, r, 1, "Net NCA / Revenue")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_NET_NCA}/{ADJ}!{adj_col}{ADJ_REV},0)"
        style_formula_cell(cell, FMT_PERCENT)
    nca_pcts = [0.40, 0.40, 0.39, 0.38, 0.37, 0.36]
    for i, val in enumerate(nca_pcts):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    NCA_PCT_ROW = r

    # Row 14: Investment Assets / Revenue
    r = 14
    write_label(ws, r, 1, "Investment Assets / Revenue")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_REV}<>0,{ADJ}!{adj_col}{ADJ_INVEST_ASSETS}/{ADJ}!{adj_col}{ADJ_REV},0)"
        style_formula_cell(cell, FMT_PERCENT)
    ia_pcts = [0.30, 0.29, 0.28, 0.27, 0.26, 0.25]
    for i, val in enumerate(ia_pcts):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    IA_PCT_ROW = r

    # Row 15: Debt / Capital
    r = 15
    write_label(ws, r, 1, "Debt / Capital")
    for ci in range(2, 5):
        adj_col = HIST_COL_MAP[ci]
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({ADJ}!{adj_col}{ADJ_INVESTED_CAP}<>0,{ADJ}!{adj_col}{ADJ_TOTAL_DEBT}/{ADJ}!{adj_col}{ADJ_INVESTED_CAP},0)"
        style_formula_cell(cell, FMT_PERCENT)
    debt_pcts = [0.01, 0.01, 0.01, 0.01, 0.01, 0.01]
    for i, val in enumerate(debt_pcts):
        cell = ws.cell(row=r, column=5 + i, value=val)
        style_assumption_cell(cell, FMT_PERCENT)
    DEBT_PCT_ROW = r

    # Row 16: Terminal Growth Rate (only 2030E has a value)
    r = 16
    write_label(ws, r, 1, "Terminal Growth Rate")
    # Leave B-I blank, only J (2030E) gets 2.5%
    cell = ws.cell(row=r, column=10, value=0.025)
    style_assumption_cell(cell, FMT_PERCENT)
    TERM_GROWTH_ROW = r

    r = 17  # blank

    # =========================================================================
    # Section 2: CONDENSED INCOME STATEMENT (rows 18-37)
    # =========================================================================
    r = 18
    style_section_header(ws, r, MAX_COL, "CONDENSED INCOME STATEMENT")

    # Row 19: Revenue
    r = 19
    write_label(ws, r, 1, "Revenue", bold=True)
    write_hist_ref(r, ADJ_REV)
    # Forecast: Revenue = prev_col_Revenue * (1 + growth_rate)
    write_fcast_formula(r,
        lambda c, ci: f"={cl(ci-1)}{r}*(1+{c}{REV_GROWTH_ROW})",
        bold=True)
    REV_ROW = r

    # Row 20: Revenue Growth
    r = 20
    write_label(ws, r, 1, "Revenue Growth", indent=2)
    # Historical: reference growth rate row
    write_hist_formula(r,
        lambda c, ci: f"={c}{REV_GROWTH_ROW}",
        fmt=FMT_PERCENT)
    # Forecast: reference assumption row
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_GROWTH_ROW}",
        fmt=FMT_PERCENT)

    # Row 21: blank
    r = 21

    # Row 22: Gross Profit
    r = 22
    write_label(ws, r, 1, "Gross Profit", bold=True)
    write_hist_ref(r, ADJ_GP)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_ROW}*{c}{GP_MARGIN_ROW}",
        bold=True)
    GP_ROW = r

    # Row 23: Gross Margin
    r = 23
    write_label(ws, r, 1, "Gross Margin", indent=2)
    write_hist_formula(r,
        lambda c, ci: f"=IF({c}{REV_ROW}<>0,{c}{GP_ROW}/{c}{REV_ROW},0)",
        fmt=FMT_PERCENT)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{GP_MARGIN_ROW}",
        fmt=FMT_PERCENT)

    # Row 24: SG&A
    r = 24
    write_label(ws, r, 1, "SG&A", indent=2)
    write_hist_ref(r, ADJ_SGA)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_ROW}*{c}{SGA_PCT_ROW}")
    SGA_ROW = r

    # Row 25: R&D
    r = 25
    write_label(ws, r, 1, "R&D", indent=2)
    write_hist_ref(r, ADJ_RD)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_ROW}*{c}{RD_PCT_ROW}")
    RD_ROW = r

    # Row 26: D&A
    r = 26
    write_label(ws, r, 1, "D&A", indent=2)
    write_hist_ref(r, ADJ_DA)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_ROW}*{c}{DA_PCT_ROW}")
    DA_ROW = r

    # Row 27: EBIT = GP - SGA - RD - DA
    r = 27
    write_label(ws, r, 1, "EBIT", bold=True)
    write_hist_ref(r, ADJ_EBIT)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{GP_ROW}-{c}{SGA_ROW}-{c}{RD_ROW}-{c}{DA_ROW}",
        bold=True)
    style_total_row(ws, r, MAX_COL)
    EBIT_ROW = r

    # Row 28: EBIT Margin
    r = 28
    write_label(ws, r, 1, "EBIT Margin", indent=2)
    # All columns: formula
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({c}{REV_ROW}<>0,{c}{EBIT_ROW}/{c}{REV_ROW},0)"
        style_formula_cell(cell, FMT_PERCENT)

    # Row 29: Tax on EBIT
    r = 29
    write_label(ws, r, 1, "Tax on EBIT", indent=2)
    write_hist_formula(r,
        lambda c, ci: f"={c}{EBIT_ROW}*{c}{TAX_RATE_ROW}")
    write_fcast_formula(r,
        lambda c, ci: f"={c}{EBIT_ROW}*{c}{TAX_RATE_ROW}")
    TAX_EBIT_ROW = r

    # Row 30: NOPAT = EBIT - Tax on EBIT
    r = 30
    write_label(ws, r, 1, "NOPAT", bold=True)
    write_hist_ref(r, ADJ_NOPAT)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{EBIT_ROW}-{c}{TAX_EBIT_ROW}",
        bold=True)
    style_total_row(ws, r, MAX_COL)
    NOPAT_ROW = r

    # Row 31: NOPAT Margin
    r = 31
    write_label(ws, r, 1, "NOPAT Margin", indent=2)
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({c}{REV_ROW}<>0,{c}{NOPAT_ROW}/{c}{REV_ROW},0)"
        style_formula_cell(cell, FMT_PERCENT)

    # Row 32: blank
    r = 32

    # Row 33: Investment Income
    r = 33
    write_label(ws, r, 1, "Investment Income", indent=2)
    write_hist_ref(r, ADJ_INV_INC)
    # Forecast: hold constant at 2024 level (col D = 2024 value)
    write_fcast_formula(r,
        lambda c, ci: f"=D{r}")
    INV_INC_ROW = r

    # Row 34: Interest Expense
    r = 34
    write_label(ws, r, 1, "Interest Expense", indent=2)
    write_hist_ref(r, ADJ_INT_EXP)
    # Forecast: hold constant at 2024 level
    write_fcast_formula(r,
        lambda c, ci: f"=D{r}")
    INT_EXP_ROW = r

    # Row 35: Net Income = NOPAT + Inv Inc - Int Exp
    r = 35
    write_label(ws, r, 1, "Net Income", bold=True)
    write_hist_ref(r, ADJ_NI)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{NOPAT_ROW}+{c}{INV_INC_ROW}+{c}{INT_EXP_ROW}",
        bold=True)
    style_double_line_row(ws, r, MAX_COL)
    NI_ROW = r

    r = 36  # blank

    # =========================================================================
    # Section 3: CONDENSED BALANCE SHEET (rows 37-50)
    # =========================================================================
    r = 37
    style_section_header(ws, r, MAX_COL, "CONDENSED BALANCE SHEET")

    # Row 38: OWC
    r = 38
    write_label(ws, r, 1, "Operating Working Capital")
    write_hist_ref(r, ADJ_OWC)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_ROW}*{c}{OWC_PCT_ROW}")
    F_OWC_ROW = r

    # Row 39: + Net NCA
    r = 39
    write_label(ws, r, 1, "+ Net Non-Current Operating Assets")
    write_hist_ref(r, ADJ_NET_NCA)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_ROW}*{c}{NCA_PCT_ROW}")
    F_NCA_ROW = r

    # Row 40: = NOA
    r = 40
    write_label(ws, r, 1, "= Net Operating Assets", bold=True)
    write_hist_ref(r, ADJ_NOA)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{F_OWC_ROW}+{c}{F_NCA_ROW}",
        bold=True)
    style_total_row(ws, r, MAX_COL)
    F_NOA_ROW = r

    # Row 41: blank
    r = 41

    # Row 42: + Investment Assets
    r = 42
    write_label(ws, r, 1, "+ Investment Assets")
    write_hist_ref(r, ADJ_INVEST_ASSETS)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{REV_ROW}*{c}{IA_PCT_ROW}")
    F_IA_ROW = r

    # Row 43: = Business Assets
    r = 43
    write_label(ws, r, 1, "= Business Assets", bold=True)
    write_hist_ref(r, ADJ_BIZ_ASSETS)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{F_NOA_ROW}+{c}{F_IA_ROW}",
        bold=True)
    style_double_line_row(ws, r, MAX_COL)
    F_BIZ_ASSETS_ROW = r

    # Row 44: blank
    r = 44

    # Row 45: Invested Capital (= Business Assets, by construction)
    r = 45
    write_label(ws, r, 1, "= Invested Capital", bold=True)
    write_hist_ref(r, ADJ_INVESTED_CAP)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{F_BIZ_ASSETS_ROW}",
        bold=True)
    F_INVESTED_CAP_ROW = r

    # Row 46: Debt
    r = 46
    write_label(ws, r, 1, "Debt", indent=2)
    write_hist_ref(r, ADJ_TOTAL_DEBT)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{F_INVESTED_CAP_ROW}*{c}{DEBT_PCT_ROW}")
    F_DEBT_ROW = r

    # Row 47: Group Equity = Invested Capital - Debt
    r = 47
    write_label(ws, r, 1, "Group Equity", indent=2)
    write_hist_ref(r, ADJ_EQUITY)
    write_fcast_formula(r,
        lambda c, ci: f"={c}{F_INVESTED_CAP_ROW}-{c}{F_DEBT_ROW}")
    F_EQUITY_ROW = r

    # Row 48: BS Check (should = 0)
    r = 48
    write_label(ws, r, 1, "BS Check (should = 0)", bold=True)
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_BIZ_ASSETS_ROW}-{c}{F_INVESTED_CAP_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
        cell.font = BLACK_BOLD
    F_BS_CHECK_ROW = r

    r = 49  # blank

    # =========================================================================
    # Section 4: FREE CASH FLOW (rows 50-65)
    # =========================================================================
    r = 50
    style_section_header(ws, r, MAX_COL, "FREE CASH FLOW")

    # Row 51: NOPAT (ref IS section)
    r = 51
    write_label(ws, r, 1, "NOPAT")
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{NOPAT_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    F_CF_NOPAT_ROW = r

    # Row 52: + D&A
    r = 52
    write_label(ws, r, 1, "+ D&A")
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{DA_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    F_CF_DA_ROW = r

    # Row 53: - Change in OWC = OWC(t) - OWC(t-1)
    r = 53
    write_label(ws, r, 1, "- Change in OWC")
    # Col B (2022): use Adj ASM 2021 (col D) as t-1
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{F_OWC_ROW}-{ADJ}!D{ADJ_OWC}"
    style_formula_cell(cell, FMT_CURRENCY)
    # Cols C onwards: OWC(t) - OWC(t-1)
    for ci in range(3, MAX_COL + 1):
        c = cl(ci)
        prev_c = cl(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_OWC_ROW}-{prev_c}{F_OWC_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    F_CHG_OWC_ROW = r

    # Row 54: - Change in Net NCA = NCA(t) - NCA(t-1)
    r = 54
    write_label(ws, r, 1, "- Change in Net NCA")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{F_NCA_ROW}-{ADJ}!D{ADJ_NET_NCA}"
    style_formula_cell(cell, FMT_CURRENCY)
    for ci in range(3, MAX_COL + 1):
        c = cl(ci)
        prev_c = cl(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_NCA_ROW}-{prev_c}{F_NCA_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    F_CHG_NCA_ROW = r

    # Row 55: = Operating CF after investment
    r = 55
    write_label(ws, r, 1, "= Oper CF after investment", bold=True)
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_CF_NOPAT_ROW}+{c}{F_CF_DA_ROW}-{c}{F_CHG_OWC_ROW}-{c}{F_CHG_NCA_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
        cell.font = BLACK_BOLD
    style_total_row(ws, r, MAX_COL)
    F_OPER_CF_ROW = r

    # Row 56: blank
    r = 56

    # Row 57: + Net Inv Profit after tax = Inv Income * (1 - tax)
    r = 57
    write_label(ws, r, 1, "+ Net Inv Profit after tax")
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{INV_INC_ROW}*(1-{c}{TAX_RATE_ROW})"
        style_formula_cell(cell, FMT_CURRENCY)
    F_NET_INV_PROFIT_ROW = r

    # Row 58: - Change in Inv Assets = IA(t) - IA(t-1)
    r = 58
    write_label(ws, r, 1, "- Change in Inv Assets")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{F_IA_ROW}-{ADJ}!D{ADJ_INVEST_ASSETS}"
    style_formula_cell(cell, FMT_CURRENCY)
    for ci in range(3, MAX_COL + 1):
        c = cl(ci)
        prev_c = cl(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_IA_ROW}-{prev_c}{F_IA_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    F_CHG_IA_ROW = r

    # Row 59: = FCF to D&E
    r = 59
    write_label(ws, r, 1, "= FCF to Debt & Equity", bold=True)
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_OPER_CF_ROW}+{c}{F_NET_INV_PROFIT_ROW}-{c}{F_CHG_IA_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
        cell.font = BLACK_BOLD
    style_total_row(ws, r, MAX_COL)
    F_FCF_DE_ROW = r

    # Row 60: - Interest after tax = Int Exp * (1 - tax)
    r = 60
    write_label(ws, r, 1, "- Interest after tax")
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{INT_EXP_ROW}*(1-{c}{TAX_RATE_ROW})"
        style_formula_cell(cell, FMT_CURRENCY)
    F_INT_AFTER_TAX_ROW = r

    # Row 61: + Change in Debt = Debt(t) - Debt(t-1)
    r = 61
    write_label(ws, r, 1, "+ Change in Debt")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{F_DEBT_ROW}-{ADJ}!D{ADJ_TOTAL_DEBT}"
    style_formula_cell(cell, FMT_CURRENCY)
    for ci in range(3, MAX_COL + 1):
        c = cl(ci)
        prev_c = cl(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_DEBT_ROW}-{prev_c}{F_DEBT_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    F_CHG_DEBT_ROW = r

    # Row 62: = FCF to Equity
    r = 62
    write_label(ws, r, 1, "= FCF to Equity", bold=True)
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_FCF_DE_ROW}-{c}{F_INT_AFTER_TAX_ROW}+{c}{F_CHG_DEBT_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
        cell.font = BLACK_BOLD
    style_double_line_row(ws, r, MAX_COL)
    F_FCF_EQ_ROW = r

    r = 63  # blank

    # =========================================================================
    # Section 5: KEY METRICS (rows 64-70)
    # =========================================================================
    r = 64
    style_section_header(ws, r, MAX_COL, "KEY METRICS")

    # Row 65: NOPAT Margin (repeated for convenience)
    r = 65
    write_label(ws, r, 1, "NOPAT Margin")
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF({c}{REV_ROW}<>0,{c}{NOPAT_ROW}/{c}{REV_ROW},0)"
        style_formula_cell(cell, FMT_PERCENT)

    # Row 66: ROE = NI / Avg Equity
    r = 66
    write_label(ws, r, 1, "ROE")
    # Col B (2022): use Adj ASM 2021 equity (col D) for average
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF((B{F_EQUITY_ROW}+{ADJ}!D{ADJ_EQUITY})/2<>0,B{NI_ROW}/((B{F_EQUITY_ROW}+{ADJ}!D{ADJ_EQUITY})/2),0)"
    style_formula_cell(cell, FMT_PERCENT)
    for ci in range(3, MAX_COL + 1):
        c = cl(ci)
        prev_c = cl(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF(({c}{F_EQUITY_ROW}+{prev_c}{F_EQUITY_ROW})/2<>0,{c}{NI_ROW}/(({c}{F_EQUITY_ROW}+{prev_c}{F_EQUITY_ROW})/2),0)"
        style_formula_cell(cell, FMT_PERCENT)

    # Row 67: RNOA = NOPAT / Avg NOA
    r = 67
    write_label(ws, r, 1, "RNOA")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF((B{F_NOA_ROW}+{ADJ}!D{ADJ_NOA})/2<>0,B{NOPAT_ROW}/((B{F_NOA_ROW}+{ADJ}!D{ADJ_NOA})/2),0)"
    style_formula_cell(cell, FMT_PERCENT)
    for ci in range(3, MAX_COL + 1):
        c = cl(ci)
        prev_c = cl(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=IF(({c}{F_NOA_ROW}+{prev_c}{F_NOA_ROW})/2<>0,{c}{NOPAT_ROW}/(({c}{F_NOA_ROW}+{prev_c}{F_NOA_ROW})/2),0)"
        style_formula_cell(cell, FMT_PERCENT)

    # Row 68: FCF to Equity (absolute, for reference)
    r = 68
    write_label(ws, r, 1, "FCF to Equity")
    for ci in range(2, MAX_COL + 1):
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{F_FCF_EQ_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)

    # --- Freeze panes at B4 ---
    freeze_panes(ws, row=4, col=2)

    return ws


# ---------------------------------------------------------------------------
# Valuation tab
# ---------------------------------------------------------------------------
def build_valuation_tab(wb):
    """Build the 'Valuation' sheet with WACC, DCF (FCFF & FCFE), PVAOI, PVAE,
    Sensitivity Tables, Comparable Company Analysis, and Football Field summary.

    All cross-sheet references point to 'Forecast' and 'Adj ASM'.
    """
    ws = wb.create_sheet(title="Valuation")
    MAX_COL = 10  # A-J to match Forecast layout

    # Quoted sheet names for formulas
    FCAST = "'Forecast'"
    ADJ = "'Adj ASM'"

    # Column widths: A=45, B-J=14
    widths = {1: 45}
    for c in range(2, MAX_COL + 1):
        widths[c] = 14
    set_column_widths(ws, widths=widths)

    # Forecast column letters: E=2025E .. J=2030E (columns 5-10)
    FCST_COLS = [5, 6, 7, 8, 9, 10]  # column indices
    FCST_YEAR_LABELS = ["2025E", "2026E", "2027E", "2028E", "2029E", "2030E"]

    def cl(col_idx):
        return get_column_letter(col_idx)

    # =========================================================================
    # Title
    # =========================================================================
    r = 1
    cell = ws.cell(row=r, column=1, value="ASM International NV - Valuation Analysis")
    cell.font = Font(size=14, bold=True, color="000000")

    r = 2
    ws.cell(row=r, column=1, value="(EUR 000)").font = Font(size=10, italic=True, color="666666")

    # =========================================================================
    # Section 1: COST OF CAPITAL (rows 3-22)
    # =========================================================================
    r = 3
    style_section_header(ws, r, MAX_COL, "COST OF CAPITAL")

    r = 4  # blank

    # Row 5: Risk-free Rate
    r = 5
    write_label(ws, r, 1, "Risk-free Rate")
    cell = ws.cell(row=r, column=2, value=0.025)
    style_assumption_cell(cell, FMT_PERCENT)
    RF_ROW = r

    # Row 6: Equity Beta
    r = 6
    write_label(ws, r, 1, "Equity Beta")
    cell = ws.cell(row=r, column=2, value=1.30)
    style_assumption_cell(cell, FMT_RATIO)
    BETA_ROW = r

    # Row 7: Equity Risk Premium
    r = 7
    write_label(ws, r, 1, "Equity Risk Premium")
    cell = ws.cell(row=r, column=2, value=0.05)
    style_assumption_cell(cell, FMT_PERCENT)
    ERP_ROW = r

    # Row 8: Cost of Equity (Re) = CAPM
    r = 8
    write_label(ws, r, 1, "Cost of Equity (Re)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{RF_ROW}+B{BETA_ROW}*B{ERP_ROW}"
    style_formula_cell(cell, FMT_PERCENT)
    RE_ROW = r

    r = 9  # blank

    # Row 10: Cost of Debt (pre-tax) = |Interest Expense| / Total Debt
    r = 10
    write_label(ws, r, 1, "Cost of Debt (pre-tax)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF({ADJ}!G78<>0,ABS({ADJ}!G29)/{ADJ}!G78,0)"
    style_formula_cell(cell, FMT_PERCENT)
    KD_PRE_ROW = r

    # Row 11: Tax Rate (from Forecast)
    r = 11
    write_label(ws, r, 1, "Tax Rate")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={FCAST}!G11"
    style_crossref_cell(cell, FMT_PERCENT)
    TAX_ROW = r

    # Row 12: Cost of Debt (after-tax)
    r = 12
    write_label(ws, r, 1, "Cost of Debt (after-tax)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{KD_PRE_ROW}*(1-B{TAX_ROW})"
    style_formula_cell(cell, FMT_PERCENT)
    KD_POST_ROW = r

    r = 13  # blank

    # Row 14: Market Cap (EUR 000)
    r = 14
    write_label(ws, r, 1, "Market Cap (EUR 000)")
    cell = ws.cell(row=r, column=2, value=25000000)
    style_assumption_cell(cell, FMT_CURRENCY)
    MCAP_ROW = r

    # Row 15: Net Debt
    r = 15
    write_label(ws, r, 1, "Net Debt")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={ADJ}!G80"
    style_crossref_cell(cell, FMT_CURRENCY)
    NET_DEBT_ROW = r

    # Row 16: E/V = MarketCap / (MarketCap + NetDebt)
    r = 16
    write_label(ws, r, 1, "E/V")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF((B{MCAP_ROW}+B{NET_DEBT_ROW})<>0,B{MCAP_ROW}/(B{MCAP_ROW}+B{NET_DEBT_ROW}),0)"
    style_formula_cell(cell, FMT_PERCENT)
    EV_RATIO_ROW = r

    # Row 17: D/V = 1 - E/V
    r = 17
    write_label(ws, r, 1, "D/V")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=1-B{EV_RATIO_ROW}"
    style_formula_cell(cell, FMT_PERCENT)
    DV_RATIO_ROW = r

    # Row 18: WACC
    r = 18
    write_label(ws, r, 1, "WACC", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{EV_RATIO_ROW}*B{RE_ROW}+B{DV_RATIO_ROW}*B{KD_POST_ROW}"
    style_formula_cell(cell, FMT_PERCENT)
    cell.font = BLACK_BOLD
    WACC_ROW = r

    r = 19  # blank

    # Row 20: Shares Outstanding (000)
    r = 20
    write_label(ws, r, 1, "Shares Outstanding (diluted)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={ADJ}!G38"
    style_crossref_cell(cell, FMT_CURRENCY)
    SHARES_ROW = r

    # Row 21: Current Share Price
    r = 21
    write_label(ws, r, 1, "Current Share Price (EUR)")
    cell = ws.cell(row=r, column=2, value=450)
    style_assumption_cell(cell, '#,##0.00')
    PRICE_ROW = r

    # Row 22: Terminal Growth Rate
    r = 22
    write_label(ws, r, 1, "Terminal Growth Rate")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={FCAST}!J16"
    style_crossref_cell(cell, FMT_PERCENT)
    TERM_G_ROW = r

    r = 23  # blank

    # =========================================================================
    # Section 2: DCF - FREE CASH FLOW TO DEBT & EQUITY (rows 25-39)
    # =========================================================================
    r = 25
    style_section_header(ws, r, MAX_COL, "DCF - FREE CASH FLOW TO DEBT & EQUITY")

    r = 26  # blank

    # Row 27: Year headers (E-J only)
    r = 27
    for i, yr in enumerate(FCST_YEAR_LABELS):
        cell = ws.cell(row=r, column=FCST_COLS[i], value=yr)
        cell.font = BLACK_BOLD
        cell.number_format = FMT_YEAR
        cell.alignment = Alignment(horizontal="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Row 28: FCF to D&E (from Forecast row 59)
    r = 28
    write_label(ws, r, 1, "FCF to Debt & Equity")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={FCAST}!{c}59"
        style_crossref_cell(cell, FMT_CURRENCY)
    FCF_DE_ROW = r

    # Row 29: Terminal Value (only in last col = J = column 10)
    r = 29
    write_label(ws, r, 1, "Terminal Value")
    cell = ws.cell(row=r, column=10)
    cell.value = f"=J{FCF_DE_ROW}*(1+$B${TERM_G_ROW})/($B${WACC_ROW}-$B${TERM_G_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    TV_DE_ROW = r

    # Row 30: Total CF
    r = 30
    write_label(ws, r, 1, "Total Cash Flow")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        if ci < 10:
            cell.value = f"={c}{FCF_DE_ROW}"
        else:
            cell.value = f"={c}{FCF_DE_ROW}+{c}{TV_DE_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    TOTAL_CF_DE_ROW = r

    # Row 31: Discount Factor = 1/(1+WACC)^n
    r = 31
    write_label(ws, r, 1, "Discount Factor")
    for i, ci in enumerate(FCST_COLS):
        n = i + 1
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=1/(1+$B${WACC_ROW})^{n}"
        style_formula_cell(cell, '0.0000')
    DF_DE_ROW = r

    # Row 32: PV of CF
    r = 32
    write_label(ws, r, 1, "PV of Cash Flow")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{TOTAL_CF_DE_ROW}*{c}{DF_DE_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    PV_CF_DE_ROW = r

    r = 33  # blank

    # Row 34: Sum of PV
    r = 34
    write_label(ws, r, 1, "Sum of PV of Cash Flows", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=SUM(E{PV_CF_DE_ROW}:J{PV_CF_DE_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    SUM_PV_DE_ROW = r

    # Row 35: Enterprise Value
    r = 35
    write_label(ws, r, 1, "Enterprise Value")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{SUM_PV_DE_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)
    EV_DE_ROW = r

    # Row 36: - Net Debt
    r = 36
    write_label(ws, r, 1, "- Net Debt")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{NET_DEBT_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # Row 37: = Equity Value
    r = 37
    write_label(ws, r, 1, "= Equity Value", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{EV_DE_ROW}-B36"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    EQ_VAL_DE_ROW = r

    # Row 38: / Shares
    r = 38
    write_label(ws, r, 1, "/ Shares Outstanding")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{SHARES_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # Row 39: Implied Price (DCF FCFF)
    r = 39
    write_label(ws, r, 1, "Implied Price (DCF FCFF)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF(B38<>0,B{EQ_VAL_DE_ROW}/B38*1000,0)"
    style_formula_cell(cell, '#,##0.00')
    cell.font = BLACK_BOLD
    style_double_line_row(ws, r, MAX_COL)
    PRICE_DCF_FCFF_ROW = r

    r = 40  # blank

    # =========================================================================
    # Section 3: DCF - FREE CASH FLOW TO EQUITY (rows 42-55)
    # =========================================================================
    r = 42
    style_section_header(ws, r, MAX_COL, "DCF - FREE CASH FLOW TO EQUITY")

    r = 43  # blank

    # Row 44: Year headers
    r = 44
    for i, yr in enumerate(FCST_YEAR_LABELS):
        cell = ws.cell(row=r, column=FCST_COLS[i], value=yr)
        cell.font = BLACK_BOLD
        cell.number_format = FMT_YEAR
        cell.alignment = Alignment(horizontal="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Row 45: FCFE (from Forecast row 62)
    r = 45
    write_label(ws, r, 1, "FCF to Equity")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={FCAST}!{c}62"
        style_crossref_cell(cell, FMT_CURRENCY)
    FCFE_ROW = r

    # Row 46: Terminal Value (FCFE)
    r = 46
    write_label(ws, r, 1, "Terminal Value")
    cell = ws.cell(row=r, column=10)
    cell.value = f"=J{FCFE_ROW}*(1+$B${TERM_G_ROW})/($B${RE_ROW}-$B${TERM_G_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    TV_FCFE_ROW = r

    # Row 47: Total CF
    r = 47
    write_label(ws, r, 1, "Total Cash Flow")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        if ci < 10:
            cell.value = f"={c}{FCFE_ROW}"
        else:
            cell.value = f"={c}{FCFE_ROW}+{c}{TV_FCFE_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    TOTAL_CF_FCFE_ROW = r

    # Row 48: Discount Factor (at Re)
    r = 48
    write_label(ws, r, 1, "Discount Factor")
    for i, ci in enumerate(FCST_COLS):
        n = i + 1
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=1/(1+$B${RE_ROW})^{n}"
        style_formula_cell(cell, '0.0000')
    DF_FCFE_ROW = r

    # Row 49: PV of CF
    r = 49
    write_label(ws, r, 1, "PV of Cash Flow")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{TOTAL_CF_FCFE_ROW}*{c}{DF_FCFE_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    PV_CF_FCFE_ROW = r

    r = 50  # blank

    # Row 51: Sum of PV
    r = 51
    write_label(ws, r, 1, "Equity Value (Sum of PV)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=SUM(E{PV_CF_FCFE_ROW}:J{PV_CF_FCFE_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    EQ_VAL_FCFE_ROW = r

    # Row 52: / Shares
    r = 52
    write_label(ws, r, 1, "/ Shares Outstanding")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{SHARES_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # Row 53: Implied Price (DCF FCFE)
    r = 53
    write_label(ws, r, 1, "Implied Price (DCF FCFE)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF(B52<>0,B{EQ_VAL_FCFE_ROW}/B52*1000,0)"
    style_formula_cell(cell, '#,##0.00')
    cell.font = BLACK_BOLD
    style_double_line_row(ws, r, MAX_COL)
    PRICE_DCF_FCFE_ROW = r

    r = 54  # blank

    # =========================================================================
    # Section 4: PV OF ABNORMAL NOPAT (PVAOI) (rows 56-76)
    # =========================================================================
    r = 56
    style_section_header(ws, r, MAX_COL, "PV OF ABNORMAL OPERATING INCOME (PVAOI)")

    r = 57  # blank

    # Row 58: Year headers
    r = 58
    for i, yr in enumerate(FCST_YEAR_LABELS):
        cell = ws.cell(row=r, column=FCST_COLS[i], value=yr)
        cell.font = BLACK_BOLD
        cell.number_format = FMT_YEAR
        cell.alignment = Alignment(horizontal="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Row 59: NOPAT (from Forecast row 30)
    r = 59
    write_label(ws, r, 1, "NOPAT")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={FCAST}!{c}30"
        style_crossref_cell(cell, FMT_CURRENCY)
    PVAOI_NOPAT_ROW = r

    # Row 60: Beginning BV of NOA (prior year NOA from Forecast row 40)
    r = 60
    write_label(ws, r, 1, "Beg BV of NOA")
    # 2025E uses 2024 (Forecast col D=row 40), 2026E uses E40, etc.
    for i, ci in enumerate(FCST_COLS):
        prev_c = cl(ci - 1)  # D for 2025E, E for 2026E, ...
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={FCAST}!{prev_c}40"
        style_crossref_cell(cell, FMT_CURRENCY)
    BEG_NOA_ROW = r

    # Row 61: Normal Earnings = WACC * Beg NOA
    r = 61
    write_label(ws, r, 1, "Normal Earnings (WACC x Beg NOA)")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=$B${WACC_ROW}*{c}{BEG_NOA_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    NORMAL_NOPAT_ROW = r

    # Row 62: Abnormal NOPAT
    r = 62
    write_label(ws, r, 1, "Abnormal NOPAT")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{PVAOI_NOPAT_ROW}-{c}{NORMAL_NOPAT_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    ABN_NOPAT_ROW = r

    # Row 63: Terminal Value (last col only)
    r = 63
    write_label(ws, r, 1, "Terminal Value")
    cell = ws.cell(row=r, column=10)
    cell.value = f"=J{ABN_NOPAT_ROW}*(1+$B${TERM_G_ROW})/($B${WACC_ROW}-$B${TERM_G_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    TV_PVAOI_ROW = r

    # Row 64: PV Factor
    r = 64
    write_label(ws, r, 1, "Discount Factor")
    for i, ci in enumerate(FCST_COLS):
        n = i + 1
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=1/(1+$B${WACC_ROW})^{n}"
        style_formula_cell(cell, '0.0000')
    DF_PVAOI_ROW = r

    # Row 65: PV of Abnormal NOPAT
    r = 65
    write_label(ws, r, 1, "PV of Abnormal NOPAT")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        if ci < 10:
            cell.value = f"={c}{ABN_NOPAT_ROW}*{c}{DF_PVAOI_ROW}"
        else:
            # Last year: include terminal value
            cell.value = f"=({c}{ABN_NOPAT_ROW}+{c}{TV_PVAOI_ROW})*{c}{DF_PVAOI_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    PV_ABN_NOPAT_ROW = r

    r = 66  # blank

    # Row 67: Sum PV Abnormal
    r = 67
    write_label(ws, r, 1, "Sum PV of Abnormal NOPAT", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=SUM(E{PV_ABN_NOPAT_ROW}:J{PV_ABN_NOPAT_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    SUM_PV_ABN_NOPAT_ROW = r

    # Row 68: + Beg BV of NOA (2024)
    r = 68
    write_label(ws, r, 1, "+ Beg BV of NOA (2024)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={FCAST}!D40"
    style_crossref_cell(cell, FMT_CURRENCY)
    BEG_NOA_2024_ROW = r

    # Row 69: = Enterprise Value
    r = 69
    write_label(ws, r, 1, "= Enterprise Value", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{SUM_PV_ABN_NOPAT_ROW}+B{BEG_NOA_2024_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    EV_PVAOI_ROW = r

    # Row 70: - Net Debt
    r = 70
    write_label(ws, r, 1, "- Net Debt")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{NET_DEBT_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # Row 71: = Equity Value
    r = 71
    write_label(ws, r, 1, "= Equity Value", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{EV_PVAOI_ROW}-B70"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    EQ_VAL_PVAOI_ROW = r

    # Row 72: / Shares
    r = 72
    write_label(ws, r, 1, "/ Shares Outstanding")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{SHARES_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # Row 73: Implied Price (PVAOI)
    r = 73
    write_label(ws, r, 1, "Implied Price (PVAOI)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF(B72<>0,B{EQ_VAL_PVAOI_ROW}/B72*1000,0)"
    style_formula_cell(cell, '#,##0.00')
    cell.font = BLACK_BOLD
    style_double_line_row(ws, r, MAX_COL)
    PRICE_PVAOI_ROW = r

    r = 74  # blank

    # =========================================================================
    # Section 5: PV OF ABNORMAL EARNINGS (PVAE) (rows 76-96)
    # =========================================================================
    r = 76
    style_section_header(ws, r, MAX_COL, "PV OF ABNORMAL EARNINGS (PVAE)")

    r = 77  # blank

    # Row 78: Year headers
    r = 78
    for i, yr in enumerate(FCST_YEAR_LABELS):
        cell = ws.cell(row=r, column=FCST_COLS[i], value=yr)
        cell.font = BLACK_BOLD
        cell.number_format = FMT_YEAR
        cell.alignment = Alignment(horizontal="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Row 79: Net Income (from Forecast row 35)
    r = 79
    write_label(ws, r, 1, "Net Income")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={FCAST}!{c}35"
        style_crossref_cell(cell, FMT_CURRENCY)
    PVAE_NI_ROW = r

    # Row 80: Beg BV of Equity (prior year equity from Forecast row 47)
    r = 80
    write_label(ws, r, 1, "Beg BV of Equity")
    for i, ci in enumerate(FCST_COLS):
        prev_c = cl(ci - 1)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={FCAST}!{prev_c}47"
        style_crossref_cell(cell, FMT_CURRENCY)
    BEG_EQ_ROW = r

    # Row 81: Normal Earnings = Re * Beg Equity
    r = 81
    write_label(ws, r, 1, "Normal Earnings (Re x Beg Equity)")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=$B${RE_ROW}*{c}{BEG_EQ_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    NORMAL_EARN_ROW = r

    # Row 82: Abnormal Earnings
    r = 82
    write_label(ws, r, 1, "Abnormal Earnings")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        cell.value = f"={c}{PVAE_NI_ROW}-{c}{NORMAL_EARN_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    ABN_EARN_ROW = r

    # Row 83: Terminal Value
    r = 83
    write_label(ws, r, 1, "Terminal Value")
    cell = ws.cell(row=r, column=10)
    cell.value = f"=J{ABN_EARN_ROW}*(1+$B${TERM_G_ROW})/($B${RE_ROW}-$B${TERM_G_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    TV_PVAE_ROW = r

    # Row 84: PV Factor (at Re)
    r = 84
    write_label(ws, r, 1, "Discount Factor")
    for i, ci in enumerate(FCST_COLS):
        n = i + 1
        cell = ws.cell(row=r, column=ci)
        cell.value = f"=1/(1+$B${RE_ROW})^{n}"
        style_formula_cell(cell, '0.0000')
    DF_PVAE_ROW = r

    # Row 85: PV of Abnormal Earnings
    r = 85
    write_label(ws, r, 1, "PV of Abnormal Earnings")
    for ci in FCST_COLS:
        c = cl(ci)
        cell = ws.cell(row=r, column=ci)
        if ci < 10:
            cell.value = f"={c}{ABN_EARN_ROW}*{c}{DF_PVAE_ROW}"
        else:
            cell.value = f"=({c}{ABN_EARN_ROW}+{c}{TV_PVAE_ROW})*{c}{DF_PVAE_ROW}"
        style_formula_cell(cell, FMT_CURRENCY)
    PV_ABN_EARN_ROW = r

    r = 86  # blank

    # Row 87: Sum PV Abnormal Earnings
    r = 87
    write_label(ws, r, 1, "Sum PV of Abnormal Earnings", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=SUM(E{PV_ABN_EARN_ROW}:J{PV_ABN_EARN_ROW})"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    SUM_PV_ABN_EARN_ROW = r

    # Row 88: + Current BV Equity (2024)
    r = 88
    write_label(ws, r, 1, "+ Current BV Equity (2024)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={ADJ}!G81"
    style_crossref_cell(cell, FMT_CURRENCY)
    BV_EQ_2024_ROW = r

    # Row 89: = Equity Value
    r = 89
    write_label(ws, r, 1, "= Equity Value", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{SUM_PV_ABN_EARN_ROW}+B{BV_EQ_2024_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    EQ_VAL_PVAE_ROW = r

    # Row 90: / Shares
    r = 90
    write_label(ws, r, 1, "/ Shares Outstanding")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{SHARES_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # Row 91: Implied Price (PVAE)
    r = 91
    write_label(ws, r, 1, "Implied Price (PVAE)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF(B90<>0,B{EQ_VAL_PVAE_ROW}/B90*1000,0)"
    style_formula_cell(cell, '#,##0.00')
    cell.font = BLACK_BOLD
    style_double_line_row(ws, r, MAX_COL)
    PRICE_PVAE_ROW = r

    r = 92  # blank

    # =========================================================================
    # Section 6: SENSITIVITY TABLES (rows 94-135)
    # =========================================================================
    r = 94
    style_section_header(ws, r, MAX_COL, "SENSITIVITY ANALYSIS")

    # --- DCF FCFF Sensitivity Table ---
    r = 96
    write_label(ws, r, 1, "DCF (FCFF) - Implied Share Price", bold=True)

    # WACC row values (7% to 12% in 0.5% steps) - 11 values
    wacc_vals = [0.070, 0.075, 0.080, 0.085, 0.090, 0.095, 0.100, 0.105, 0.110, 0.115, 0.120]
    # Terminal growth col values (1.0% to 4.0% in 0.5% steps) - 7 values
    tg_vals = [0.010, 0.015, 0.020, 0.025, 0.030, 0.035, 0.040]

    # Write corner label
    r = 97
    write_label(ws, r, 1, "WACC \\ Terminal Growth", bold=True)
    cell = ws.cell(row=r, column=1)
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

    # Write terminal growth headers (columns B-H = cols 2-8)
    for j, tg in enumerate(tg_vals):
        cell = ws.cell(row=r, column=2 + j, value=tg)
        style_assumption_cell(cell, FMT_PERCENT)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Sensitivity grid rows
    SENS1_START_ROW = 98
    for i, wacc in enumerate(wacc_vals):
        row = SENS1_START_ROW + i
        # WACC label in col A
        cell = ws.cell(row=row, column=1, value=wacc)
        style_assumption_cell(cell, FMT_PERCENT)
        cell.border = THIN_BORDER

        for j, tg in enumerate(tg_vals):
            # Col for this tg value
            col = 2 + j
            tg_cell_ref = f"{cl(col)}{97}"  # terminal growth from header row
            wacc_cell_ref = f"$A${row}"       # WACC from row label

            # Formula: sum of PV of FCFs discounted at this WACC + TV at this WACC/g - NetDebt, / Shares * 1000
            # FCF refs: Forecast E59:J59
            # = (FCF_2025/(1+wacc)^1 + ... + FCF_2030/(1+wacc)^6 + FCF_2030*(1+g)/(wacc-g)/(1+wacc)^6 - NetDebt) / Shares * 1000
            parts = []
            for k in range(6):
                fcf_col = cl(5 + k)  # E through J
                yr_n = k + 1
                parts.append(f"{FCAST}!{fcf_col}59/(1+{wacc_cell_ref})^{yr_n}")
            # Terminal value PV
            parts.append(f"{FCAST}!J59*(1+{tg_cell_ref})/({wacc_cell_ref}-{tg_cell_ref})/(1+{wacc_cell_ref})^6")

            formula = f"=({'+'.join(parts)}-$B${NET_DEBT_ROW})/$B${SHARES_ROW}*1000"
            cell = ws.cell(row=row, column=col, value=formula)
            style_formula_cell(cell, '#,##0.00')

    SENS1_END_ROW = SENS1_START_ROW + len(wacc_vals) - 1

    r = SENS1_END_ROW + 2  # blank row then next table

    # --- PVAOI Sensitivity Table ---
    r = SENS1_END_ROW + 3
    write_label(ws, r, 1, "PVAOI - Implied Share Price", bold=True)
    PVAOI_SENS_TITLE_ROW = r

    r = PVAOI_SENS_TITLE_ROW + 1
    write_label(ws, r, 1, "WACC \\ Terminal Growth", bold=True)
    cell = ws.cell(row=r, column=1)
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    PVAOI_HEADER_ROW = r

    # Write terminal growth headers
    for j, tg in enumerate(tg_vals):
        cell = ws.cell(row=r, column=2 + j, value=tg)
        style_assumption_cell(cell, FMT_PERCENT)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # PVAOI Sensitivity grid
    SENS2_START_ROW = PVAOI_HEADER_ROW + 1
    for i, wacc in enumerate(wacc_vals):
        row = SENS2_START_ROW + i
        cell = ws.cell(row=row, column=1, value=wacc)
        style_assumption_cell(cell, FMT_PERCENT)
        cell.border = THIN_BORDER

        for j, tg in enumerate(tg_vals):
            col = 2 + j
            tg_cell_ref = f"{cl(col)}{PVAOI_HEADER_ROW}"
            wacc_cell_ref = f"$A${row}"

            # PVAOI = sum of PV of abnormal NOPAT + TV + Beg NOA - NetDebt, / Shares * 1000
            # Abnormal NOPAT = NOPAT - WACC * BegNOA for each year
            # BegNOA: Forecast D40 for 2025, E40 for 2026, ... I40 for 2030
            # NOPAT: Forecast E30:J30
            parts_abn = []
            for k in range(6):
                nopat_col = cl(5 + k)       # E through J
                beg_noa_col = cl(4 + k)     # D through I (prior year)
                yr_n = k + 1
                # PV of abnormal NOPAT for this year
                abn_expr = f"({FCAST}!{nopat_col}30-{wacc_cell_ref}*{FCAST}!{beg_noa_col}40)"
                parts_abn.append(f"{abn_expr}/(1+{wacc_cell_ref})^{yr_n}")

            # Terminal value of abnormal NOPAT (based on last year's abnormal)
            last_abn = f"({FCAST}!J30-{wacc_cell_ref}*{FCAST}!I40)"
            tv_pv = f"{last_abn}*(1+{tg_cell_ref})/({wacc_cell_ref}-{tg_cell_ref})/(1+{wacc_cell_ref})^6"

            # Beg NOA (2024) = Forecast D40
            formula = f"=({'+'.join(parts_abn)}+{tv_pv}+{FCAST}!D40-$B${NET_DEBT_ROW})/$B${SHARES_ROW}*1000"
            cell = ws.cell(row=row, column=col, value=formula)
            style_formula_cell(cell, '#,##0.00')

    SENS2_END_ROW = SENS2_START_ROW + len(wacc_vals) - 1

    r = SENS2_END_ROW + 2  # blank

    # =========================================================================
    # Section 7: COMPARABLE COMPANY ANALYSIS (rows ~SENS2_END+3)
    # =========================================================================
    COMPS_START = SENS2_END_ROW + 3
    r = COMPS_START
    style_section_header(ws, r, MAX_COL, "COMPARABLE COMPANY ANALYSIS")

    r = COMPS_START + 2
    # Headers: col A = label, cols B, C, D = AIXTRON, AMAT, LRCX
    write_label(ws, r, 1, "Peer Company", bold=True)
    cell = ws.cell(row=r, column=1)
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

    peers = ["AIXTRON SE", "Applied Materials", "Lam Research"]
    for j, peer in enumerate(peers):
        cell = ws.cell(row=r, column=2 + j, value=peer)
        cell.font = BLACK_BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    COMPS_HDR_ROW = r

    # Peer EV (EUR 000) - hardcoded blue inputs (approximate values)
    r = COMPS_HDR_ROW + 1
    write_label(ws, r, 1, "Enterprise Value (EUR 000)")
    peer_evs = [3500000, 150000000, 95000000]  # ~EUR 3.5bn, 150bn, 95bn
    for j, ev in enumerate(peer_evs):
        cell = ws.cell(row=r, column=2 + j, value=ev)
        style_assumption_cell(cell, FMT_CURRENCY)
    PEER_EV_ROW = r

    # Peer EBITDA
    r = COMPS_HDR_ROW + 2
    write_label(ws, r, 1, "EBITDA (EUR 000)")
    peer_ebitda = [120000, 9500000, 5800000]  # approximate
    for j, val in enumerate(peer_ebitda):
        cell = ws.cell(row=r, column=2 + j, value=val)
        style_assumption_cell(cell, FMT_CURRENCY)
    PEER_EBITDA_ROW = r

    # Peer EBIT
    r = COMPS_HDR_ROW + 3
    write_label(ws, r, 1, "EBIT (EUR 000)")
    peer_ebit = [85000, 8200000, 4900000]  # approximate
    for j, val in enumerate(peer_ebit):
        cell = ws.cell(row=r, column=2 + j, value=val)
        style_assumption_cell(cell, FMT_CURRENCY)
    PEER_EBIT_ROW = r

    # EV/EBITDA multiples
    r = COMPS_HDR_ROW + 4
    write_label(ws, r, 1, "EV/EBITDA")
    for j in range(3):
        c = cl(2 + j)
        cell = ws.cell(row=r, column=2 + j)
        cell.value = f"=IF({c}{PEER_EBITDA_ROW}<>0,{c}{PEER_EV_ROW}/{c}{PEER_EBITDA_ROW},0)"
        style_formula_cell(cell, FMT_MULTIPLE)
    PEER_EV_EBITDA_ROW = r

    # EV/EBIT multiples
    r = COMPS_HDR_ROW + 5
    write_label(ws, r, 1, "EV/EBIT")
    for j in range(3):
        c = cl(2 + j)
        cell = ws.cell(row=r, column=2 + j)
        cell.value = f"=IF({c}{PEER_EBIT_ROW}<>0,{c}{PEER_EV_ROW}/{c}{PEER_EBIT_ROW},0)"
        style_formula_cell(cell, FMT_MULTIPLE)
    PEER_EV_EBIT_ROW = r

    r = COMPS_HDR_ROW + 7  # blank row then medians

    # Median EV/EBITDA
    write_label(ws, r, 1, "Median EV/EBITDA", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=MEDIAN(B{PEER_EV_EBITDA_ROW}:D{PEER_EV_EBITDA_ROW})"
    style_formula_cell(cell, FMT_MULTIPLE)
    cell.font = BLACK_BOLD
    MED_EV_EBITDA_ROW = r

    # Median EV/EBIT
    r = COMPS_HDR_ROW + 8
    write_label(ws, r, 1, "Median EV/EBIT", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=MEDIAN(B{PEER_EV_EBIT_ROW}:D{PEER_EV_EBIT_ROW})"
    style_formula_cell(cell, FMT_MULTIPLE)
    cell.font = BLACK_BOLD
    MED_EV_EBIT_ROW = r

    r = COMPS_HDR_ROW + 10  # blank then ASM implied
    write_label(ws, r, 1, "ASM International - Implied Valuation", bold=True)
    ASM_IMPL_TITLE_ROW = r

    # ASM EBITDA (2024) = EBIT + D&A from Adj ASM
    r = ASM_IMPL_TITLE_ROW + 1
    write_label(ws, r, 1, "ASM EBITDA (2024, EUR 000)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={ADJ}!G25+{ADJ}!G14"
    style_crossref_cell(cell, FMT_CURRENCY)
    ASM_EBITDA_ROW = r

    # ASM EBIT (2024)
    r = ASM_IMPL_TITLE_ROW + 2
    write_label(ws, r, 1, "ASM EBIT (2024, EUR 000)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"={ADJ}!G25"
    style_crossref_cell(cell, FMT_CURRENCY)
    ASM_EBIT_ROW = r

    # Implied EV (EV/EBITDA)
    r = ASM_IMPL_TITLE_ROW + 4
    write_label(ws, r, 1, "Implied EV (EV/EBITDA)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{ASM_EBITDA_ROW}*B{MED_EV_EBITDA_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)
    IMP_EV_EBITDA_ROW = r

    # - Net Debt
    r = ASM_IMPL_TITLE_ROW + 5
    write_label(ws, r, 1, "- Net Debt")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{NET_DEBT_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # = Equity Value (EV/EBITDA)
    r = ASM_IMPL_TITLE_ROW + 6
    write_label(ws, r, 1, "= Equity Value (EV/EBITDA)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{IMP_EV_EBITDA_ROW}-B{ASM_IMPL_TITLE_ROW + 5}"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    EQ_VAL_COMPS_EBITDA_ROW = r

    # Implied Price (EV/EBITDA)
    r = ASM_IMPL_TITLE_ROW + 7
    write_label(ws, r, 1, "Implied Price (EV/EBITDA)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF(B{SHARES_ROW}<>0,B{EQ_VAL_COMPS_EBITDA_ROW}/B{SHARES_ROW}*1000,0)"
    style_formula_cell(cell, '#,##0.00')
    cell.font = BLACK_BOLD
    PRICE_COMPS_EBITDA_ROW = r

    r = ASM_IMPL_TITLE_ROW + 9  # blank

    # Implied EV (EV/EBIT)
    write_label(ws, r, 1, "Implied EV (EV/EBIT)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{ASM_EBIT_ROW}*B{MED_EV_EBIT_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)
    IMP_EV_EBIT_ROW = r

    # - Net Debt
    r = ASM_IMPL_TITLE_ROW + 10
    write_label(ws, r, 1, "- Net Debt")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{NET_DEBT_ROW}"
    style_formula_cell(cell, FMT_CURRENCY)

    # = Equity Value (EV/EBIT)
    r = ASM_IMPL_TITLE_ROW + 11
    write_label(ws, r, 1, "= Equity Value (EV/EBIT)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{IMP_EV_EBIT_ROW}-B{ASM_IMPL_TITLE_ROW + 10}"
    style_formula_cell(cell, FMT_CURRENCY)
    cell.font = BLACK_BOLD
    EQ_VAL_COMPS_EBIT_ROW = r

    # Implied Price (EV/EBIT)
    r = ASM_IMPL_TITLE_ROW + 12
    write_label(ws, r, 1, "Implied Price (EV/EBIT)", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=IF(B{SHARES_ROW}<>0,B{EQ_VAL_COMPS_EBIT_ROW}/B{SHARES_ROW}*1000,0)"
    style_formula_cell(cell, '#,##0.00')
    cell.font = BLACK_BOLD
    style_double_line_row(ws, r, MAX_COL)
    PRICE_COMPS_EBIT_ROW = r

    # =========================================================================
    # Section 8: FOOTBALL FIELD SUMMARY
    # =========================================================================
    FF_START = ASM_IMPL_TITLE_ROW + 14
    r = FF_START
    style_section_header(ws, r, MAX_COL, "FOOTBALL FIELD SUMMARY")

    r = FF_START + 2
    # Headers
    write_label(ws, r, 1, "Valuation Method", bold=True)
    cell = ws.cell(row=r, column=1)
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell = ws.cell(row=r, column=2, value="Implied Price (EUR)")
    cell.font = BLACK_BOLD
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    FF_HDR_ROW = r

    # DCF (FCFF)
    r = FF_HDR_ROW + 1
    write_label(ws, r, 1, "DCF (FCFF)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{PRICE_DCF_FCFF_ROW}"
    style_formula_cell(cell, '#,##0.00')
    FF_FCFF_ROW = r

    # DCF (FCFE)
    r = FF_HDR_ROW + 2
    write_label(ws, r, 1, "DCF (FCFE)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{PRICE_DCF_FCFE_ROW}"
    style_formula_cell(cell, '#,##0.00')
    FF_FCFE_ROW = r

    # PVAOI
    r = FF_HDR_ROW + 3
    write_label(ws, r, 1, "PVAOI")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{PRICE_PVAOI_ROW}"
    style_formula_cell(cell, '#,##0.00')
    FF_PVAOI_ROW = r

    # PVAE
    r = FF_HDR_ROW + 4
    write_label(ws, r, 1, "PVAE")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{PRICE_PVAE_ROW}"
    style_formula_cell(cell, '#,##0.00')
    FF_PVAE_ROW = r

    # Comps EV/EBITDA
    r = FF_HDR_ROW + 5
    write_label(ws, r, 1, "Comps (EV/EBITDA)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{PRICE_COMPS_EBITDA_ROW}"
    style_formula_cell(cell, '#,##0.00')
    FF_COMPS_EBITDA_ROW = r

    # Comps EV/EBIT
    r = FF_HDR_ROW + 6
    write_label(ws, r, 1, "Comps (EV/EBIT)")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{PRICE_COMPS_EBIT_ROW}"
    style_formula_cell(cell, '#,##0.00')
    FF_COMPS_EBIT_ROW = r

    # Current Price
    r = FF_HDR_ROW + 7
    write_label(ws, r, 1, "Current Share Price")
    cell = ws.cell(row=r, column=2)
    cell.value = f"=B{PRICE_ROW}"
    style_formula_cell(cell, '#,##0.00')
    FF_CURRENT_ROW = r

    r = FF_HDR_ROW + 8  # blank
    style_total_row(ws, r, 2)

    # Average Implied Price (excluding current price)
    r = FF_HDR_ROW + 9
    write_label(ws, r, 1, "Average Implied Price", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f"=AVERAGE(B{FF_FCFF_ROW}:B{FF_COMPS_EBIT_ROW})"
    style_formula_cell(cell, '#,##0.00')
    cell.font = BLACK_BOLD
    FF_AVG_ROW = r

    # Recommendation
    r = FF_HDR_ROW + 10
    write_label(ws, r, 1, "Recommendation", bold=True)
    cell = ws.cell(row=r, column=2)
    cell.value = f'=IF(B{FF_AVG_ROW}>B{FF_CURRENT_ROW},"BUY","SELL")'
    cell.font = Font(bold=True, color="000000", size=12)
    cell.alignment = Alignment(horizontal="center")
    style_double_line_row(ws, r, MAX_COL)

    # --- Freeze panes at B4 ---
    freeze_panes(ws, row=4, col=2)

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading source data...")
    source_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    print("Creating output workbook...")
    wb = openpyxl.Workbook()

    print("Building Cover...")
    build_cover(wb)

    print("Building Input tabs...")
    build_input_tabs(wb, source_wb)

    print("Building Std ASM...")
    build_standardized_asm(wb)

    print("Building Std AIXTRON...")
    build_standardized_tab(wb, 'AIXA', "'Input AIXTRON'", None, 'Std AIXTRON')

    print("Building Std AMAT...")
    build_standardized_tab(wb, 'AMAT', "'Input AMAT'", None, 'Std AMAT')

    print("Building Std LRCX...")
    build_standardized_tab(wb, 'LRCX', "'Input LRCX'", None, 'Std LRCX')

    print("Building Adjustments tab...")
    build_adjustments_tab(wb)

    print("Building Adj ASM...")
    build_adjusted_tab(wb, 'ASM', 'Std ASM', 'Adj ASM')

    print("Building Adj AIXTRON...")
    build_adjusted_tab(wb, 'AIXA', 'Std AIXTRON', 'Adj AIXTRON')

    print("Building Adj AMAT...")
    build_adjusted_tab(wb, 'AMAT', 'Std AMAT', 'Adj AMAT')

    print("Building Adj LRCX...")
    build_adjusted_tab(wb, 'LRCX', 'Std LRCX', 'Adj LRCX')

    print("Building Ratio Definitions...")
    build_ratio_definitions(wb)

    print("Building Ratios ASM...")
    build_ratio_tab(wb, 'Adj ASM', 'Ratios ASM')

    print("Building Ratios AIXTRON...")
    build_ratio_tab(wb, 'Adj AIXTRON', 'Ratios AIXTRON')

    print("Building Ratios AMAT...")
    build_ratio_tab(wb, 'Adj AMAT', 'Ratios AMAT')

    print("Building Ratios LRCX...")
    build_ratio_tab(wb, 'Adj LRCX', 'Ratios LRCX')

    print("Building Ratio Comparison...")
    build_ratio_comparison(wb)

    print("Building Forecast...")
    build_forecast_tab(wb)

    print("Building Valuation...")
    build_valuation_tab(wb)

    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done!")

    # Print summary
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")

    source_wb.close()
    wb.close()


if __name__ == "__main__":
    main()
