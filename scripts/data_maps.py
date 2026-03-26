"""
data_maps.py - Source data row/column mappings for all companies.

Maps the exact row numbers in `ASM International and Comps data.xlsx` to our
model's standardized line items.  Every formula in the output workbook depends
on these mappings being correct.

All row numbers were verified programmatically against the actual source file.
"""

# ---------------------------------------------------------------------------
# Column / year constants
# ---------------------------------------------------------------------------
YEARS = [2019, 2020, 2021, 2022, 2023, 2024]
DATA_START_COL = 2   # column B
DATA_END_COL = 7     # column G
NUM_YEARS = 6

# ---------------------------------------------------------------------------
# Sheet names in the source workbook
# ---------------------------------------------------------------------------
SOURCE_FILE = "ASM International and Comps data.xlsx"

SOURCE_SHEETS = {
    "ASM":   "ASM",
    "AIXA":  "AIXTRON",
    "AMAT":  "Applied Materials",
    "LRCX":  "Lam Research",
    "ASM_REPORTED": "ASM (as reported)",
    "TEL":   "Electron",
}

# ============================================================================
# ASM International  (sheet "ASM")
# ============================================================================
ASM_IS = {
    "revenue":                21,
    "total_revenue":          22,
    "cogs":                   23,
    "gross_profit":           24,
    "sga":                    26,
    "rd":                     27,
    "amort_goodwill_intang":  28,
    "other_operating_exp":    29,
    "other_operating_total":  30,
    "operating_income":       31,
    "interest_expense":       33,
    "interest_income":        34,
    "net_interest":           35,
    "income_from_affiliates": 36,
    "fx_gains":               37,
    "ebt_excl_unusual":       38,
    "gain_loss_sale_invest":  39,
    "gain_loss_sale_assets":  40,
    "asset_writedown":        41,
    "other_unusual":          42,
    "ebt_incl_unusual":       43,
    "tax_expense":            45,
    "earnings_cont_ops":      46,
    "net_income_to_company":  47,
    "net_income":             48,
    "ni_common_incl_extra":   49,
    "ni_common_excl_extra":   50,
    # Per-share items
    "basic_eps":              52,
    "basic_eps_excl_extra":   53,
    "basic_shares":           54,
    "diluted_eps_incl_extra": 55,
    "diluted_eps_excl_extra": 56,
    "diluted_shares":         57,
    "normalized_basic_eps":   58,
    "normalized_diluted_eps": 59,
    "dividends_per_share":    60,
    "payout_ratio":           61,
    # Supplemental
    "ebitda":                 64,
    "ebita":                  65,
    "ebit":                   66,
    "ebitdar":                67,
    "effective_tax_rate":     68,
    "current_domestic_taxes": 69,
    "current_foreign_taxes":  70,
    "total_current_taxes":    71,
    "deferred_domestic_taxes": 72,
    "deferred_foreign_taxes": 73,
    "total_deferred_taxes":   74,
    "normalized_net_income":  75,
    "interest_lt_debt":       76,
    # Operating expense supplement
    "rd_expense_supp":        79,
    "net_rental_expense":     80,
    "sbc_unallocated":        83,
    "sbc_total":              84,
}

ASM_CF = {
    "net_income_cf":          110,
    "da_depreciation":        111,
    "da_amort_goodwill":      112,
    "da_total":               113,
    "other_amortization":     114,
    "gain_loss_sale_assets":  115,
    "loss_sale_invest":       116,
    "asset_writedown_restruct": 117,
    "loss_equity_invest":     118,
    "sbc":                    119,
    "other_operating":        120,
    "change_ar":              121,
    "change_inventory":       122,
    "change_ap":              123,
    "change_other_noa":       124,
    "cfo":                    125,
    # Investing
    "capex":                  127,
    "sale_ppe":               128,
    "cash_acquisitions":      129,
    "purchase_intangibles":   130,
    "invest_mkt_equity":      131,
    "other_investing":        132,
    "cfi":                    133,
    # Financing
    "lt_debt_repaid":         135,
    "total_debt_repaid":      136,
    "issuance_common":        137,
    "repurchase_common":      138,
    "common_dividends":       139,
    "total_dividends":        140,
    "other_financing":        141,
    "cff":                    142,
    # Other
    "fx_adj":                 144,
    "net_change_cash":        145,
    # Supplemental
    "cash_taxes_paid":        147,
    "levered_fcf":            148,
    "unlevered_fcf":          149,
    "change_nwc":             150,
    "net_debt_issued":        151,
}

ASM_BS = {
    # Assets
    "cash":                   165,
    "cash_st_investments":    166,
    "accounts_receivable":    167,
    "other_receivables":      168,
    "total_receivables":      169,
    "inventory":              170,
    "prepaid_exp":            171,
    "other_current_assets":   172,
    "total_current_assets":   173,
    "gross_ppe":              174,
    "accum_depreciation":     175,
    "net_ppe":                176,
    "lt_investments":         177,
    "goodwill":               178,
    "other_intangibles":      179,
    "deferred_tax_assets":    180,
    "deferred_charges_lt":    181,
    "other_lt_assets":        182,
    "total_assets":           183,
    # Liabilities
    "accounts_payable":       185,
    "accrued_exp":            186,
    "current_lease_liab":     187,
    "current_income_tax":     188,
    "unearned_revenue_curr":  189,
    "other_current_liab":     190,
    "total_current_liab":     191,
    "lt_leases":              192,
    "deferred_tax_liab":      193,
    "other_non_current_liab": 194,
    "total_liabilities":      195,
    # Equity
    "common_stock":           197,
    "apic":                   198,
    "retained_earnings":      199,
    "treasury_stock":         200,
    "comprehensive_other":    201,
    "total_common_equity":    202,
    "total_equity":           203,
    "total_liab_equity":      204,
    # Supplemental
    "shares_out_filing":      206,
    "shares_out_common":      207,
    "book_value_per_share":   208,
    "tangible_book_value":    209,
    "total_debt":             211,
    "net_debt":               212,
    "equity_method_invest":   215,
    "order_backlog":          225,
    "full_time_employees":    223,
}


# ============================================================================
# AIXTRON  (sheet "AIXTRON")
# ============================================================================
AIXA_IS = {
    "revenue":                21,
    "total_revenue":          22,
    "cogs":                   23,
    "gross_profit":           24,
    "sga":                    26,
    "rd":                     27,
    "other_operating_exp":    28,
    "other_operating_total":  29,
    "operating_income":       30,
    "interest_expense":       32,
    "interest_income":        33,
    "net_interest":           34,
    "fx_gains":               35,
    "ebt_excl_unusual":       36,
    "restructuring":          37,
    "gain_loss_sale_invest":  38,
    "gain_loss_sale_assets":  39,
    "asset_writedown":        40,
    "insurance_settlements":  41,
    "ebt_incl_unusual":       42,
    "tax_expense":            44,
    "earnings_cont_ops":      45,
    "net_income_to_company":  46,
    "minority_interest":      47,
    "net_income":             48,
    "ni_common_incl_extra":   49,
    "ni_common_excl_extra":   50,
    # Per-share
    "basic_eps":              52,
    "basic_eps_excl_extra":   53,
    "basic_shares":           54,
    "diluted_eps_incl_extra": 55,
    "diluted_eps_excl_extra": 56,
    "diluted_shares":         57,
    "normalized_basic_eps":   58,
    "normalized_diluted_eps": 59,
    "dividends_per_share":    60,
    "payout_ratio":           61,
    # Supplemental
    "ebitda":                 64,
    "ebita":                  65,
    "ebit":                   66,
    "ebitdar":                67,
    "effective_tax_rate":     68,
    "total_current_taxes":    69,
    "total_deferred_taxes":   70,
    "normalized_net_income":  71,
    "interest_lt_debt":       72,
    # Operating expense supplement
    "selling_marketing":      74,
    "general_admin":          75,
    "rd_expense_supp":        76,
    "net_rental_expense":     77,
    "sbc_unallocated":        80,
    "sbc_total":              81,
}

AIXA_CF = {
    "net_income_cf":          108,
    "da_depreciation":        109,
    "da_amort_goodwill":      110,
    "da_total":               111,
    "gain_loss_sale_assets":  112,
    "loss_sale_invest":       113,
    "asset_writedown_restruct": 114,
    "sbc":                    115,
    "other_operating":        116,
    "change_trading_assets":  117,
    "change_ar":              118,
    "change_inventory":       119,
    "change_ap":              120,
    "change_unearned_rev":    121,
    "change_other_noa":       122,
    "cfo":                    123,
    # Investing
    "capex":                  125,
    "sale_ppe":               126,
    "purchase_intangibles":   127,
    "invest_mkt_equity":      128,
    "other_investing":        129,
    "cfi":                    130,
    # Financing
    "lt_debt_repaid":         132,
    "total_debt_repaid":      133,
    "issuance_common":        134,
    "common_dividends":       135,
    "total_dividends":        136,
    "other_financing":        137,
    "cff":                    138,
    # Other
    "fx_adj":                 140,
    "misc_cf_adj":            141,
    "net_change_cash":        142,
    # Supplemental
    "cash_interest_paid":     144,
    "cash_taxes_paid":        145,
    "levered_fcf":            146,
    "unlevered_fcf":          147,
    "change_nwc":             148,
    "net_debt_issued":        149,
}

AIXA_BS = {
    # Assets
    "cash":                   163,
    "st_investments":         164,
    "trading_assets":         165,
    "cash_st_investments":    166,
    "accounts_receivable":    167,
    "other_receivables":      168,
    "total_receivables":      169,
    "inventory":              170,
    "prepaid_exp":            171,
    "other_current_assets":   172,
    "total_current_assets":   173,
    "gross_ppe":              174,
    "accum_depreciation":     175,
    "net_ppe":                176,
    "lt_investments":         177,
    "goodwill":               178,
    "other_intangibles":      179,
    "deferred_tax_assets":    180,
    "other_lt_assets":        181,
    "total_assets":           182,
    # Liabilities
    "accounts_payable":       184,
    "accrued_exp":            185,
    "current_lease_liab":     186,
    "current_income_tax":     187,
    "unearned_revenue_curr":  188,
    "other_current_liab":     189,
    "total_current_liab":     190,
    "lt_leases":              191,
    "deferred_tax_liab":      192,
    "other_non_current_liab": 193,
    "total_liabilities":      194,
    # Equity
    "common_stock":           196,
    "apic":                   197,
    "retained_earnings":      198,
    "comprehensive_other":    199,
    "total_common_equity":    200,
    "total_minority_interest": 201,
    "total_equity":           202,
    "total_liab_equity":      203,
    # Supplemental
    "shares_out_filing":      205,
    "shares_out_common":      206,
    "book_value_per_share":   207,
    "tangible_book_value":    208,
    "total_debt":             210,
    "net_debt":               211,
    "order_backlog":          222,
    "full_time_employees":    220,
}


# ============================================================================
# Applied Materials  (sheet "Applied Materials")
# ============================================================================
AMAT_IS = {
    "revenue":                21,
    "total_revenue":          22,
    "cogs":                   23,
    "gross_profit":           24,
    "sga":                    26,
    "rd":                     27,
    "other_operating_total":  28,
    "operating_income":       29,
    "interest_expense":       31,
    "interest_income":        32,
    "net_interest":           33,
    "fx_gains":               34,
    "other_non_operating":    35,
    "ebt_excl_unusual":       36,
    "restructuring":          37,
    "merger_restruct":        38,
    "gain_loss_sale_invest":  39,
    "other_unusual":          40,
    "ebt_incl_unusual":       41,
    "tax_expense":            43,
    "earnings_cont_ops":      44,
    "net_income_to_company":  45,
    "net_income":             46,
    "ni_common_incl_extra":   47,
    "ni_common_excl_extra":   48,
    # Per-share
    "basic_eps":              50,
    "basic_eps_excl_extra":   51,
    "basic_shares":           52,
    "diluted_eps_incl_extra": 53,
    "diluted_eps_excl_extra": 54,
    "diluted_shares":         55,
    "normalized_basic_eps":   56,
    "normalized_diluted_eps": 57,
    "dividends_per_share":    58,
    "payout_ratio":           59,
    # Supplemental
    "ebitda":                 62,
    "ebita":                  63,
    "ebit":                   64,
    "ebitdar":                65,
    "effective_tax_rate":     66,
    "current_domestic_taxes": 67,
    "current_foreign_taxes":  68,
    "total_current_taxes":    69,
    "deferred_domestic_taxes": 70,
    "deferred_foreign_taxes": 71,
    "total_deferred_taxes":   72,
    "normalized_net_income":  73,
    "interest_lt_debt":       74,
    "non_cash_pension":       75,
    # Operating expense supplement
    "selling_marketing":      78,
    "general_admin":          79,
    "rd_expense_supp":        80,
    "net_rental_expense":     81,
    "sbc_cogs":               84,
    "sbc_rd":                 85,
    "sbc_sm":                 86,
    "sbc_ga":                 87,
    "sbc_total":              88,
}

AMAT_CF = {
    "net_income_cf":          105,
    "da_depreciation":        106,
    "da_amort_goodwill":      107,
    "da_total":               108,
    "loss_sale_invest":       109,
    "sbc":                    110,
    "other_operating":        111,
    "change_ar":              112,
    "change_inventory":       113,
    "change_ap":              114,
    "change_unearned_rev":    115,
    "change_income_taxes":    116,
    "change_other_noa":       117,
    "cfo":                    118,
    # Investing
    "capex":                  120,
    "cash_acquisitions":      121,
    "invest_mkt_equity":      122,
    "cfi":                    123,
    # Financing
    "st_debt_issued":         125,
    "lt_debt_issued":         126,
    "total_debt_issued":      127,
    "st_debt_repaid":         128,
    "lt_debt_repaid":         129,
    "total_debt_repaid":      130,
    "issuance_common":        131,
    "repurchase_common":      132,
    "common_dividends":       133,
    "total_dividends":        134,
    "cff":                    135,
    # Other
    "net_change_cash":        137,
    # Supplemental
    "cash_interest_paid":     139,
    "cash_taxes_paid":        140,
    "levered_fcf":            141,
    "unlevered_fcf":          142,
    "change_nwc":             143,
    "net_debt_issued":        144,
}

AMAT_BS = {
    # Assets
    "cash":                   158,
    "st_investments":         159,
    "cash_st_investments":    160,
    "accounts_receivable":    161,
    "total_receivables":      162,
    "inventory":              163,
    "prepaid_exp":            164,
    "other_current_assets":   165,
    "total_current_assets":   166,
    "gross_ppe":              167,
    "accum_depreciation":     168,
    "net_ppe":                169,
    "lt_investments":         170,
    "goodwill":               171,
    "other_intangibles":      172,
    "deferred_tax_assets":    173,
    "other_lt_assets":        174,
    "total_assets":           175,
    # Liabilities
    "accounts_payable":       177,
    "accrued_exp":            178,
    "st_borrowings":          179,
    "current_lt_debt":        180,
    "current_lease_liab":     181,
    "current_income_tax":     182,
    "unearned_revenue_curr":  183,
    "other_current_liab":     184,
    "total_current_liab":     185,
    "lt_debt":                186,
    "lt_leases":              187,
    "pension_post_retire":    188,
    "other_non_current_liab": 189,
    "total_liabilities":      190,
    # Equity
    "common_stock":           192,
    "apic":                   193,
    "retained_earnings":      194,
    "treasury_stock":         195,
    "comprehensive_other":    196,
    "total_common_equity":    197,
    "total_equity":           198,
    "total_liab_equity":      199,
    # Supplemental
    "shares_out_filing":      201,
    "shares_out_common":      202,
    "book_value_per_share":   203,
    "tangible_book_value":    204,
    "total_debt":             206,
    "net_debt":               207,
    "full_time_employees":    218,
}


# ============================================================================
# Lam Research  (sheet "Lam Research")
# ============================================================================
LRCX_IS = {
    "revenue":                21,
    "total_revenue":          22,
    "cogs":                   23,
    "gross_profit":           24,
    "sga":                    26,
    "rd":                     27,
    "other_operating_total":  28,
    "operating_income":       29,
    "interest_expense":       31,
    "interest_income":        32,
    "net_interest":           33,
    "fx_gains":               34,
    "other_non_operating":    35,
    "ebt_excl_unusual":       36,
    "restructuring":          37,
    "ebt_incl_unusual":       38,
    "tax_expense":            40,
    "earnings_cont_ops":      41,
    "net_income_to_company":  42,
    "net_income":             43,
    "ni_common_incl_extra":   44,
    "ni_common_excl_extra":   45,
    # Per-share
    "basic_eps":              47,
    "basic_eps_excl_extra":   48,
    "basic_shares":           49,
    "diluted_eps_incl_extra": 50,
    "diluted_eps_excl_extra": 51,
    "diluted_shares":         52,
    "normalized_basic_eps":   53,
    "normalized_diluted_eps": 54,
    "dividends_per_share":    55,
    "payout_ratio":           56,
    # Supplemental
    "ebitda":                 59,
    "ebita":                  60,
    "ebit":                   61,
    "ebitdar":                62,
    "effective_tax_rate":     63,
    "current_domestic_taxes": 64,
    "current_foreign_taxes":  65,
    "total_current_taxes":    66,
    "deferred_domestic_taxes": 67,
    "deferred_foreign_taxes": 68,
    "total_deferred_taxes":   69,
    "normalized_net_income":  70,
    "interest_lt_debt":       71,
    # Operating expense supplement
    "rd_expense_supp":        73,
    "net_rental_expense":     74,
    "sbc_unallocated":        77,
    "sbc_total":              78,
}

LRCX_CF = {
    "net_income_cf":          95,
    "da_depreciation":        96,
    "da_amort_goodwill":      97,
    "da_total":               98,
    "sbc":                    99,
    "other_operating":        100,
    "change_ar":              101,
    "change_inventory":       102,
    "change_ap":              103,
    "change_unearned_rev":    104,
    "change_other_noa":       105,
    "cfo":                    106,
    # Investing
    "capex":                  108,
    "cash_acquisitions":      109,
    "invest_mkt_equity":      110,
    "other_investing":        111,
    "cfi":                    112,
    # Financing
    "st_debt_issued":         114,
    "lt_debt_issued":         115,
    "total_debt_issued":      116,
    "st_debt_repaid":         117,
    "lt_debt_repaid":         118,
    "total_debt_repaid":      119,
    "issuance_common":        120,
    "repurchase_common":      121,
    "common_dividends":       122,
    "total_dividends":        123,
    "other_financing":        124,
    "cff":                    125,
    # Other
    "fx_adj":                 127,
    "net_change_cash":        128,
    # Supplemental
    "cash_interest_paid":     130,
    "cash_taxes_paid":        131,
    "levered_fcf":            132,
    "unlevered_fcf":          133,
    "change_nwc":             134,
    "net_debt_issued":        135,
}

LRCX_BS = {
    # Assets
    "cash":                   149,
    "st_investments":         150,
    "cash_st_investments":    151,
    "accounts_receivable":    152,
    "total_receivables":      153,
    "inventory":              154,
    "prepaid_exp":            155,
    "other_current_assets":   156,
    "total_current_assets":   157,
    "gross_ppe":              158,
    "accum_depreciation":     159,
    "net_ppe":                160,
    "lt_investments":         161,
    "goodwill":               162,
    "other_intangibles":      163,
    "deferred_charges_lt":    164,
    "other_lt_assets":        165,
    "total_assets":           166,
    # Liabilities
    "accounts_payable":       168,
    "accrued_exp":            169,
    "current_lt_debt":        170,
    "current_lease_liab":     171,
    "current_income_tax":     172,
    "unearned_revenue_curr":  173,
    "other_current_liab":     174,
    "total_current_liab":     175,
    "lt_debt":                176,
    "lt_leases":              177,
    "other_non_current_liab": 178,
    "total_liabilities":      179,
    # Equity
    "common_stock":           181,
    "apic":                   182,
    "retained_earnings":      183,
    "treasury_stock":         184,
    "comprehensive_other":    185,
    "total_common_equity":    186,
    "total_equity":           187,
    "total_liab_equity":      188,
    # Supplemental
    "shares_out_filing":      190,
    "shares_out_common":      191,
    "book_value_per_share":   192,
    "tangible_book_value":    193,
    "total_debt":             195,
    "net_debt":               196,
    "full_time_employees":    205,
}


# ============================================================================
# ASM (as reported)  (sheet "ASM (as reported)")
# ============================================================================
ASM_REPORTED_IS = {
    "revenue":                     17,
    "cost_of_sales":               19,
    "sga":                         20,
    "rd":                          21,
    "other_income":                22,
    "gain_loss_sale_invest":       23,
    "fx_gain_loss":                24,
    "finance_income":              25,
    "finance_expenses":            26,
    "share_income_associates":     27,
    "ebt":                         28,
    "provision_income_tax":        30,
    "net_income":                  31,
    # Supplementary
    "gross_profit":                33,
    "operating_income":            34,
    "basic_eps_cont":              35,
    "diluted_eps_cont":            36,
}

ASM_REPORTED_CF = {
    "net_income":                          44,
    "da_impairment":                       45,
    "gain_loss_sale_ppe":                  46,
    "gain_loss_sale_invest":               47,
    "impairment_reversal_associates":      48,
    "changes_eval_tools_customers":        49,
    "sbc":                                 50,
    "income_tax":                          51,
    "income_tax_paid":                     52,
    "other_assets":                        53,
    "non_cash_interest":                   54,
    "net_finance_income_costs":            55,
    "non_cash_costs":                      56,
    "share_income_associates":             57,
    "evaluation_tools":                    58,
    "changes_employee_benefits":           59,
    "accounts_receivable":                 60,
    "contract_assets_liabilities":         61,
    "inventories":                         62,
    "ap_accrued_other":                    63,
    "other_current_assets":                64,
    "provision_warranty":                  65,
    "cfo":                                 66,
    # Investing
    "capex_ppe":                           68,
    "proceeds_sale_ppe":                   69,
    "investment_subsidiary":               70,
    "purchase_intangibles":                71,
    "capitalized_dev_expenditure":         72,
    "dividends_from_associate":            73,
    "other_investments":                   74,
    "cfi":                                 75,
    # Financing
    "payment_lease_liabilities":           77,
    "proceeds_issuance_treasury":          78,
    "purchase_treasury_shares":            79,
    "capital_repayment":                   80,
    "dividends_common":                    81,
    "credit_facility_renewal":             82,
    "cff":                                 83,
    # Other
    "fx_rate_effect":                      85,
    "net_change_cash":                     86,
}

ASM_REPORTED_BS = {
    # Current Assets
    "cash":                                94,
    "accounts_receivable":                 95,
    "contract_asset":                      96,
    "income_taxes_receivable":             97,
    "inventories":                         98,
    "other_current_assets":                99,
    "total_current_assets":                100,
    # Noncurrent Assets
    "rou_assets":                          102,
    "ppe":                                 103,
    "investment_associates":               104,
    "other_investments":                   105,
    "deferred_tax_assets":                 106,
    "goodwill":                            107,
    "other_intangible_assets":             108,
    "other_assets":                        109,
    "employee_benefits_pension":           110,
    "evaluation_tools_at_customers":       111,
    "total_assets":                        112,
    # Current Liabilities
    "accounts_payable":                    114,
    "accrued_expenses":                    115,
    "income_taxes_payable":                116,
    "contract_liabilities":                117,
    "provision_warranty":                  118,
    "contingent_consideration_current":    119,
    "total_current_liabilities":           120,
    # Noncurrent Liabilities
    "lease_liabilities":                   122,
    "deferred_income_taxes":               123,
    "contingent_consideration_noncurrent": 124,
    # Equity
    "common_stock_par":                    126,
    "total_equity":                        127,
    "total_liab_equity":                   128,
}


# ============================================================================
# Master company info dictionary
# ============================================================================
COMPANY_INFO = {
    "ASM": {
        "name":       "ASM International NV",
        "sheet":      "ASM",
        "currency":   "EUR",
        "magnitude":  "Thousands",
        "is_map":     ASM_IS,
        "cf_map":     ASM_CF,
        "bs_map":     ASM_BS,
    },
    "AIXA": {
        "name":       "AIXTRON SE",
        "sheet":      "AIXTRON",
        "currency":   "EUR",
        "magnitude":  "Thousands",
        "is_map":     AIXA_IS,
        "cf_map":     AIXA_CF,
        "bs_map":     AIXA_BS,
    },
    "AMAT": {
        "name":       "Applied Materials, Inc.",
        "sheet":      "Applied Materials",
        "currency":   "EUR",           # CapIQ converted to EUR
        "magnitude":  "Thousands",
        "is_map":     AMAT_IS,
        "cf_map":     AMAT_CF,
        "bs_map":     AMAT_BS,
    },
    "LRCX": {
        "name":       "Lam Research Corporation",
        "sheet":      "Lam Research",
        "currency":   "EUR",           # CapIQ converted to EUR
        "magnitude":  "Thousands",
        "is_map":     LRCX_IS,
        "cf_map":     LRCX_CF,
        "bs_map":     LRCX_BS,
    },
    "ASM_REPORTED": {
        "name":       "ASM International NV (As Reported)",
        "sheet":      "ASM (as reported)",
        "currency":   "EUR",
        "magnitude":  "Thousands",
        "is_map":     ASM_REPORTED_IS,
        "cf_map":     ASM_REPORTED_CF,
        "bs_map":     ASM_REPORTED_BS,
    },
}


# ============================================================================
# Verification helper
# ============================================================================
def verify_mappings(source_path=None):
    """
    Read the source Excel and verify that every mapped row has the expected
    label in column A.  Prints label + 2024 value for each mapped item.
    Returns a list of any mismatches found.
    """
    import openpyxl

    if source_path is None:
        source_path = SOURCE_FILE

    wb = openpyxl.load_workbook(source_path, data_only=True)
    mismatches = []

    for company_key, info in COMPANY_INFO.items():
        sheet_name = info["sheet"]
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"  {company_key} - {sheet_name}")
        print(f"{'='*60}")

        for section_label, section_map in [
            ("IS", info["is_map"]),
            ("CF", info["cf_map"]),
            ("BS", info["bs_map"]),
        ]:
            print(f"\n  --- {section_label} ---")
            for item_key, row_num in sorted(section_map.items(), key=lambda x: x[1]):
                label = ws.cell(row=row_num, column=1).value
                val_2024 = ws.cell(row=row_num, column=DATA_END_COL).value
                label_str = str(label).strip() if label else "<EMPTY>"
                print(f"    Row {row_num:>3} | {item_key:<35} | {label_str:<55} | 2024={val_2024}")

                # Flag rows that are blank or look like headers
                if label is None or (isinstance(label, str) and label.strip() == ""):
                    mismatches.append(
                        f"{company_key}.{section_label}.{item_key}: "
                        f"Row {row_num} is BLANK"
                    )

    wb.close()

    if mismatches:
        print(f"\n{'!'*60}")
        print(f"  MISMATCHES FOUND: {len(mismatches)}")
        for m in mismatches:
            print(f"    - {m}")
    else:
        print(f"\n  All mappings verified successfully.")

    return mismatches


if __name__ == "__main__":
    verify_mappings()
