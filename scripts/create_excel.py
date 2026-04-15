#!/usr/bin/env python3
"""
Create a professionally formatted Stanford Professors Outreach Excel workbook.
McKinsey/Goldman Sachs caliber formatting with Stanford branding.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Color Palette ──────────────────────────────────────────────────────────────
DARK_NAVY      = "1B2A4A"
CARDINAL_RED   = "8C1515"
STANFORD_WHITE = "FFFFFF"
LIGHT_GRAY     = "F5F5F5"
MEDIUM_GRAY    = "D9D9D9"
BORDER_GRAY    = "BFBFBF"
DARK_GRAY      = "4A4A4A"
AREA_BLUE      = "1F4E79"       # Engineering area
AREA_PURPLE    = "6B3FA0"       # Interdisciplinary area
GREEN_STATUS   = "C6EFCE"
GREEN_FONT     = "006100"
YELLOW_STATUS  = "FFEB9C"
YELLOW_FONT    = "9C6500"
RED_STATUS     = "FFC7CE"
RED_FONT       = "9C0006"
UNAVAIL_FILL   = "FDDCCA"       # Light salmon for unavailable
PRIORITY_HIGH  = "E2EFDA"       # Light green
PRIORITY_MED   = "FCE4D6"       # Light peach
PRIORITY_LOW   = "D6DCE4"       # Blue-gray
SUBTITLE_BG    = "2C3E6B"       # Slightly lighter navy for subtitle

# ── Border Styles ──────────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)
header_border = Border(
    left=Side(style="thin", color=STANFORD_WHITE),
    right=Side(style="thin", color=STANFORD_WHITE),
    top=Side(style="medium", color=STANFORD_WHITE),
    bottom=Side(style="medium", color=STANFORD_WHITE),
)
bottom_accent = Border(
    bottom=Side(style="medium", color=CARDINAL_RED),
)

# ── Professor Data ─────────────────────────────────────────────────────────────
PROFESSORS = [
    {
        "num": 1, "name": "Kathleen Eisenhardt", "dept": "MS&E / Strategy",
        "area": "Engineering + GSB", "focus": "Strategy in tech companies, simple rules, business model design",
        "recent": '"Decentralization in Organizations" (AMJ Annals 2025), "Measure Twice Cut Once" (Org Science 2024)',
        "url": "https://profiles.stanford.edu/kathleen-eisenhardt",
        "email": "kme@stanford.edu", "website": "https://www.stanford.edu/~kme/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 2, "name": "Chuck Eesley", "dept": "MS&E",
        "area": "Engineering", "focus": "AI & entrepreneurship, technology entrepreneurship",
        "recent": '"Companies inadvertently fund online misinformation" (Nature 2024), CROSSROADS paper (Org Science 2025)',
        "url": "https://profiles.stanford.edu/chuck-eesley",
        "email": "cee@stanford.edu", "website": "https://www.stanford.edu/~cee/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 3, "name": "Ramesh Johari", "dept": "MS&E",
        "area": "Engineering", "focus": "Online platforms, marketplace design, causal inference",
        "recent": '"Quality Selection in Two-Sided Markets" (OR 2024), "Switchback Price Experiments" (2024)',
        "url": "https://profiles.stanford.edu/ramesh-johari",
        "email": "rjohari@stanford.edu", "website": "https://web.stanford.edu/~rjohari/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 4, "name": "Pamela Hinds", "dept": "MS&E (Chair)",
        "area": "Engineering", "focus": "Technology's effect on teams, collaboration, innovation",
        "recent": '"Generative AI and Progressive Encapsulation" (Org Studies 2024)',
        "url": "https://profiles.stanford.edu/pamela-hinds",
        "email": "phinds@stanford.edu", "website": "https://web.stanford.edu/~phinds/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 5, "name": "Melissa Valentine", "dept": "MS&E",
        "area": "Engineering", "focus": "AI transforming work, algorithmic management",
        "recent": '"Agentforce" paper (J Org Design 2025), Flash Teams book (MIT Press 2025)',
        "url": "https://profiles.stanford.edu/melissa-valentine",
        "email": "mav@stanford.edu", "website": "https://web.stanford.edu/~mav/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 6, "name": "Ashish Goel", "dept": "MS&E (Chair)",
        "area": "Engineering", "focus": "Algorithms, social networks, democracy, crowdsourcing",
        "recent": "2024 ACM Fellow, Participatory Budgeting Platform (100+ elections)",
        "url": "https://profiles.stanford.edu/ashish-goel",
        "email": "ashishg@stanford.edu", "website": "https://web.stanford.edu/~ashishg/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 7, "name": "Itai Ashlagi", "dept": "MS&E",
        "area": "Engineering", "focus": "Market design, matching markets, kidney exchange",
        "recent": '2024 Lanchester Prize, "Congested Waiting Lists" (2025)',
        "url": "https://profiles.stanford.edu/itai-ashlagi",
        "email": "iashlagi@stanford.edu", "website": "https://web.stanford.edu/~iashlagi/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 8, "name": "Amin Saberi", "dept": "MS&E",
        "area": "Engineering", "focus": "Algorithms for economics, market design",
        "recent": '"MAGNOLIA" (ICML 2024), "Ride Hailing Matching" (OR 2024), back-to-back Test of Time Awards',
        "url": "https://profiles.stanford.edu/amin-saberi",
        "email": "saberi@stanford.edu", "website": "https://web.stanford.edu/~saberi/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 9, "name": "Fei-Fei Li", "dept": "Computer Science",
        "area": "Engineering", "focus": "AI, computer vision, spatial intelligence",
        "recent": "World Labs ($5B valuation), Time Person of Year 2025, Queen Elizabeth Prize 2025",
        "url": "https://profiles.stanford.edu/fei-fei-li",
        "email": "feifeili@stanford.edu", "website": "https://vision.stanford.edu/feifeili/",
        "status": "", "notes": "On partial leave", "flag": "On partial leave",
    },
    {
        "num": 10, "name": "Percy Liang", "dept": "Computer Science",
        "area": "Engineering", "focus": "Foundation models, LLMs, benchmarking",
        "recent": "HELM benchmark, CRFM director, Marin open foundation models",
        "url": "https://profiles.stanford.edu/percy-liang",
        "email": "pliang@cs.stanford.edu", "website": "https://cs.stanford.edu/~pliang/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 11, "name": "Michael Bernstein", "dept": "Computer Science",
        "area": "Engineering", "focus": "HCI, generative AI agents",
        "recent": '"Generative Agent Simulations of 1,000 People" (2024), Flash Teams book',
        "url": "https://profiles.stanford.edu/michael-bernstein",
        "email": "msb@cs.stanford.edu", "website": "https://hci.stanford.edu/msb/",
        "status": "", "notes": "On leave until Fall 2026", "flag": "On leave until Fall 2026",
    },
    {
        "num": 12, "name": "Sachin Katti", "dept": "EE + CS",
        "area": "Engineering", "focus": "Wireless networks, AI for mobile",
        "recent": "LEFT STANFORD - Now Intel CTO/CAO (April 2025)",
        "url": "https://profiles.stanford.edu/sachin-katti",
        "email": "skatti@stanford.edu", "website": "",
        "status": "", "notes": "Left Stanford, now Intel CTO/CAO (April 2025)", "flag": "Left Stanford",
    },
    {
        "num": 13, "name": "Larry Leifer", "dept": "Mechanical Engineering",
        "area": "Engineering", "focus": "Design thinking, NeuroDesign, innovation leadership",
        "recent": "CDR 40-year anniversary (2024), NeuroDesign research",
        "url": "https://profiles.stanford.edu/larry-leifer",
        "email": "larry.leifer@stanford.edu", "website": "https://me.stanford.edu/people/larry-leifer",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 14, "name": "Mark Cutkosky", "dept": "Mechanical Engineering",
        "area": "Engineering", "focus": "Robotics, biomimetics, design innovation",
        "recent": "ReachBot for planetary exploration, gecko-inspired space adhesives",
        "url": "https://profiles.stanford.edu/mark-cutkosky",
        "email": "cutkosky@stanford.edu", "website": "https://me.stanford.edu/people/mark-cutkosky",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 15, "name": "Allison Okamura", "dept": "Mechanical Engineering",
        "area": "Engineering", "focus": "Haptics, medical robotics, soft robotics",
        "recent": '"Haptiknit" wearable haptics (2024), "Fourigami" haptic device (2025)',
        "url": "https://profiles.stanford.edu/allison-okamura",
        "email": "aokamura@stanford.edu", "website": "https://charm.stanford.edu/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 16, "name": "Erik Brynjolfsson", "dept": "HAI / SIEPR",
        "area": "Interdisciplinary", "focus": "Economics of AI, digital economy, productivity",
        "recent": '"Generative AI at Work" (QJE 2025, 14% productivity boost), AI J-Curve framework',
        "url": "https://profiles.stanford.edu/erik-brynjolfsson",
        "email": "erikb@stanford.edu", "website": "https://www.brynjolfsson.com/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 17, "name": "Susan Athey", "dept": "GSB Economics",
        "area": "GSB", "focus": "Economics of digitization, platform economics, causal inference",
        "recent": 'AER Presidential Address (April 2025), "ML Who to Nudge" (J Econometrics 2025)',
        "url": "https://profiles.stanford.edu/susan-athey",
        "email": "athey@stanford.edu", "website": "https://athey.people.stanford.edu/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 18, "name": "Tom Byers", "dept": "MS&E",
        "area": "Engineering", "focus": "Principled entrepreneurship, responsible tech innovation",
        "recent": "2024 GCEC Legacy Award, Mayfield Fellows, Hacking for Defense",
        "url": "https://profiles.stanford.edu/tom-byers",
        "email": "tbyers@stanford.edu", "website": "https://stvp.stanford.edu/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 19, "name": "James Zou", "dept": "Biomedical Data Science",
        "area": "Engineering", "focus": "Fair AI, ML for biomedicine",
        "recent": '"Virtual Scientists" (Nature 2025), multiple Nature-family papers 2024',
        "url": "https://profiles.stanford.edu/james-zou",
        "email": "jamesz@stanford.edu", "website": "https://www.james-zou.com/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 20, "name": "Nicholas Bloom", "dept": "Economics / SIEPR",
        "area": "Interdisciplinary", "focus": "Remote work, management practices, innovation",
        "recent": '"Hybrid WFH improves retention" (Nature 2024, landmark RCT)',
        "url": "https://profiles.stanford.edu/nicholas-bloom",
        "email": "nbloom@stanford.edu", "website": "https://nbloom.people.stanford.edu/",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 21, "name": "Robert Burgelman", "dept": "Strategic Management",
        "area": "GSB", "focus": "Strategy dynamics, corporate venturing",
        "recent": '"Fading Corporate Survival Prospects" (SMJ 2024), co-selection bias framework',
        "url": "https://profiles.stanford.edu/robert-burgelman",
        "email": "profrab@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/robert-burgelman",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 22, "name": "Jesper Sorensen", "dept": "Organizational Behavior",
        "area": "GSB", "focus": "Entrepreneurship, organizational culture",
        "recent": "Stanford Seed director, career frustration to entrepreneurship research",
        "url": "https://profiles.stanford.edu/jesper-sorensen",
        "email": "sorensen@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/jesper-sorensen",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 23, "name": "William Barnett", "dept": "Strategic Management",
        "area": "GSB", "focus": "Competitive strategy, Red Queen theory",
        "recent": '"The Red Queen Among Organizations" new edition (June 2025)',
        "url": "https://profiles.stanford.edu/william-barnett",
        "email": "william.barnett@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/william-barnett",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 24, "name": "Ilya Strebulaev", "dept": "Finance",
        "area": "GSB", "focus": "Venture capital, entrepreneurial finance",
        "recent": '"The Venture Mindset" bestseller (2024), VC gender/race bias (Mgmt Science 2024)',
        "url": "https://profiles.stanford.edu/ilya-strebulaev",
        "email": "istrebulaev@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/ilya-strebulaev",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 25, "name": "Peter DeMarzo", "dept": "Finance",
        "area": "GSB", "focus": "Corporate finance, financial contracting",
        "recent": "SVB collapse analysis (2025), was interim GSB Dean 2024-25",
        "url": "https://profiles.stanford.edu/peter-demarzo",
        "email": "pdemarzo@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/peter-demarzo",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 26, "name": "Jeffrey Pfeffer", "dept": "Organizational Behavior",
        "area": "GSB", "focus": "Power in organizations, evidence-based management",
        "recent": '"Pfeffer on Power" podcast, nurse burnout paper (JAMA 2025)',
        "url": "https://profiles.stanford.edu/jeffrey-pfeffer",
        "email": "pfeff@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/jeffrey-pfeffer",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 27, "name": "Hayagreeva Rao", "dept": "Organizational Behavior",
        "area": "GSB", "focus": "Institutional change, social movements, scaling",
        "recent": '"The Friction Project" (2024 bestseller with Sutton)',
        "url": "https://profiles.stanford.edu/hayagreeva-rao",
        "email": "hrao@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/hayagreeva-rao",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 28, "name": "Lindred Greer", "dept": "Organizational Behavior",
        "area": "GSB", "focus": "Team dynamics, power/hierarchy",
        "recent": '"Hierarchical flexing" research',
        "url": "https://profiles.stanford.edu/lindred-greer",
        "email": "lgreer@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/lindred-greer",
        "status": "", "notes": "May have moved to Michigan Ross", "flag": "May have moved to Michigan Ross",
    },
    {
        "num": 29, "name": "Robert Sutton", "dept": "MS&E / Org Behavior",
        "area": "Interdisciplinary", "focus": "Innovation, organizational change, scaling",
        "recent": '"The Friction Project" (2024)',
        "url": "https://profiles.stanford.edu/robert-sutton",
        "email": "bobsut@stanford.edu", "website": "https://www.bobsutton.net/",
        "status": "", "notes": "Professor Emeritus (retired 2023)", "flag": "Emeritus (retired 2023)",
    },
    {
        "num": 30, "name": "Haim Mendelson", "dept": "Operations/IT",
        "area": "GSB", "focus": "IT strategy, digital transformation, platforms",
        "recent": "AI-Powered Innovation speaker series (2025), Organizational IQ framework",
        "url": "https://profiles.stanford.edu/haim-mendelson",
        "email": "haim@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/haim-mendelson",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 31, "name": "Hau Lee", "dept": "Operations/IT",
        "area": "GSB", "focus": "Supply chain management, value chain innovation",
        "recent": '"AI-driven Supply Chain" (2025), Stanford-Amazon Supply Chain Summit (2025)',
        "url": "https://profiles.stanford.edu/hau-lee",
        "email": "haulee@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/hau-lee",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 32, "name": "Stefanos Zenios", "dept": "Operations/IT",
        "area": "GSB", "focus": "Healthcare innovation, biodesign, entrepreneurship",
        "recent": "Startup Garage (DoorDash origin, $3B+ VC raised), Ecopreneurship program",
        "url": "https://profiles.stanford.edu/stefanos-zenios",
        "email": "stefzen@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/stefanos-zenios",
        "status": "", "notes": "", "flag": None,
    },
    {
        "num": 33, "name": "Sarah Soule", "dept": "Org Behavior / Strategy",
        "area": "GSB", "focus": "Social movements, CSR, institutional change",
        "recent": "Appointed GSB Dean June 2025 (first woman)",
        "url": "https://profiles.stanford.edu/sarah-soule",
        "email": "soule@stanford.edu", "website": "https://www.gsb.stanford.edu/faculty-research/faculty/sarah-soule",
        "status": "", "notes": "Now GSB Dean (administrator)", "flag": "Now GSB Dean (administrator)",
    },
]

# ── Column Configuration ──────────────────────────────────────────────────────
COLUMNS = [
    {"header": "#",                  "key": "num",     "width": 5},
    {"header": "Name",               "key": "name",    "width": 24},
    {"header": "Department",         "key": "dept",    "width": 24},
    {"header": "Area",               "key": "area",    "width": 20},
    {"header": "Research Focus",     "key": "focus",   "width": 42},
    {"header": "Key Recent Work (2024-2025)", "key": "recent", "width": 56},
    {"header": "Profile URL",        "key": "url",     "width": 38},
    {"header": "Email",              "key": "email",   "width": 28},
    {"header": "Personal/Lab Website", "key": "website", "width": 38},
    {"header": "Status",             "key": "status",  "width": 16},
    {"header": "Priority",           "key": "priority","width": 14},
    {"header": "Notes",              "key": "notes",   "width": 40},
]

NUM_COLS = len(COLUMNS)


def determine_priority(prof):
    """Assign priority based on availability and relevance signals."""
    if prof["flag"]:
        return "Unavailable"
    # High priority: top-tier recent publications or awards, strong AI/entrepreneurship focus
    high_keywords = [
        "Nature", "QJE", "bestseller", "Time Person", "ACM Fellow",
        "Lanchester", "AER Presidential", "Queen Elizabeth",
    ]
    if any(kw.lower() in prof["recent"].lower() for kw in high_keywords):
        return "High"
    # Medium priority: solid publications
    med_keywords = ["2025", "ICML", "Org Science", "AMJ", "JAMA"]
    if any(kw.lower() in prof["recent"].lower() for kw in med_keywords):
        return "Medium"
    return "Standard"


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Outreach Tracker"

    # ── Sheet settings ─────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = CARDINAL_RED

    # ── Row 1: Title ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Stanford Professors - Visiting Research Position Outreach"
    title_cell.font = Font(name="Calibri", size=20, bold=True, color=STANFORD_WHITE)
    title_cell.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    # Fill remaining merged cells with the same background
    for c in range(2, NUM_COLS + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")

    # ── Row 2: Subtitle ────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = "Engineering, Business, Management & Entrepreneurship  |  April 2026"
    sub_cell.font = Font(name="Calibri", size=12, italic=True, color="C0C8D8")
    sub_cell.fill = PatternFill(start_color=SUBTITLE_BG, end_color=SUBTITLE_BG, fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    for c in range(2, NUM_COLS + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = PatternFill(start_color=SUBTITLE_BG, end_color=SUBTITLE_BG, fill_type="solid")

    # ── Row 3: Spacer ──────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6
    for c in range(1, NUM_COLS + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = PatternFill(start_color=STANFORD_WHITE, end_color=STANFORD_WHITE, fill_type="solid")
        cell.border = Border(bottom=Side(style="medium", color=CARDINAL_RED))

    # ── Row 4: Column Headers ──────────────────────────────────────────────
    HEADER_ROW = 4
    header_font = Font(name="Calibri", size=11, bold=True, color=STANFORD_WHITE)
    header_fill = PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, col_def in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx)
        cell.value = col_def["header"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
    ws.row_dimensions[HEADER_ROW].height = 36

    # ── Column Widths ──────────────────────────────────────────────────────
    for idx, col_def in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_def["width"]

    # ── Area fill mapping ──────────────────────────────────────────────────
    area_fills = {
        "Engineering":     PatternFill(start_color=AREA_BLUE, end_color=AREA_BLUE, fill_type="solid"),
        "Engineering + GSB": PatternFill(start_color="4A2272", end_color="4A2272", fill_type="solid"),
        "GSB":             PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid"),
        "Interdisciplinary": PatternFill(start_color=AREA_PURPLE, end_color=AREA_PURPLE, fill_type="solid"),
    }
    area_fonts = {
        "Engineering":       Font(name="Calibri", size=10, bold=True, color=STANFORD_WHITE),
        "Engineering + GSB": Font(name="Calibri", size=10, bold=True, color=STANFORD_WHITE),
        "GSB":               Font(name="Calibri", size=10, bold=True, color=STANFORD_WHITE),
        "Interdisciplinary": Font(name="Calibri", size=10, bold=True, color=STANFORD_WHITE),
    }

    unavail_fill = PatternFill(start_color=UNAVAIL_FILL, end_color=UNAVAIL_FILL, fill_type="solid")
    white_fill   = PatternFill(start_color=STANFORD_WHITE, end_color=STANFORD_WHITE, fill_type="solid")
    alt_fill     = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    # Priority fills
    priority_styles = {
        "High":        (PatternFill(start_color=PRIORITY_HIGH, end_color=PRIORITY_HIGH, fill_type="solid"),
                        Font(name="Calibri", size=10, bold=True, color="375623")),
        "Medium":      (PatternFill(start_color=PRIORITY_MED, end_color=PRIORITY_MED, fill_type="solid"),
                        Font(name="Calibri", size=10, bold=True, color="843C0C")),
        "Standard":    (PatternFill(start_color=PRIORITY_LOW, end_color=PRIORITY_LOW, fill_type="solid"),
                        Font(name="Calibri", size=10, color="44546A")),
        "Unavailable": (PatternFill(start_color="F2DBDB", end_color="F2DBDB", fill_type="solid"),
                        Font(name="Calibri", size=10, italic=True, color="953734")),
    }

    # ── Data Rows ──────────────────────────────────────────────────────────
    DATA_START = 5
    default_font  = Font(name="Calibri", size=10, color=DARK_GRAY)
    name_font     = Font(name="Calibri", size=11, bold=True, color="1B2A4A")
    link_font     = Font(name="Calibri", size=10, color="0563C1", underline="single")
    email_font    = Font(name="Calibri", size=10, color="0563C1", underline="single")
    unavail_name_font = Font(name="Calibri", size=11, bold=True, color="953734", italic=True)
    unavail_text_font = Font(name="Calibri", size=10, color="953734", italic=True)

    for i, prof in enumerate(PROFESSORS):
        row = DATA_START + i
        is_unavail = prof["flag"] is not None
        is_alt = (i % 2 == 1)
        priority = determine_priority(prof)

        # Determine row background
        if is_unavail:
            row_fill = unavail_fill
        elif is_alt:
            row_fill = alt_fill
        else:
            row_fill = white_fill

        # Populate each column
        for col_idx, col_def in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=col_idx)
            key = col_def["key"]

            if key == "priority":
                cell.value = priority
            else:
                cell.value = prof.get(key, "")

            # Default styling
            cell.font = unavail_text_font if is_unavail else default_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center", wrap_text=(key in ("focus", "recent", "notes")),
                horizontal="center" if key in ("num", "status", "priority") else "left",
            )

            # Special column styling
            if key == "num":
                cell.font = Font(name="Calibri", size=10, bold=True, color=CARDINAL_RED)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif key == "name":
                cell.font = unavail_name_font if is_unavail else name_font

            elif key == "area":
                area_val = prof["area"]
                if area_val in area_fills:
                    cell.fill = area_fills[area_val]
                    cell.font = area_fonts[area_val]
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            elif key == "url" and prof["url"]:
                cell.font = link_font if not is_unavail else unavail_text_font
                cell.hyperlink = prof["url"]

            elif key == "email" and prof["email"]:
                cell.font = email_font if not is_unavail else unavail_text_font
                cell.hyperlink = f"mailto:{prof['email']}"

            elif key == "website" and prof["website"]:
                cell.font = link_font if not is_unavail else unavail_text_font
                cell.hyperlink = prof["website"]

            elif key == "priority":
                pfill, pfont = priority_styles.get(priority, (row_fill, default_font))
                cell.fill = pfill
                cell.font = pfont
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 32

    LAST_DATA_ROW = DATA_START + len(PROFESSORS) - 1

    # ── Status Data Validation (dropdown) ──────────────────────────────────
    status_col_letter = get_column_letter(COLUMNS.index(
        next(c for c in COLUMNS if c["key"] == "status")
    ) + 1)
    dv = DataValidation(
        type="list",
        formula1='"Contacted,Pending,No Response,Scheduled,Declined"',
        allow_blank=True,
    )
    dv.error = "Please select a valid status."
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select outreach status"
    dv.promptTitle = "Status"
    ws.add_data_validation(dv)
    dv.add(f"{status_col_letter}{DATA_START}:{status_col_letter}{LAST_DATA_ROW}")

    # ── Conditional Formatting for Status ──────────────────────────────────
    status_range = f"{status_col_letter}{DATA_START}:{status_col_letter}{LAST_DATA_ROW}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal", formula=['"Contacted"'],
            fill=PatternFill(start_color=GREEN_STATUS, end_color=GREEN_STATUS, fill_type="solid"),
            font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal", formula=['"Pending"'],
            fill=PatternFill(start_color=YELLOW_STATUS, end_color=YELLOW_STATUS, fill_type="solid"),
            font=Font(color=YELLOW_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal", formula=['"No Response"'],
            fill=PatternFill(start_color=RED_STATUS, end_color=RED_STATUS, fill_type="solid"),
            font=Font(color=RED_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal", formula=['"Scheduled"'],
            fill=PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            font=Font(color="1F3864", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal", formula=['"Declined"'],
            fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            font=Font(color="4A4A4A", bold=True, italic=True),
        ),
    )

    # ── Auto-filter ────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(NUM_COLS)}{LAST_DATA_ROW}"

    # ── Freeze Panes (below header row) ────────────────────────────────────
    ws.freeze_panes = f"A{DATA_START}"

    # ════════════════════════════════════════════════════════════════════════
    # ── Summary Statistics Sheet ───────────────────────────────────────────
    # ════════════════════════════════════════════════════════════════════════
    ss = wb.create_sheet(title="Summary Statistics")
    ss.sheet_properties.tabColor = DARK_NAVY

    # Count by area
    area_counts = {}
    dept_counts = {}
    priority_counts = {"High": 0, "Medium": 0, "Standard": 0, "Unavailable": 0}
    unavail_list = []

    for prof in PROFESSORS:
        area = prof["area"]
        dept = prof["dept"]
        pri = determine_priority(prof)
        area_counts[area] = area_counts.get(area, 0) + 1
        dept_counts[dept] = dept_counts.get(dept, 0) + 1
        priority_counts[pri] = priority_counts.get(pri, 0) + 1
        if prof["flag"]:
            unavail_list.append((prof["name"], prof["flag"]))

    # ── Title ──────────────────────────────────────────────────────────────
    ss.merge_cells("A1:F1")
    c = ss.cell(row=1, column=1)
    c.value = "Outreach Summary Dashboard"
    c.font = Font(name="Calibri", size=18, bold=True, color=STANFORD_WHITE)
    c.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ss.row_dimensions[1].height = 44
    for col in range(2, 7):
        ss.cell(row=1, column=col).fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")

    ss.merge_cells("A2:F2")
    c = ss.cell(row=2, column=1)
    c.value = "Stanford Visiting Research Position  |  April 2026"
    c.font = Font(name="Calibri", size=11, italic=True, color="C0C8D8")
    c.fill = PatternFill(start_color=SUBTITLE_BG, end_color=SUBTITLE_BG, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ss.row_dimensions[2].height = 28
    for col in range(2, 7):
        ss.cell(row=2, column=col).fill = PatternFill(start_color=SUBTITLE_BG, end_color=SUBTITLE_BG, fill_type="solid")

    # Helper to write a section header
    def section_header(sheet, row, text, span=2):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = sheet.cell(row=row, column=1)
        c.value = text
        c.font = Font(name="Calibri", size=13, bold=True, color=STANFORD_WHITE)
        c.fill = PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(2, span + 1):
            sheet.cell(row=row, column=col).fill = PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid")
        sheet.row_dimensions[row].height = 30

    def stat_row(sheet, row, label, value, is_alt=False):
        bg = LIGHT_GRAY if is_alt else STANFORD_WHITE
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c1 = sheet.cell(row=row, column=1)
        c1.value = label
        c1.font = Font(name="Calibri", size=11, color=DARK_GRAY)
        c1.fill = fill
        c1.border = thin_border
        c1.alignment = Alignment(vertical="center", indent=1)

        c2 = sheet.cell(row=row, column=2)
        c2.value = value
        c2.font = Font(name="Calibri", size=11, bold=True, color=DARK_NAVY)
        c2.fill = fill
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row].height = 26

    # ── Overall ────────────────────────────────────────────────────────────
    r = 4
    section_header(ss, r, "OVERVIEW")
    r += 1
    stat_row(ss, r, "Total Professors", len(PROFESSORS)); r += 1
    stat_row(ss, r, "Available for Outreach", len(PROFESSORS) - len(unavail_list), True); r += 1
    stat_row(ss, r, "Unavailable / Flagged", len(unavail_list)); r += 1

    r += 1  # spacer
    section_header(ss, r, "BY AREA")
    r += 1
    for idx, (area, cnt) in enumerate(sorted(area_counts.items(), key=lambda x: -x[1])):
        stat_row(ss, r, area, cnt, idx % 2 == 1)
        r += 1

    r += 1
    section_header(ss, r, "BY PRIORITY")
    r += 1
    for idx, pri in enumerate(["High", "Medium", "Standard", "Unavailable"]):
        stat_row(ss, r, pri, priority_counts[pri], idx % 2 == 1)
        # Color the priority label
        pfill, pfont = priority_styles.get(pri, (None, None))
        if pfill:
            ss.cell(row=r, column=2).fill = pfill
            ss.cell(row=r, column=2).font = pfont
        r += 1

    r += 1
    section_header(ss, r, "FLAGGED PROFESSORS", span=3)
    r += 1
    # Sub-headers
    for ci, hdr in enumerate(["Name", "Flag/Reason", ""], start=1):
        c = ss.cell(row=r, column=ci)
        c.value = hdr
        c.font = Font(name="Calibri", size=10, bold=True, color=DARK_NAVY)
        c.fill = PatternFill(start_color=MEDIUM_GRAY, end_color=MEDIUM_GRAY, fill_type="solid")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    for idx, (name, reason) in enumerate(unavail_list):
        bg = LIGHT_GRAY if idx % 2 == 1 else STANFORD_WHITE
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c1 = ss.cell(row=r, column=1)
        c1.value = name
        c1.font = Font(name="Calibri", size=10, bold=True, color="953734")
        c1.fill = fill
        c1.border = thin_border
        c1.alignment = Alignment(vertical="center", indent=1)

        c2 = ss.cell(row=r, column=2)
        c2.value = reason
        c2.font = Font(name="Calibri", size=10, italic=True, color="953734")
        c2.fill = fill
        c2.border = thin_border
        c2.alignment = Alignment(vertical="center", indent=1)

        ss.cell(row=r, column=3).fill = fill
        ss.cell(row=r, column=3).border = thin_border
        ss.row_dimensions[r].height = 24
        r += 1

    # ── Top departments ────────────────────────────────────────────────────
    r += 1
    section_header(ss, r, "BY DEPARTMENT", span=2)
    r += 1
    for ci, hdr in enumerate(["Department", "Count"], start=1):
        c = ss.cell(row=r, column=ci)
        c.value = hdr
        c.font = Font(name="Calibri", size=10, bold=True, color=DARK_NAVY)
        c.fill = PatternFill(start_color=MEDIUM_GRAY, end_color=MEDIUM_GRAY, fill_type="solid")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1
    for idx, (dept, cnt) in enumerate(sorted(dept_counts.items(), key=lambda x: -x[1])):
        stat_row(ss, r, dept, cnt, idx % 2 == 1)
        r += 1

    # Column widths for summary
    ss.column_dimensions["A"].width = 32
    ss.column_dimensions["B"].width = 18
    ss.column_dimensions["C"].width = 18
    ss.column_dimensions["D"].width = 14
    ss.column_dimensions["E"].width = 14
    ss.column_dimensions["F"].width = 14

    # ════════════════════════════════════════════════════════════════════════
    # ── Legend / Instructions Sheet ────────────────────────────────────────
    # ════════════════════════════════════════════════════════════════════════
    ls = wb.create_sheet(title="Legend & Instructions")
    ls.sheet_properties.tabColor = "4A2272"

    ls.merge_cells("A1:D1")
    c = ls.cell(row=1, column=1)
    c.value = "How to Use This Workbook"
    c.font = Font(name="Calibri", size=16, bold=True, color=STANFORD_WHITE)
    c.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ls.row_dimensions[1].height = 40
    for col in range(2, 5):
        ls.cell(row=1, column=col).fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")

    instructions = [
        ("Status Tracking", "Use the dropdown in the Status column to track outreach progress. Colors update automatically."),
        ("Priority Levels", "High = top-tier publications/awards. Medium = strong recent output. Standard = solid candidates. Unavailable = flagged professors."),
        ("Color Coding - Areas", "Blue = Engineering. Cardinal Red = GSB. Purple = Interdisciplinary. Deep purple = Engineering + GSB cross-appointment."),
        ("Flagged Rows", "Salmon/orange rows indicate professors who may be unavailable (on leave, left Stanford, emeritus, or in admin roles)."),
        ("Filtering", "Use the auto-filter arrows in the header row to filter by department, area, priority, or status."),
        ("Notes Column", "Use the Notes column for personal observations, meeting dates, and follow-up reminders."),
        ("Hyperlinks", "Profile URLs, emails, and websites are clickable hyperlinks. Click to open directly."),
    ]

    r = 3
    for idx, (title, desc) in enumerate(instructions):
        bg = LIGHT_GRAY if idx % 2 == 1 else STANFORD_WHITE
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        c1 = ls.cell(row=r, column=1)
        c1.value = title
        c1.font = Font(name="Calibri", size=11, bold=True, color=CARDINAL_RED)
        c1.fill = fill
        c1.border = thin_border
        c1.alignment = Alignment(vertical="top", indent=1)

        ls.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c2 = ls.cell(row=r, column=2)
        c2.value = desc
        c2.font = Font(name="Calibri", size=10, color=DARK_GRAY)
        c2.fill = fill
        c2.border = thin_border
        c2.alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(3, 5):
            ls.cell(row=r, column=col).fill = fill
            ls.cell(row=r, column=col).border = thin_border
        ls.row_dimensions[r].height = 36
        r += 1

    # Status legend
    r += 1
    ls.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ls.cell(row=r, column=1)
    c.value = "STATUS COLOR LEGEND"
    c.font = Font(name="Calibri", size=12, bold=True, color=STANFORD_WHITE)
    c.fill = PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 5):
        ls.cell(row=r, column=col).fill = PatternFill(start_color=CARDINAL_RED, end_color=CARDINAL_RED, fill_type="solid")
    ls.row_dimensions[r].height = 30
    r += 1

    status_legend = [
        ("Contacted", GREEN_STATUS, GREEN_FONT),
        ("Pending", YELLOW_STATUS, YELLOW_FONT),
        ("No Response", RED_STATUS, RED_FONT),
        ("Scheduled", "D9E2F3", "1F3864"),
        ("Declined", "E2EFDA", "4A4A4A"),
    ]
    for s_label, s_bg, s_fg in status_legend:
        c1 = ls.cell(row=r, column=1)
        c1.value = s_label
        c1.font = Font(name="Calibri", size=11, bold=True, color=s_fg)
        c1.fill = PatternFill(start_color=s_bg, end_color=s_bg, fill_type="solid")
        c1.border = thin_border
        c1.alignment = Alignment(horizontal="center", vertical="center")
        ls.row_dimensions[r].height = 26
        r += 1

    ls.column_dimensions["A"].width = 24
    ls.column_dimensions["B"].width = 28
    ls.column_dimensions["C"].width = 20
    ls.column_dimensions["D"].width = 20

    # ── Print settings for main sheet ──────────────────────────────────────
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"1:{HEADER_ROW}"

    # ── Save ───────────────────────────────────────────────────────────────
    output_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "stanford_professors_outreach.xlsx"
    )
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    print(f"  - {len(PROFESSORS)} professors across {len(area_counts)} areas")
    print(f"  - {len(unavail_list)} flagged as unavailable")
    print(f"  - Sheets: 'Outreach Tracker', 'Summary Statistics', 'Legend & Instructions'")


if __name__ == "__main__":
    build_workbook()
